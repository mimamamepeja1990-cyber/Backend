from fastapi import status
from fastapi import (
    FastAPI, Depends, HTTPException, UploadFile, File,
    WebSocket, WebSocketDisconnect, Request, BackgroundTasks, Body
)
from fastapi.responses import RedirectResponse, PlainTextResponse, Response, FileResponse, StreamingResponse
from fastapi.responses import JSONResponse
from fastapi.encoders import jsonable_encoder
from fastapi.middleware.cors import CORSMiddleware
from fastapi.exceptions import RequestValidationError
from fastapi.staticfiles import StaticFiles
from starlette.middleware.gzip import GZipMiddleware
from contextlib import asynccontextmanager

from typing import List, Optional, Dict, Any
import os
import asyncio
import logging
import csv
import json
import re
import time
import datetime
import math
import unicodedata
from io import StringIO, BytesIO
from html import escape as html_escape
from urllib.parse import quote_plus
from zoneinfo import ZoneInfo
import anyio
import sys
import traceback
import threading

from sqlalchemy.orm import Session
from sqlalchemy import text
from sqlalchemy import func
from sqlalchemy import inspect
from sqlalchemy import or_
from types import SimpleNamespace

from app import models, schemas, crud, utils
from app.database import engine, Base, get_db, SessionLocal
from sqlalchemy.exc import IntegrityError, OperationalError, DBAPIError
from fastapi.security import OAuth2PasswordRequestForm, OAuth2PasswordBearer
from typing import Tuple

# optional remote backup (GitHub Gist) ? configured via env vars
import httpx
try:
    import mercadopago  # type: ignore
except Exception:
    mercadopago = None
GIST_TOKEN = os.environ.get('BACKUP_GIST_TOKEN')
GIST_ID = os.environ.get('BACKUP_GIST_ID')
BACKUP_URL = os.environ.get('CATALOG_BACKUP_URL')  # optional public URL to fetch a snapshot from

async def push_snapshot_to_gist(content: str) -> bool:
    """Update the configured Gist with the provided JSON content. Requires BACKUP_GIST_TOKEN and BACKUP_GIST_ID."""
    if not GIST_TOKEN or not GIST_ID:
        return False
    url = f"https://api.github.com/gists/{GIST_ID}"
    headers = {"Authorization": f"token {GIST_TOKEN}", "Accept": "application/vnd.github.v3+json"}
    payload = {"files": {"products.json": {"content": content}}}
    try:
        resp = httpx.patch(url, json=payload, headers=headers, timeout=15)
        resp.raise_for_status()
        return True
    except Exception as e:
        logger.warning('push_snapshot_to_gist failed: %s', e)
        return False

def fetch_snapshot_from_gist() -> Optional[str]:
    """Fetch the products.json content from the configured Gist (if available)."""
    if GIST_ID:
        url = f"https://api.github.com/gists/{GIST_ID}"
        try:
            resp = httpx.get(url, timeout=15)
            resp.raise_for_status()
            data = resp.json()
            files = data.get('files') or {}
            pj = files.get('products.json') or files.get('products')
            if pj and pj.get('content'):
                return pj.get('content')
        except Exception as e:
            logger.warning('fetch_snapshot_from_gist failed: %s', e)
    # try a generic public URL if provided
    if BACKUP_URL:
        try:
            resp = httpx.get(BACKUP_URL, timeout=15)
            resp.raise_for_status()
            return resp.text
        except Exception as e:
            logger.warning('fetch_snapshot_from_url failed: %s', e)
    return None


def fetch_promotions_from_gist() -> Optional[str]:
    """Fetch the promotions.json content from the configured Gist (if available)."""
    if GIST_ID:
        url = f"https://api.github.com/gists/{GIST_ID}"
        try:
            resp = httpx.get(url, timeout=15)
            resp.raise_for_status()
            data = resp.json()
            files = data.get('files') or {}
            pj = files.get('promotions.json') or files.get('promotions')
            if pj and pj.get('content'):
                return pj.get('content')
        except Exception as e:
            logger.warning('fetch_promotions_from_gist failed: %s', e)
    # try a generic public URL if provided (reuse BACKUP_URL if it points to a promotions snapshot)
    if BACKUP_URL:
        try:
            resp = httpx.get(BACKUP_URL, timeout=15)
            resp.raise_for_status()
            return resp.text
        except Exception as e:
            logger.warning('fetch_promotions_from_url failed: %s', e)
    return None


# LOGGING
# -------------------------------------------------------------------
logger = logging.getLogger("catalog_api")
handler = logging.StreamHandler()
handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
logger.addHandler(handler)
logger.setLevel(logging.INFO)

# -------------------------------------------------------------------
# Engine-level safe helpers ðŸ”§
# -------------------------------------------------------------------

def _invalidate_conn(conn):
    try:
        # mark connection as invalid so it's removed from pool
        conn.invalidate()
    except Exception:
        try:
            conn.close()
        except Exception:
            pass


def _safe_engine_fetchall(sql, params=None):
    params = params or {}
    try:
        with engine.connect() as conn:
            return conn.execute(text(sql), params).fetchall()
    except Exception as e:
        msg = str(e)
        logger.exception('safe_engine_fetchall initial failed: %s', e)
        if 'current transaction is aborted' in msg.lower():
            try:
                tb = traceback.format_exc()
                # Persist diagnostic so we can find the originating failure in Render logs
                base = os.path.dirname(os.path.dirname(__file__))
                logpath = os.path.join(base, 'server_log.txt')
                with open(logpath, 'a', encoding='utf-8') as f:
                    f.write(f"{datetime.datetime.utcnow().isoformat()} - safe_engine_fetchall aborted transaction: stmt={sql} msg={msg[:300]}\n")
                    f.write(tb + "\n\n")
            except Exception:
                pass
        # retry once with a fresh connection
        try:
            with engine.connect() as conn2:
                return conn2.execute(text(sql), params).fetchall()
        except Exception as e2:
            logger.exception('safe_engine_fetchall retry failed: %s', e2)
            return None


def _safe_engine_fetchone(sql, params=None):
    params = params or {}
    try:
        with engine.connect() as conn:
            return conn.execute(text(sql), params).fetchone()
    except Exception as e:
        msg = str(e)
        logger.exception('safe_engine_fetchone initial failed: %s', e)
        if 'current transaction is aborted' in msg.lower():
            try:
                tb = traceback.format_exc()
                base = os.path.dirname(os.path.dirname(__file__))
                logpath = os.path.join(base, 'server_log.txt')
                with open(logpath, 'a', encoding='utf-8') as f:
                    f.write(f"{datetime.datetime.utcnow().isoformat()} - safe_engine_fetchone aborted transaction: stmt={sql} msg={msg[:300]}\n")
                    f.write(tb + "\n\n")
            except Exception:
                pass
        try:
            with engine.connect() as conn2:
                return conn2.execute(text(sql), params).fetchone()
        except Exception as e2:
            logger.exception('safe_engine_fetchone retry failed: %s', e2)
            return None

def _is_truthy(value: Optional[str]) -> bool:
    try:
        return str(value or '').strip().lower() in {'1', 'true', 'yes', 'y', 'on'}
    except Exception:
        return False


# --- Geocoding (optional, for delivery route optimization) ---
GEOCODE_PROVIDER = str(os.environ.get('GEOCODE_PROVIDER') or '').strip().lower() or 'nominatim'
GEOCODE_GOOGLE_API_KEY = str(os.environ.get('GEOCODE_GOOGLE_API_KEY') or '').strip() or None
GEOCODE_ENABLED = _is_truthy(os.environ.get('GEOCODE_ENABLED')) or bool(GEOCODE_GOOGLE_API_KEY)
GEOCODE_NOMINATIM_URL = str(os.environ.get('GEOCODE_NOMINATIM_URL') or 'https://nominatim.openstreetmap.org/search').strip()
GEOCODE_NOMINATIM_EMAIL = str(os.environ.get('GEOCODE_NOMINATIM_EMAIL') or '').strip() or None
GEOCODE_USER_AGENT = str(os.environ.get('GEOCODE_USER_AGENT') or 'catalog-admin').strip()
GEOCODE_RATE_LIMIT_MS = int(float(os.environ.get('GEOCODE_RATE_LIMIT_MS') or 1100))
GEOCODE_TIMEOUT_SEC = float(os.environ.get('GEOCODE_TIMEOUT_SEC') or 6)
GEOCODE_CACHE_TTL_SEC = int(float(os.environ.get('GEOCODE_CACHE_TTL_SEC') or (60 * 60 * 24 * 30)))
GEOCODE_CACHE_MAX = int(float(os.environ.get('GEOCODE_CACHE_MAX') or 2500))
GEOCODE_MAX_PER_RUN = int(float(os.environ.get('GEOCODE_MAX_PER_RUN') or 60))
GEOCODE_RUN_WINDOW_SEC = int(float(os.environ.get('GEOCODE_RUN_WINDOW_SEC') or 60))
GEOCODE_PERSIST_CACHE = _is_truthy(os.environ.get('GEOCODE_PERSIST_CACHE'))
ROUTE_START_LAT_RAW = os.environ.get('ROUTE_START_LAT') or '-32.819094'
ROUTE_START_LON_RAW = os.environ.get('ROUTE_START_LON') or '-68.804286'
ROUTE_START_ADDRESS = str(os.environ.get('ROUTE_START_ADDRESS') or '').strip()
GOOGLE_MAPS_JS_API_KEY = str(os.environ.get('GOOGLE_MAPS_JS_API_KEY') or os.environ.get('GOOGLE_MAPS_API_KEY') or '').strip()
DRIVER_LOCATION_MAX_AGE_SEC = int(float(os.environ.get('DRIVER_LOCATION_MAX_AGE_SEC') or 60 * 60))
DRIVER_LOCATION_STORE_HISTORY = _is_truthy(os.environ.get('DRIVER_LOCATION_STORE_HISTORY') or os.environ.get('DRIVER_LOCATION_HISTORY'))

_GEOCODE_CACHE: Dict[str, Dict[str, Any]] = {}
_GEOCODE_CACHE_LOCK = threading.Lock()
_GEOCODE_CACHE_LOADED = False
_GEOCODE_LAST_CALL_TS = 0.0
_GEOCODE_RUN_STATS = {'window_start': 0.0, 'count': 0}
_GEOCODE_CACHE_DIRTY = False
_ROUTE_START_CACHE: Dict[str, Any] = {}
DRIVER_LOCATIONS: Dict[str, Dict[str, Any]] = {}
DRIVER_LOCATIONS_LOCK = threading.Lock()


def _running_on_render() -> bool:
    return bool(
        os.environ.get('RENDER')
        or os.environ.get('RENDER_SERVICE_ID')
        or os.environ.get('RENDER_EXTERNAL_HOSTNAME')
    )


def _startup_skip_enabled() -> bool:
    return _is_truthy(os.environ.get('SKIP_STARTUP_BOOTSTRAP') or os.environ.get('SKIP_STARTUP'))


def _startup_background_enabled() -> bool:
    if os.environ.get('STARTUP_BACKGROUND') is not None:
        return _is_truthy(os.environ.get('STARTUP_BACKGROUND'))
    return _running_on_render()


def _ensure_default_admin_owner() -> None:
    """Create a default admin owner user if none exists (DB-backed admin auth)."""
    username = str(os.environ.get('ADMIN_OWNER_USERNAME') or 'emiliano').strip()
    password = str(os.environ.get('ADMIN_OWNER_PASSWORD') or 'badbunni33').strip()
    force_reset = _is_truthy(os.environ.get('ADMIN_OWNER_RESET') or os.environ.get('ADMIN_OWNER_FORCE_RESET'))
    if not username or not password:
        return
    db = SessionLocal()
    try:
        try:
            existing_owner = db.query(models.AdminUser).filter(models.AdminUser.role == 'owner').first()
        except Exception:
            existing_owner = None
        if existing_owner:
            if force_reset:
                try:
                    changed = False
                    if username and str(existing_owner.username or '').strip().lower() != username.lower():
                        existing_owner.username = username
                        changed = True
                    if password:
                        existing_owner.hashed_password = utils.hash_password(password)
                        changed = True
                    if getattr(existing_owner, 'is_active', True) is False:
                        existing_owner.is_active = True
                        changed = True
                    if changed:
                        db.add(existing_owner)
                        db.commit()
                except Exception:
                    try:
                        db.rollback()
                    except Exception:
                        pass
            return
        existing_user = crud.get_admin_user_by_username(db, username)
        if existing_user:
            try:
                existing_user.role = 'owner'
                if force_reset and password:
                    existing_user.hashed_password = utils.hash_password(password)
                db.add(existing_user)
                db.commit()
                return
            except Exception:
                try:
                    db.rollback()
                except Exception:
                    pass
        payload = schemas.AdminUserCreate(
            username=username,
            password=password,
            role='owner',
            full_name='Owner',
        )
        hashed = utils.hash_password(password)
        crud.create_admin_user(db, payload, hashed_password=hashed, created_by='system', force_role='owner')
    finally:
        try:
            db.close()
        except Exception:
            pass


def _ensure_admin_user_columns() -> None:
    try:
        insp = inspect(engine)
        if 'admin_users' not in insp.get_table_names():
            return
        existing = {c['name'] for c in insp.get_columns('admin_users')}
    except Exception:
        existing = set()
    if 'zone' in existing:
        return
    try:
        dialect_local = getattr(engine, 'dialect', None)
        dialect_name = getattr(dialect_local, 'name', '') if dialect_local else ''
        with engine.begin() as conn:
            if 'postgres' in dialect_name:
                try:
                    conn.execute(text("ALTER TABLE admin_users ADD COLUMN IF NOT EXISTS zone VARCHAR(120)"))
                except Exception:
                    conn.execute(text("ALTER TABLE admin_users ADD COLUMN zone VARCHAR(120)"))
            else:
                conn.execute(text("ALTER TABLE admin_users ADD COLUMN zone VARCHAR(120)"))
    except Exception:
        logger.exception('ensure_admin_user_columns failed')


def _startup_bootstrap_sync() -> None:
    """Run blocking startup tasks in a background thread when desired."""
    # Startup
    Base.metadata.create_all(bind=engine)

    # Ensure admin owner user exists (so panel users persist in DB).
    try:
        _ensure_admin_user_columns()
        _ensure_default_admin_owner()
    except Exception:
        logger.exception('ensure_default_admin_owner failed')

    # Ensure legacy DBs get required columns on the `orders` table.
    # Use SQLAlchemy inspector (cross-dialect) to detect missing columns and
    # issue ALTER TABLE ADD COLUMN statements for any that are absent. This
    # avoids relying on information_schema schema assumptions.
    try:
        with engine.begin() as conn:
            needed = {
                'status': "VARCHAR(50) DEFAULT 'recibido'",
                'customer_type': "VARCHAR(50) DEFAULT 'mayorista'",
                'user_id': 'INTEGER',
                'user_full_name': 'VARCHAR(200)',
                'user_email': 'VARCHAR(320)',
                'user_barrio': 'VARCHAR(200)',
                'user_calle': 'VARCHAR(200)',
                'user_numeracion': 'VARCHAR(100)',
                'user_postal_code': 'VARCHAR(20)',
                'user_department': 'VARCHAR(120)',
                '_token_received': 'BOOLEAN',
                '_token_preview': 'TEXT',
                # Ensure 'source' has a server-side default so rows without explicit
                # source end up with 'web' in the DB and survive restarts.
                'source': "VARCHAR(50) DEFAULT 'web'",
                'payment_method': 'VARCHAR(50)',
                'payment_status': 'VARCHAR(50)',
                'payment_reference': 'VARCHAR(200)',
                'scheduled_delivery_date': 'VARCHAR(10)',
                'delivery_cutoff_applied': 'BOOLEAN',
                'delivery_timezone': 'VARCHAR(80)',
                'delivery_cutoff_hour': 'INTEGER',
                'assigned_driver_id': 'INTEGER',
                'assigned_driver_username': 'VARCHAR(80)',
                'assigned_driver_name': 'VARCHAR(200)',
                'assigned_driver_zone': 'VARCHAR(120)',
                'assigned_at': 'TIMESTAMP',
                'delivery_lat': 'FLOAT',
                'delivery_lon': 'FLOAT',
                'route_id': 'VARCHAR(120)',
                'route_order': 'INTEGER',
                'route_generated_at': 'TIMESTAMP',
                'delivered_at': 'TIMESTAMP',
                'delivered_by_id': 'INTEGER',
                'delivered_by_username': 'VARCHAR(80)',
            }

            try:
                insp = inspect(engine)
                existing_cols = {c['name'] for c in insp.get_columns('orders')}
            except Exception:
                existing_cols = set()

            dialect = getattr(engine, 'dialect', None)
            dialect_name = getattr(dialect, 'name', '') if dialect else ''

            for col, coltype in needed.items():
                if col in existing_cols:
                    continue
                try:
                    # Prefer ALTER TABLE ... ADD COLUMN IF NOT EXISTS on Postgres
                    if 'postgres' in dialect_name:
                        try:
                            conn.execute(text(f"ALTER TABLE orders ADD COLUMN IF NOT EXISTS {col} {coltype}"))
                        except Exception:
                            conn.execute(text(f"ALTER TABLE orders ADD COLUMN {col} {coltype}"))
                    else:
                        # SQLite and other dialects generally support simple ADD COLUMN
                        conn.execute(text(f"ALTER TABLE orders ADD COLUMN {col} {coltype}"))
                    logger.info('Added missing column to orders: %s', col)
                except Exception as e:
                    logger.warning('Could not add column %s to orders: %s', col, e)
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS idx_orders_created_at ON orders(created_at);"))
            except Exception:
                pass
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS idx_orders_status ON orders(status);"))
            except Exception:
                pass
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS idx_orders_source ON orders(source);"))
            except Exception:
                pass
    except Exception:
        logger.exception('ensure orders columns step failed')

    # Run a second, explicit migration pass for legacy deployments where
    # the orders table schema may drift between releases.
    try:
        mig = _run_add_user_columns()
        if isinstance(mig, dict):
            added = mig.get('added') or []
            failed = mig.get('failed') or []
            if added:
                logger.info('orders migration pass added columns: %s', added)
            if failed:
                logger.warning('orders migration pass failed for: %s', failed)
    except Exception:
        logger.exception('secondary orders migration pass failed')

    # Ensure legacy DBs get `stock` and `discount` columns on the `products` table.
    try:
        with engine.begin() as conn:
            dialect = getattr(engine, 'dialect', None)
            dialect_name = getattr(dialect, 'name', '') if dialect else ''
            discount_type = 'REAL DEFAULT 0' if 'postgres' in dialect_name else 'FLOAT DEFAULT 0'
            prod_needed = {
                'code': 'VARCHAR(100)',
                'brand': 'VARCHAR(200)',
                'stock': 'INTEGER DEFAULT 0',
                'min_stock': 'INTEGER DEFAULT 0',
                'stock_kg': 'REAL DEFAULT 0',
                'kg_per_unit': 'REAL DEFAULT 1',
                'discount': discount_type,
                "sale_unit": "VARCHAR(20) DEFAULT 'unit'",
                "price_retail": "REAL",
                "cost": "REAL",
            }
            try:
                insp = inspect(engine)
                existing_prod_cols = {c['name'] for c in insp.get_columns('products')}
            except Exception:
                existing_prod_cols = set()

            dialect = getattr(engine, 'dialect', None)
            dialect_name = getattr(dialect, 'name', '') if dialect else ''
            for col, coltype in prod_needed.items():
                if col in existing_prod_cols:
                    continue
                try:
                    if 'postgres' in dialect_name:
                        try:
                            conn.execute(text(f"ALTER TABLE products ADD COLUMN IF NOT EXISTS {col} {coltype}"))
                        except Exception:
                            conn.execute(text(f"ALTER TABLE products ADD COLUMN {col} {coltype}"))
                    else:
                        conn.execute(text(f"ALTER TABLE products ADD COLUMN {col} {coltype}"))
                    logger.info('Added missing column to products: %s', col)
                except Exception as e:
                    try:
                        _invalidate_conn(conn)
                    except Exception:
                        pass
                    logger.warning('Could not add column %s to products: %s', col, e)
            # Backfill for legacy kg products: if stock_kg is empty but stock has value,
            # copy stock into stock_kg so existing inventory remains available.
            try:
                if {'stock_kg', 'stock', 'sale_unit'}.issubset(existing_prod_cols | set(prod_needed.keys())):
                    conn.execute(text("""
                        UPDATE products
                        SET stock_kg = stock
                        WHERE (stock_kg IS NULL OR stock_kg <= 0)
                          AND stock IS NOT NULL
                          AND stock > 0
                          AND LOWER(COALESCE(sale_unit, 'unit')) IN ('kg','kilo','kilos','kilogram','kilograms','kilogramo','kilogramos')
                    """))
            except Exception as e:
                logger.warning('Could not backfill stock_kg from stock: %s', e)
            # Add lightweight indexes for common filters/sorts
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS idx_products_category ON products(category);"))
            except Exception:
                pass
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS idx_products_active ON products(active);"))
            except Exception:
                pass
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS idx_products_name ON products(name);"))
            except Exception:
                pass
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS idx_products_price ON products(price);"))
            except Exception:
                pass
    except Exception:
        logger.exception('ensure products columns step failed')

    # Log which database we are using (mask credentials) and test connection
    try:
        db_env = os.environ.get('DATABASE_URL')
        masked = 'sqlite (local file)'
        if db_env:
            try:
                from urllib.parse import urlparse
                u = urlparse(db_env)
                user = u.username or ''
                host = u.hostname or ''
                port = u.port or ''
                path = u.path or ''
                masked = f"{u.scheme}://{user + ':****@' if user else ''}{host}{(':'+str(port)) if port else ''}{path}"
            except Exception:
                masked = 'postgres (masked)'
        logger.info('Database URL: %s', masked)
        try:
            conn = engine.connect()
            try:
                conn.execute(text('SELECT 1'))
            finally:
                conn.close()
            logger.info('Database connection test succeeded')
        except Exception as e:
            logger.exception('Database connection test failed: %s', e)
        # If running on Render with SQLite, require a persistent disk path.
        # Otherwise data (users/addresses/orders) is lost on redeploy.
        running_on_render = bool(
            os.environ.get('RENDER')
            or os.environ.get('RENDER_SERVICE_ID')
            or os.environ.get('RENDER_EXTERNAL_HOSTNAME')
        )
        using_sqlite = (db_env is None) or (str(db_env).strip() == '') or (str(db_env).lower().startswith('sqlite'))
        if running_on_render and using_sqlite:
            engine_url = str(getattr(engine, 'url', '') or '')
            env_markers = [
                str(os.environ.get('RENDER_DISK_MOUNT_PATH') or '').strip(),
                str(os.environ.get('RENDER_DISK_PATH') or '').strip(),
                str(os.environ.get('RENDER_PERSISTENT_DISK_PATH') or '').strip(),
                str(os.environ.get('DB_PATH') or '').strip(),
                str(os.environ.get('DATA_DIR') or '').strip(),
                str(os.environ.get('PERSISTENT_DATA_DIR') or '').strip(),
            ]
            mount_markers = []
            for marker in ('/var/data',):
                try:
                    if os.path.ismount(marker):
                        mount_markers.append(marker)
                except Exception:
                    continue
            persistent_markers = [m for m in (env_markers + mount_markers) if m]
            sqlite_on_persistent_disk = any(marker in engine_url for marker in persistent_markers)
            if not sqlite_on_persistent_disk:
                logger.error(
                    'Detected Render + SQLite without persistent disk path in DB URL (%s). '
                    'Configure DATABASE_URL (Postgres) or mount persistent disk and point DB_PATH/DATA_DIR to it.',
                    engine_url,
                )
                raise RuntimeError(
                    'Render deployment requires persistent storage for the database. '
                    'Configure DATABASE_URL (Postgres) or set DB_PATH/DATA_DIR to a persistent disk mount.'
                )
            logger.warning(
                'Running on Render with SQLite backed by persistent disk path (%s). '
                'For production resilience, Postgres is still recommended.',
                engine_url,
            )
    except RuntimeError:
        raise
    except Exception:
        logger.exception('Database startup check failed')

    # Log current email delivery configuration once at startup.
    try:
        logger.info('Resend config snapshot: %s', _resend_status_snapshot())
    except Exception:
        logger.exception('Could not log Resend config snapshot')

    db = SessionLocal()
    try:
        try:
            # Use a raw count select to avoid SQLAlchemy attempting to select
            # mapped columns that may not yet exist in legacy DB schemas.
            r = crud._safe_scalar(db, 'SELECT count(*) FROM products')
            prod_count = int(r or 0)
        except Exception as e:
            # Avoid any ORM-based product queries here because the ORM may
            # reference model columns (e.g. `stock`) that are missing in the
            # live DB and cause a ProgrammingError. We log and treat as empty.
            logger.warning('Raw products count failed; skipping ORM count to avoid schema errors: %s', e)
            prod_count = 0
        if prod_count == 0:
            # try to restore from remote backup (gist or public URL) or from local snapshot before seeding demo data
            restored = False
            try:
                content = fetch_snapshot_from_gist()
                if content:
                    logger.info('Restoring products from configured backup');
                    items = json.loads(content)
                    for p in items:
                        db.add(models.Product(code=(str(p.get('code') or p.get('codigo') or '').strip() or None), name=p.get('name'), price=p.get('price') or 0, price_retail=p.get('price_retail'), description=p.get('description') or '', category=p.get('category') or '', image_url=p.get('image_url') or None, active=bool(p.get('active', True))))
                    db.commit()
                    restored = True
            except Exception as _err:
                logger.warning('restore-from-backup (gist) failed: %s', _err)

            # If no remote backup restored, try local catalog snapshot file (written by write_catalog_snapshot)
            if not restored:
                try:
                    local_path = os.path.join(CATALOG_DIR, 'products.json')
                    if os.path.exists(local_path):
                        with open(local_path, 'r', encoding='utf-8') as f:
                            items = json.load(f)
                        if items and isinstance(items, list):
                            logger.info('Restoring products from local snapshot %s', local_path)
                            for p in items:
                                db.add(models.Product(code=(str(p.get('code') or p.get('codigo') or '').strip() or None), name=p.get('name'), price=p.get('price') or 0, price_retail=p.get('price_retail'), description=p.get('description') or '', category=p.get('category') or '', image_url=p.get('image_url') or None, active=bool(p.get('active', True))))
                            db.commit()
                            restored = True
                except Exception as _err:
                    logger.warning('restore-from-local-snapshot failed: %s', _err)

            # fallback to demo data only if nothing was restored
            if not restored:
                try:
                    db.add_all([
                        models.Product(name="Camiseta", price=19.99, description="Camiseta", category="Ropa"),
                        models.Product(name="Taza", price=8.5, description="Taza", category="Accesorios"),
                    ])
                    db.commit()
                except Exception as _err:
                    logger.warning('seeding demo data failed: %s', _err)
        # Try to restore promotions snapshot from remote backup (gist) or local snapshot
        try:
            restored_promos = False
            try:
                prom_content = fetch_promotions_from_gist()
                if prom_content:
                    try:
                        # ensure catalog dir exists
                        os.makedirs(CATALOG_DIR, exist_ok=True)
                        with open(os.path.join(CATALOG_DIR, 'promotions.json'), 'w', encoding='utf-8') as f:
                            f.write(prom_content)
                        logger.info('Restored promotions from configured backup');
                        restored_promos = True
                    except Exception as e:
                        logger.warning('writing restored promotions failed: %s', e)
            except Exception as _err:
                logger.warning('restore-promotions-from-backup failed: %s', _err)

            # If no remote promotions restored, prefer existing local promotions.json if present (no-op)
        except Exception:
            logger.exception('promotions restore step failed')
    finally:
        db.close()


# -------------------------------------------------------------------
# LIFESPAN (startup / shutdown)
# -------------------------------------------------------------------
@asynccontextmanager
async def lifespan(app: FastAPI):
    if _startup_skip_enabled():
        logger.warning('Skipping startup bootstrap (SKIP_STARTUP_BOOTSTRAP enabled).')
        yield
        return

    if _startup_background_enabled():
        logger.info('Running startup bootstrap in background thread.')

        async def _run_background() -> None:
            try:
                await asyncio.to_thread(_startup_bootstrap_sync)
            except Exception:
                logger.exception('Background startup bootstrap failed')

        try:
            app.state.startup_task = asyncio.create_task(_run_background())
        except Exception:
            asyncio.create_task(_run_background())
        yield
        return

    await asyncio.to_thread(_startup_bootstrap_sync)
    yield
    # Shutdown (nada)

# -------------------------------------------------------------------
# APP
# -------------------------------------------------------------------
app = FastAPI(title="Catálogo API", lifespan=lifespan)
app.add_middleware(GZipMiddleware, minimum_size=int(os.environ.get('GZIP_MIN_SIZE') or 1000))

# In-memory cache of recently created order payloads (id -> { payload, ts })
# Used to surface token previews in the admin list when the DB lacks persisted user_* columns.
ORDER_PAYLOAD_CACHE = {}
ORDER_PAYLOAD_CACHE_MAX_AGE = 60 * 60 * 2  # keep for 2 hours

# In-memory cache for order status overrides (id -> { status, ts })
# Used as a fallback if DB status column is missing or update fails.
ORDER_STATUS_CACHE = {}
ORDER_STATUS_CACHE_MAX_AGE = 60 * 60 * 24  # keep for 24 hours

# Catalog endpoint caches (short TTL to reduce load during bursts)
PRODUCTS_CACHE: Dict[Any, Dict[str, Any]] = {}
PRODUCTS_CACHE_MAX = int(os.environ.get('PRODUCTS_CACHE_MAX') or 200)
PRODUCTS_CACHE_TTL = float(os.environ.get('PRODUCTS_CACHE_TTL') or 10.0)
PRODUCTS_CACHE_LIMIT_MAX = int(os.environ.get('PRODUCTS_CACHE_LIMIT_MAX') or 1000)
PRODUCTS_COUNT_CACHE: Dict[Any, Dict[str, Any]] = {}
PRODUCTS_COUNT_CACHE_TTL = float(os.environ.get('PRODUCTS_COUNT_CACHE_TTL') or 15.0)
PROMOTIONS_CACHE: Dict[str, Any] = {'ts': 0.0, 'mtime': None, 'value': None}
PROMOTIONS_CACHE_TTL = float(os.environ.get('PROMOTIONS_CACHE_TTL') or 15.0)

# Auto-image progress tracking (in-memory, last job wins).
AUTO_IMAGE_PROGRESS_LOCK = threading.Lock()
AUTO_IMAGE_PROGRESS: Dict[str, Any] = {
    'status': 'idle',
    'total': 0,
    'processed': 0,
    'attached': 0,
    'started_at': None,
    'finished_at': None,
    'last_update': None,
    'source': None,
    'last_error': None,
}


def _cache_get(cache: Dict[Any, Dict[str, Any]], key: Any, ttl: float):
    try:
        entry = cache.get(key)
        if not entry:
            return None
        if (time.time() - float(entry.get('ts') or 0)) > ttl:
            cache.pop(key, None)
            return None
        return entry.get('value')
    except Exception:
        return None


def _cache_set(cache: Dict[Any, Dict[str, Any]], key: Any, value: Any, max_size: int):
    try:
        cache[key] = {'value': value, 'ts': time.time()}
        if len(cache) > max_size:
            # prune oldest entries
            ordered = sorted(cache.items(), key=lambda item: item[1].get('ts', 0))
            for k, _ in ordered[: max(0, len(cache) - max_size)]:
                cache.pop(k, None)
    except Exception:
        pass


def _normalize_cache_text(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    try:
        return str(value).strip().lower() or None
    except Exception:
        return None


def _products_cache_key(skip: int, limit: int, q: Optional[str], category: Optional[str], active: Optional[bool], sort: Optional[str]):
    return (
        int(skip or 0),
        int(limit or 0),
        _normalize_cache_text(q),
        _normalize_cache_text(category),
        None if active is None else bool(active),
        _normalize_cache_text(sort),
    )


def _serialize_products(items: Any) -> List[Dict[str, Any]]:
    try:
        return jsonable_encoder(items)  # detach ORM instances
    except Exception:
        out = []
        for it in (items or []):
            if isinstance(it, dict):
                out.append(it)
                continue
            try:
                data = {k: v for k, v in vars(it).items() if not k.startswith('_')}
                out.append(data)
            except Exception:
                continue
        return out


def _invalidate_products_cache() -> None:
    PRODUCTS_CACHE.clear()
    PRODUCTS_COUNT_CACHE.clear()


def _invalidate_promotions_cache() -> None:
    PROMOTIONS_CACHE.update({'ts': 0.0, 'mtime': None, 'value': None})


def _get_promotions_cached() -> List[Dict[str, Any]]:
    path = os.path.join(CATALOG_DIR, 'promotions.json')
    try:
        if not os.path.exists(path):
            return []
        mtime = os.path.getmtime(path)
        now = time.time()
        cached = PROMOTIONS_CACHE.get('value')
        if (
            cached is not None
            and PROMOTIONS_CACHE.get('mtime') == mtime
            and (now - float(PROMOTIONS_CACHE.get('ts') or 0)) < PROMOTIONS_CACHE_TTL
        ):
            return cached
        with open(path, 'r', encoding='utf-8') as f:
            data = _normalize_promotions_payload(json.load(f))
        PROMOTIONS_CACHE.update({'value': data, 'mtime': mtime, 'ts': now})
        return data
    except Exception as e:
        logger.exception('promotions cache read failed: %s', e)
        cached = PROMOTIONS_CACHE.get('value')
        return cached if isinstance(cached, list) else []

def _prune_order_cache():
    try:
        now = time.time()
        for k in list(ORDER_PAYLOAD_CACHE.keys()):
            try:
                if ORDER_PAYLOAD_CACHE[k] and (now - ORDER_PAYLOAD_CACHE[k].get('ts', 0) > ORDER_PAYLOAD_CACHE_MAX_AGE):
                    del ORDER_PAYLOAD_CACHE[k]
            except Exception:
                pass
    except Exception:
        pass

def _prune_status_cache():
    try:
        now = time.time()
        for k in list(ORDER_STATUS_CACHE.keys()):
            try:
                if ORDER_STATUS_CACHE[k] and (now - ORDER_STATUS_CACHE[k].get('ts', 0) > ORDER_STATUS_CACHE_MAX_AGE):
                    del ORDER_STATUS_CACHE[k]
            except Exception:
                pass
    except Exception:
        pass


# --- Orders: lifecycle normalization + auto milestones ---
# Canonical lifecycle: recibido -> visto -> preparado -> enviado -> entregado
_ORDER_LIFECYCLE = ('recibido', 'visto', 'preparado', 'enviado', 'entregado')
_ORDER_STATUS_ALIASES = {
    # legacy / english / variants
    'nuevo': 'recibido',
    'new': 'recibido',
    'pendiente': 'recibido',
    'pending': 'recibido',
    'seen': 'visto',
    'viewed': 'visto',
    'preparando': 'preparado',
    'preparing': 'preparado',
    'prepared': 'preparado',
    'en_camino': 'enviado',
    'encamino': 'enviado',
    'delivering': 'enviado',
    'shipped': 'enviado',
    'delivered': 'entregado',
    'canceled': 'cancelado',
    'cancelled': 'cancelado',
}
_ORDER_STATUS_RANK = {
    'recibido': 1,
    'visto': 2,
    'preparado': 3,
    'enviado': 4,
    'entregado': 5,
}
_ORDER_TZINFO_CACHE: Dict[str, Any] = {}


def _normalize_order_status(value: Any) -> str:
    try:
        raw = str(value or '').strip().lower()
    except Exception:
        raw = ''
    if not raw:
        return 'recibido'
    # Normalize separators to keep stored values consistent.
    raw = re.sub(r'[\s\-]+', '_', raw)
    raw = _ORDER_STATUS_ALIASES.get(raw, raw)
    if raw == 'cancelado':
        return 'cancelado'
    if raw in _ORDER_LIFECYCLE:
        return raw
    return 'recibido'


def _normalize_order_status_strict(value: Any) -> Optional[str]:
    """Normalize a status but return None when the input is empty/unknown."""
    try:
        raw = str(value or '').strip().lower()
    except Exception:
        raw = ''
    if not raw:
        return None
    raw = re.sub(r'[\s\-]+', '_', raw)
    raw = _ORDER_STATUS_ALIASES.get(raw, raw)
    if raw == 'cancelado':
        return 'cancelado'
    if raw in _ORDER_LIFECYCLE:
        return raw
    return None


def _order_status_rank(value: Any) -> int:
    v = _normalize_order_status(value)
    return int(_ORDER_STATUS_RANK.get(v, 0) or 0)


def _resolve_order_tzinfo(tz_name_hint: Optional[str] = None):
    tz_name = str(
        tz_name_hint
        or os.environ.get('ORDER_STATUS_TIMEZONE')
        or os.environ.get('ORDER_EMAIL_TIMEZONE')
        or os.environ.get('ORDER_DELIVERY_TIMEZONE')
        or 'America/Argentina/Mendoza'
    ).strip() or 'America/Argentina/Mendoza'
    cached = _ORDER_TZINFO_CACHE.get(tz_name)
    if cached is not None:
        return cached
    try:
        tzinfo = ZoneInfo(tz_name)
    except Exception:
        tz_name = 'UTC'
        tzinfo = datetime.timezone.utc
    _ORDER_TZINFO_CACHE[tz_name] = tzinfo
    return tzinfo


def _coerce_datetime(value: Any) -> Optional[datetime.datetime]:
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        # Accept typical ISO strings coming from JSON encoders.
        s = s.replace('Z', '+00:00')
        try:
            return datetime.datetime.fromisoformat(s)
        except Exception:
            return None
    return None


def _compute_order_auto_status(
    created_at_value: Any,
    scheduled_delivery_date_value: Any = None,
    status_value: Any = None,
    now_utc: Optional[datetime.datetime] = None,
    tz_name_hint: Optional[str] = None,
) -> Optional[str]:
    """Auto-advance status for delivery milestones (PedidosYa-style).

    - enviado: 09:00 local time (scheduled delivery day)
    - entregado: 16:00 local time (scheduled delivery day)

    Important: The auto milestones are applied only after the order reached
    at least `preparado`. This prevents time-based jumps from blocking the
    admin workflow (recibido -> visto -> preparado).
    """
    base_status = _normalize_order_status(status_value)
    if base_status == 'cancelado':
        return None
    if _order_status_rank(base_status) < _order_status_rank('preparado'):
        return None

    base_utc = now_utc or datetime.datetime.now(datetime.timezone.utc)
    if base_utc.tzinfo is None:
        base_utc = base_utc.replace(tzinfo=datetime.timezone.utc)

    tzinfo = _resolve_order_tzinfo(tz_name_hint)
    now_local = base_utc.astimezone(tzinfo)

    target_date = None
    schedule_key = _normalize_iso_date_key(scheduled_delivery_date_value)
    if schedule_key:
        try:
            target_date = datetime.date.fromisoformat(schedule_key)
        except Exception:
            target_date = None
    if target_date is None:
        created_at = _coerce_datetime(created_at_value)
        if created_at is None:
            return None
        # Treat naive created_at as UTC for consistency across sqlite/postgres.
        if created_at.tzinfo is None:
            created_at = created_at.replace(tzinfo=datetime.timezone.utc)
        created_local = created_at.astimezone(tzinfo)
        target_date = created_local.date() + datetime.timedelta(days=1)

    enviado_at = datetime.datetime.combine(target_date, datetime.time(9, 0), tzinfo=tzinfo)
    entregado_at = datetime.datetime.combine(target_date, datetime.time(16, 0), tzinfo=tzinfo)

    if now_local >= entregado_at:
        return 'entregado'
    if now_local >= enviado_at:
        return 'enviado'
    return None


def _merge_order_statuses(*values: Any) -> str:
    normed = []
    for v in values:
        if v is None:
            continue
        vv = _normalize_order_status(v)
        if vv:
            normed.append(vv)
    if any(v == 'cancelado' for v in normed):
        return 'cancelado'
    best = 'recibido'
    best_rank = _order_status_rank(best)
    for v in normed:
        r = _order_status_rank(v)
        if r > best_rank:
            best = v
            best_rank = r
    return best


def _normalize_email(value: Any) -> Optional[str]:
    if value is None:
        return None
    try:
        email = str(value).strip().lower()
    except Exception:
        return None
    if not email or '@' not in email or ' ' in email:
        return None
    return email


def _extract_customer_email_for_order(
    order_data: Optional[Dict[str, Any]],
    request_data: Optional[Dict[str, Any]] = None,
    token_payload: Optional[Dict[str, Any]] = None,
) -> Optional[str]:
    candidates: List[Any] = []

    if isinstance(order_data, dict):
        candidates.append(order_data.get('user_email'))
        token_preview = order_data.get('_token_preview')
        if isinstance(token_preview, dict):
            candidates.append(token_preview.get('email'))

    if isinstance(request_data, dict):
        candidates.append(request_data.get('user_email'))
        token_preview = request_data.get('_token_preview')
        if isinstance(token_preview, dict):
            candidates.append(token_preview.get('email'))

    if isinstance(token_payload, dict):
        candidates.append(token_payload.get('email'))
        candidates.append(token_payload.get('sub'))

    for candidate in candidates:
        normalized = _normalize_email(candidate)
        if normalized:
            return normalized
    return None


def _resolve_user_email_by_id(user_id: Any) -> Optional[str]:
    try:
        uid = int(user_id)
    except Exception:
        return None
    db = SessionLocal()
    try:
        try:
            from app import models as _models
            user = db.query(_models.User).filter(_models.User.id == uid).first()
            if not user:
                return None
            return _normalize_email(getattr(user, 'email', None))
        except Exception:
            return None
    finally:
        try:
            db.close()
        except Exception:
            pass


def _enrich_order_contact_fields(order_data: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    """Best-effort enrich order payload with contact fields when DB lacks user_* columns."""
    payload = dict(order_data) if isinstance(order_data, dict) else {}
    order_id = payload.get('id')
    order_id_text = str(order_id).strip() if order_id is not None else ''

    def _apply_candidate(candidate: Any) -> None:
        if not isinstance(candidate, dict):
            return
        for field in ('user_id', 'user_full_name', 'user_email', 'user_barrio', 'user_calle', 'user_numeracion', 'user_postal_code', 'user_department'):
            if not payload.get(field) and candidate.get(field):
                payload[field] = candidate.get(field)
        for field in ('scheduled_delivery_date', 'delivery_cutoff_applied', 'delivery_timezone', 'delivery_cutoff_hour'):
            current = payload.get(field)
            if (current is None or (isinstance(current, str) and not current.strip())) and (field in candidate):
                payload[field] = candidate.get(field)

        token_preview = candidate.get('_token_preview')
        if isinstance(token_preview, dict):
            nested_address = token_preview.get('address') if isinstance(token_preview.get('address'), dict) else {}
            if not payload.get('_token_preview'):
                payload['_token_preview'] = token_preview
            if not payload.get('user_full_name') and token_preview.get('name'):
                payload['user_full_name'] = token_preview.get('name')
            if not payload.get('user_email') and token_preview.get('email'):
                payload['user_email'] = token_preview.get('email')
            if not payload.get('user_postal_code'):
                payload['user_postal_code'] = (
                    token_preview.get('postal_code') or token_preview.get('user_postal_code') or
                    nested_address.get('postal_code') or nested_address.get('postcode')
                )
            if not payload.get('user_department'):
                payload['user_department'] = token_preview.get('department') or token_preview.get('user_department') or nested_address.get('department')
            if '_token_received' not in payload:
                payload['_token_received'] = True

    # 1) In-memory payload cache (fast path).
    if order_id_text and (not payload.get('user_email') or not payload.get('user_full_name')):
        try:
            cached_entry = ORDER_PAYLOAD_CACHE.get(order_id_text) or {}
            _apply_candidate(cached_entry.get('payload'))
        except Exception:
            pass

    # 2) Persisted preview table (durable path).
    if order_id_text and (not payload.get('user_email') or not payload.get('user_full_name')):
        try:
            row = _safe_engine_fetchone(
                'SELECT token_preview, token_received FROM order_token_previews WHERE order_id = :id ORDER BY created_at DESC LIMIT 1',
                {'id': order_id_text},
            )
            if row:
                tp_raw = row[0]
                token_received = bool(row[1]) if len(row) > 1 else False
                token_preview = {}
                if tp_raw:
                    try:
                        token_preview = json.loads(tp_raw) if isinstance(tp_raw, str) else (tp_raw if isinstance(tp_raw, dict) else {})
                    except Exception:
                        token_preview = {}
                candidate = {'_token_preview': token_preview, '_token_received': token_received}
                if isinstance(token_preview, dict):
                    nested_address = token_preview.get('address') if isinstance(token_preview.get('address'), dict) else {}
                    if token_preview.get('email'):
                        candidate['user_email'] = token_preview.get('email')
                    if token_preview.get('name'):
                        candidate['user_full_name'] = token_preview.get('name')
                    if token_preview.get('barrio'):
                        candidate['user_barrio'] = token_preview.get('barrio')
                    if token_preview.get('calle'):
                        candidate['user_calle'] = token_preview.get('calle')
                    if token_preview.get('numeracion'):
                        candidate['user_numeracion'] = token_preview.get('numeracion')
                    candidate['user_postal_code'] = (
                        token_preview.get('postal_code') or token_preview.get('user_postal_code') or
                        nested_address.get('postal_code') or nested_address.get('postcode')
                    )
                    candidate['user_department'] = (
                        token_preview.get('department') or token_preview.get('user_department') or
                        nested_address.get('department')
                    )
                _apply_candidate(candidate)
        except Exception:
            pass

    # 3) Last fallback: resolve by user_id from users table.
    if not payload.get('user_email') and payload.get('user_id') is not None:
        resolved = _resolve_user_email_by_id(payload.get('user_id'))
        if resolved:
            payload['user_email'] = resolved

    return payload


MENDOZA_GEO_BOUNDS = {
    'min_lat': -37.7,
    'max_lat': -31.0,
    'min_lon': -70.7,
    'max_lon': -66.2,
}

_MENDOZA_POSTAL_TO_DEPARTMENTS = {
    '5500': ['Capital'],
    '5501': ['Godoy Cruz'],
    '5502': ['Godoy Cruz'],
    '5503': ['Godoy Cruz'],
    '5507': ['Lujan de Cuyo'],
    '5509': ['Lujan de Cuyo'],
    '5511': ['Lujan de Cuyo'],
    '5513': ['Maipu'],
    '5515': ['Maipu'],
    '5517': ['Maipu'],
    '5519': ['Guaymallen'],
    '5521': ['Guaymallen'],
    '5523': ['Guaymallen'],
    '5525': ['Guaymallen'],
    '5533': ['Lavalle'],
    '5535': ['Lavalle'],
    '5539': ['Las Heras'],
    '5540': ['Las Heras'],
    '5541': ['Las Heras'],
    '5549': ['Lujan de Cuyo'],
    '5560': ['Tunuyan'],
    '5561': ['Tupungato'],
    '5569': ['San Carlos'],
    '5570': ['San Martin'],
    '5573': ['San Martin', 'Junin'],
    '5575': ['Junin'],
    '5577': ['Rivadavia'],
    '5590': ['La Paz'],
    '5596': ['Santa Rosa'],
    '5600': ['San Rafael'],
    '5603': ['San Rafael'],
    '5613': ['Malargue'],
    '5620': ['General Alvear'],
}


def _normalize_region_token(value: Any) -> str:
    try:
        text_value = str(value or '').strip().lower()
    except Exception:
        text_value = ''
    if not text_value:
        return ''
    try:
        text_value = ''.join(
            ch for ch in unicodedata.normalize('NFD', text_value)
            if unicodedata.category(ch) != 'Mn'
        )
    except Exception:
        pass
    return re.sub(r'\s+', ' ', text_value).strip()


_MENDOZA_DEPARTMENTS = sorted(
    {dep for deps in _MENDOZA_POSTAL_TO_DEPARTMENTS.values() for dep in deps},
    key=lambda value: len(_normalize_region_token(value)),
    reverse=True,
)


def _extract_postal_digits(value: Any) -> str:
    try:
        raw = str(value or '').strip()
    except Exception:
        raw = ''
    if not raw:
        return ''
    match = re.search(r'\bM?\s*(\d{4})\b', raw, re.IGNORECASE)
    if not match:
        return ''
    return str(match.group(1) or '').strip()


def _normalize_postal_code(value: Any) -> str:
    digits = _extract_postal_digits(value)
    return f'M{digits}' if digits else ''


def _extract_postal_from_text(value: Any) -> str:
    try:
        text_value = str(value or '')
    except Exception:
        text_value = ''
    return _normalize_postal_code(text_value)


def _departments_for_postal(postal_code: Any) -> List[str]:
    digits = _extract_postal_digits(postal_code)
    if not digits:
        return []
    return list(_MENDOZA_POSTAL_TO_DEPARTMENTS.get(digits) or [])


def _department_from_hints(values: Any) -> str:
    candidates = values if isinstance(values, (list, tuple, set)) else [values]
    probe = _normalize_region_token(' '.join(str(v or '') for v in candidates if v))
    if not probe:
        return ''
    for department in _MENDOZA_DEPARTMENTS:
        department_token = _normalize_region_token(department)
        if department_token and department_token in probe:
            return str(department or '')
    return ''


def _postal_matches_department(postal_code: Any, department: Any) -> bool:
    postal = _normalize_postal_code(postal_code)
    department_token = _normalize_region_token(department)
    if not postal or not department_token:
        return True
    departments = _departments_for_postal(postal)
    if not departments:
        return True
    return any(_normalize_region_token(dep) == department_token for dep in departments)


def _select_postal_for_department(postal_code: Any, department: Any, hints: Optional[List[Any]] = None) -> str:
    resolved_department = str(
        department or _department_from_hints(hints or []) or ''
    ).strip()
    direct_postal = _normalize_postal_code(postal_code)
    if direct_postal and _postal_matches_department(direct_postal, resolved_department):
        return direct_postal
    for hint in (hints or []):
        candidate = _normalize_postal_code(hint)
        if candidate and _postal_matches_department(candidate, resolved_department):
            return candidate
    return direct_postal if direct_postal and not resolved_department else ''


def _coerce_coord(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        if isinstance(value, str):
            value = value.strip().replace(',', '.')
        n = float(value)
    except Exception:
        return None
    if not (-180.0 <= n <= 180.0):
        return None
    return round(n, 6)


def _is_mendoza_point(lat: Any, lon: Any) -> bool:
    lat_n = _coerce_coord(lat)
    lon_n = _coerce_coord(lon)
    if lat_n is None or lon_n is None:
        return False
    return (
        MENDOZA_GEO_BOUNDS['min_lat'] <= lat_n <= MENDOZA_GEO_BOUNDS['max_lat'] and
        MENDOZA_GEO_BOUNDS['min_lon'] <= lon_n <= MENDOZA_GEO_BOUNDS['max_lon']
    )


def _extract_order_address_snapshot(order_data: Dict[str, Any]) -> Dict[str, Any]:
    payload = order_data if isinstance(order_data, dict) else {}
    token_preview = payload.get('_token_preview')
    if isinstance(token_preview, str):
        try:
            token_preview = json.loads(token_preview)
        except Exception:
            token_preview = {}
    if not isinstance(token_preview, dict):
        token_preview = {}
    nested_address = token_preview.get('address')
    if not isinstance(nested_address, dict):
        nested_address = {}

    barrio = (
        payload.get('user_barrio') or payload.get('barrio') or payload.get('user_neighborhood') or
        token_preview.get('barrio') or nested_address.get('barrio') or token_preview.get('city') or nested_address.get('city') or
        ''
    )
    calle = (
        payload.get('user_calle') or payload.get('calle') or payload.get('user_street') or
        token_preview.get('calle') or nested_address.get('calle') or nested_address.get('street') or nested_address.get('road') or
        ''
    )
    numeracion = (
        payload.get('user_numeracion') or payload.get('numeracion') or payload.get('user_number') or
        token_preview.get('numeracion') or nested_address.get('numeracion') or nested_address.get('number') or nested_address.get('house_number') or
        ''
    )
    raw_address = (
        payload.get('user_address') or payload.get('user_direccion') or payload.get('delivery_address') or
        payload.get('shipping_address') or payload.get('address') or payload.get('user_full_address') or
        payload.get('full_address') or payload.get('direccion') or
        payload.get('query_hint') or payload.get('user_address_label') or
        token_preview.get('user_address') or token_preview.get('direccion') or token_preview.get('query_hint') or token_preview.get('full_text') or token_preview.get('label') or
        nested_address.get('direccion') or nested_address.get('query_hint') or nested_address.get('full_text') or nested_address.get('display_name') or
        ''
    )

    postal_candidates = [
        payload.get('user_postal_code'),
        payload.get('postal_code'),
        payload.get('postcode'),
        payload.get('zip_code'),
        payload.get('zip'),
        token_preview.get('postal_code'),
        token_preview.get('postcode'),
        token_preview.get('user_postal_code'),
        token_preview.get('query_hint'),
        nested_address.get('postal_code'),
        nested_address.get('postcode'),
        nested_address.get('zip_code'),
        nested_address.get('query_hint'),
        raw_address,
        barrio,
    ]
    postal_code = ''
    for candidate in postal_candidates:
        postal_code = _normalize_postal_code(candidate)
        if postal_code:
            break

    department = (
        payload.get('user_department') or payload.get('department') or
        token_preview.get('department') or token_preview.get('user_department') or
        nested_address.get('department') or nested_address.get('county') or nested_address.get('state_district') or
        ''
    )
    department = str(department or '').strip()
    hinted_department = _department_from_hints([
        department,
        barrio,
        raw_address,
        payload.get('query_hint'),
        token_preview.get('query_hint'),
        token_preview.get('label'),
        token_preview.get('full_text'),
        nested_address.get('query_hint'),
        nested_address.get('display_name'),
    ])
    if not department and hinted_department:
        department = hinted_department

    department_candidates = _departments_for_postal(postal_code)
    if department_candidates:
        probe = _normalize_region_token(' '.join([
            str(department or ''),
            str(barrio or ''),
            str(raw_address or ''),
            str(token_preview.get('query_hint') or ''),
            str(token_preview.get('label') or ''),
            str(token_preview.get('full_text') or ''),
            str(nested_address.get('query_hint') or ''),
        ]))
        matched_department = ''
        if probe:
            for dep_name in department_candidates:
                dep_token = _normalize_region_token(dep_name)
                if dep_token and dep_token in probe:
                    matched_department = dep_name
                    break
        if not matched_department and hinted_department:
            hinted_token = _normalize_region_token(hinted_department)
            for dep_name in department_candidates:
                if hinted_token and hinted_token == _normalize_region_token(dep_name):
                    matched_department = dep_name
                    break
        if not matched_department:
            matched_department = ''
        if not department:
            department = matched_department or hinted_department or ''

    postal_code = _select_postal_for_department(postal_code, department, [
        raw_address,
        barrio,
        payload.get('query_hint'),
        token_preview.get('query_hint'),
        token_preview.get('label'),
        token_preview.get('full_text'),
        nested_address.get('query_hint'),
        nested_address.get('display_name'),
    ])

    lat_candidates = [
        payload.get('user_lat'),
        payload.get('user_latitude'),
        payload.get('delivery_lat'),
        payload.get('lat'),
        token_preview.get('user_lat'),
        token_preview.get('lat'),
        token_preview.get('latitude'),
        token_preview.get('delivery_lat'),
        nested_address.get('lat'),
        nested_address.get('latitude'),
    ]
    lon_candidates = [
        payload.get('user_lon'),
        payload.get('user_lng'),
        payload.get('user_longitude'),
        payload.get('delivery_lon'),
        payload.get('delivery_lng'),
        payload.get('lon'),
        payload.get('lng'),
        token_preview.get('user_lon'),
        token_preview.get('user_lng'),
        token_preview.get('lon'),
        token_preview.get('lng'),
        token_preview.get('longitude'),
        token_preview.get('delivery_lon'),
        token_preview.get('delivery_lng'),
        nested_address.get('lon'),
        nested_address.get('lng'),
        nested_address.get('longitude'),
    ]
    lat = next((v for v in (_coerce_coord(c) for c in lat_candidates) if v is not None), None)
    lon = next((v for v in (_coerce_coord(c) for c in lon_candidates) if v is not None), None)

    return {
        'barrio': str(barrio or '').strip(),
        'calle': str(calle or '').strip(),
        'numeracion': str(numeracion or '').strip(),
        'raw_address': str(raw_address or '').strip(),
        'postal_code': postal_code,
        'department': str(department or '').strip(),
        'lat': lat,
        'lon': lon,
    }


def _compose_order_maps_query(snapshot: Dict[str, Any]) -> str:
    if not isinstance(snapshot, dict):
        return ''
    street = ' '.join(
        part for part in [
            str(snapshot.get('calle') or '').strip(),
            str(snapshot.get('numeracion') or '').strip(),
        ] if part
    ).strip()
    barrio = str(snapshot.get('barrio') or '').strip()
    raw_address = str(snapshot.get('raw_address') or '').strip()
    postal_code = _normalize_postal_code(snapshot.get('postal_code'))
    department = str(snapshot.get('department') or '').strip()
    if not department:
        department = _department_from_hints([barrio, raw_address, street])

    if not department:
        deps = _departments_for_postal(postal_code)
        if deps:
            department = deps[0]

    postal_code = _select_postal_for_department(postal_code, department, [raw_address, barrio, street])

    def _looks_like_street(value: Any) -> bool:
        try:
            text_value = str(value or '').strip()
        except Exception:
            text_value = ''
        if not text_value:
            return False
        token = _normalize_region_token(text_value)
        if not token:
            return False
        if any(bad in token for bad in ('direccion', 'domicilio', 'sin direccion', 'ubicacion')):
            return False
        has_letter = bool(re.search(r'[A-Za-z\u00c0-\u024f]', text_value))
        has_number = bool(re.search(r'\b\d{1,6}\b', text_value))
        if has_letter and has_number:
            return True
        # If we have contextual fields, allow streets without number.
        if has_letter and (barrio or department or postal_code):
            return True
        return False

    first_raw_segment = str(raw_address.split(',')[0] or '').strip()
    street_for_query = ''
    if _looks_like_street(street):
        street_for_query = street
    elif _looks_like_street(first_raw_segment):
        street_for_query = first_raw_segment

    parts: List[str] = []
    seen: set = set()

    def _push(value: Any) -> None:
        text_value = str(value or '').strip()
        if not text_value:
            return
        token = _normalize_region_token(text_value)
        if not token or token in seen:
            return
        seen.add(token)
        parts.append(text_value)

    if street_for_query:
        _push(street_for_query)
    _push(barrio)
    _push(department)
    _push(postal_code)
    if raw_address and not street_for_query:
        # Keep broader text only when no usable street was found.
        _push(raw_address)
    _push('Mendoza')
    _push('Argentina')

    # Avoid sending low-quality ambiguous queries to Maps.
    quality_pieces = [
        bool(street_for_query),
        bool(barrio),
        bool(department),
        bool(postal_code),
    ]
    if sum(1 for flag in quality_pieces if flag) < 2:
        return ''

    return ', '.join(parts)


def _normalize_geocode_query(query: Any) -> str:
    try:
        return ' '.join(str(query or '').strip().split())
    except Exception:
        return ''


def _geocode_cache_key(provider: str, query: str) -> str:
    return f"{provider}:{_normalize_geocode_query(query).lower()}"


def _load_geocode_cache(db: Optional[Session]) -> None:
    global _GEOCODE_CACHE_LOADED, _GEOCODE_CACHE
    if not GEOCODE_PERSIST_CACHE or _GEOCODE_CACHE_LOADED or db is None:
        return
    try:
        cached = crud.get_setting(db, 'geocode_cache_v1')
        cleaned: Dict[str, Dict[str, Any]] = {}
        if isinstance(cached, dict):
            for key, val in cached.items():
                if not isinstance(val, dict):
                    continue
                lat = _coerce_coord(val.get('lat'))
                lon = _coerce_coord(val.get('lon'))
                try:
                    ts = float(val.get('ts') or 0.0)
                except Exception:
                    ts = 0.0
                ok = bool(val.get('ok')) if 'ok' in val else _is_mendoza_point(lat, lon)
                cleaned[str(key)] = {'lat': lat, 'lon': lon, 'ts': ts, 'ok': ok}
        with _GEOCODE_CACHE_LOCK:
            _GEOCODE_CACHE.update(cleaned)
    except Exception:
        logger.exception('load geocode cache failed')
    finally:
        _GEOCODE_CACHE_LOADED = True


def _prune_geocode_cache_unlocked() -> None:
    if GEOCODE_CACHE_MAX <= 0:
        _GEOCODE_CACHE.clear()
        return
    if len(_GEOCODE_CACHE) <= GEOCODE_CACHE_MAX:
        return
    items = sorted(_GEOCODE_CACHE.items(), key=lambda kv: float(kv[1].get('ts') or 0.0))
    for key, _ in items[: max(0, len(items) - GEOCODE_CACHE_MAX)]:
        _GEOCODE_CACHE.pop(key, None)


def _geocode_cache_get(key: str, db: Optional[Session] = None) -> Optional[Dict[str, Any]]:
    global _GEOCODE_CACHE_DIRTY
    if GEOCODE_CACHE_MAX <= 0:
        return None
    if GEOCODE_PERSIST_CACHE:
        _load_geocode_cache(db)
    now = time.time()
    with _GEOCODE_CACHE_LOCK:
        entry = _GEOCODE_CACHE.get(key)
        if not entry:
            return None
        ts = float(entry.get('ts') or 0.0)
        if GEOCODE_CACHE_TTL_SEC > 0 and ts and (now - ts) > GEOCODE_CACHE_TTL_SEC:
            _GEOCODE_CACHE.pop(key, None)
            _GEOCODE_CACHE_DIRTY = True
            return None
        return dict(entry)


def _persist_geocode_cache(db: Optional[Session]) -> None:
    global _GEOCODE_CACHE_DIRTY
    if not GEOCODE_PERSIST_CACHE or not _GEOCODE_CACHE_DIRTY or db is None:
        return
    try:
        with _GEOCODE_CACHE_LOCK:
            snapshot = dict(_GEOCODE_CACHE)
            _GEOCODE_CACHE_DIRTY = False
        crud.set_setting(db, 'geocode_cache_v1', snapshot)
    except Exception:
        logger.exception('persist geocode cache failed')


def _geocode_cache_set(key: str, lat: Optional[float], lon: Optional[float], ok: bool, db: Optional[Session] = None) -> None:
    global _GEOCODE_CACHE_DIRTY
    if GEOCODE_CACHE_MAX <= 0:
        return
    with _GEOCODE_CACHE_LOCK:
        _GEOCODE_CACHE[key] = {
            'lat': _coerce_coord(lat),
            'lon': _coerce_coord(lon),
            'ts': time.time(),
            'ok': bool(ok),
        }
        _prune_geocode_cache_unlocked()
        _GEOCODE_CACHE_DIRTY = True
    _persist_geocode_cache(db)


def _geocode_budget_ok() -> bool:
    if GEOCODE_MAX_PER_RUN <= 0:
        return False
    now = time.time()
    with _GEOCODE_CACHE_LOCK:
        window_start = float(_GEOCODE_RUN_STATS.get('window_start') or 0.0)
        if (now - window_start) > GEOCODE_RUN_WINDOW_SEC:
            _GEOCODE_RUN_STATS['window_start'] = now
            _GEOCODE_RUN_STATS['count'] = 0
        count = int(_GEOCODE_RUN_STATS.get('count') or 0)
        if count >= GEOCODE_MAX_PER_RUN:
            return False
        _GEOCODE_RUN_STATS['count'] = count + 1
    return True


def _geocode_rate_limit() -> None:
    global _GEOCODE_LAST_CALL_TS
    if GEOCODE_RATE_LIMIT_MS <= 0:
        return
    now = time.time()
    wait_for = (GEOCODE_RATE_LIMIT_MS / 1000.0) - (now - _GEOCODE_LAST_CALL_TS)
    if wait_for > 0:
        time.sleep(wait_for)
    _GEOCODE_LAST_CALL_TS = time.time()


def _geocode_fetch_nominatim(query: str) -> Tuple[Optional[float], Optional[float], bool]:
    params = {
        'q': query,
        'format': 'json',
        'limit': 1,
        'addressdetails': 0,
        'countrycodes': 'ar',
    }
    try:
        params['viewbox'] = f"{MENDOZA_GEO_BOUNDS['min_lon']},{MENDOZA_GEO_BOUNDS['max_lat']},{MENDOZA_GEO_BOUNDS['max_lon']},{MENDOZA_GEO_BOUNDS['min_lat']}"
        params['bounded'] = 1
    except Exception:
        pass
    if GEOCODE_NOMINATIM_EMAIL:
        params['email'] = GEOCODE_NOMINATIM_EMAIL
    headers = {'User-Agent': GEOCODE_USER_AGENT}
    try:
        resp = httpx.get(GEOCODE_NOMINATIM_URL, params=params, headers=headers, timeout=GEOCODE_TIMEOUT_SEC)
        resp.raise_for_status()
        data = resp.json()
        if isinstance(data, list) and data:
            lat = _coerce_coord(data[0].get('lat'))
            lon = _coerce_coord(data[0].get('lon'))
            return lat, lon, True
        return None, None, True
    except Exception:
        logger.warning('geocode nominatim failed')
        return None, None, False


def _geocode_fetch_google(query: str) -> Tuple[Optional[float], Optional[float], bool]:
    if not GEOCODE_GOOGLE_API_KEY:
        return None, None, False
    params = {
        'address': query,
        'key': GEOCODE_GOOGLE_API_KEY,
        'region': 'ar',
    }
    try:
        params['bounds'] = f"{MENDOZA_GEO_BOUNDS['min_lat']},{MENDOZA_GEO_BOUNDS['min_lon']}|{MENDOZA_GEO_BOUNDS['max_lat']},{MENDOZA_GEO_BOUNDS['max_lon']}"
    except Exception:
        pass
    try:
        resp = httpx.get('https://maps.googleapis.com/maps/api/geocode/json', params=params, timeout=GEOCODE_TIMEOUT_SEC)
        resp.raise_for_status()
        data = resp.json()
        if not isinstance(data, dict):
            return None, None, True
        status = str(data.get('status') or '').upper()
        if status != 'OK':
            return None, None, True
        results = data.get('results') or []
        if results:
            loc = (results[0].get('geometry') or {}).get('location') or {}
            lat = _coerce_coord(loc.get('lat'))
            lon = _coerce_coord(loc.get('lng'))
            return lat, lon, True
        return None, None, True
    except Exception:
        logger.warning('geocode google failed')
        return None, None, False


def _geocode_lookup(query: str, db: Optional[Session] = None) -> Tuple[Optional[float], Optional[float]]:
    if not GEOCODE_ENABLED:
        return None, None
    q = _normalize_geocode_query(query)
    if not q:
        return None, None
    provider = GEOCODE_PROVIDER
    if provider == 'google' and not GEOCODE_GOOGLE_API_KEY:
        provider = 'nominatim'
    key = _geocode_cache_key(provider, q)
    cached = _geocode_cache_get(key, db=db)
    if cached:
        if cached.get('ok') and _is_mendoza_point(cached.get('lat'), cached.get('lon')):
            return cached.get('lat'), cached.get('lon')
        return None, None
    if not _geocode_budget_ok():
        return None, None
    _geocode_rate_limit()
    if provider == 'google':
        lat, lon, ok_resp = _geocode_fetch_google(q)
    else:
        lat, lon, ok_resp = _geocode_fetch_nominatim(q)
    if ok_resp:
        ok = _is_mendoza_point(lat, lon)
        _geocode_cache_set(key, lat, lon, ok, db=db)
    if _is_mendoza_point(lat, lon):
        return float(lat), float(lon)
    return None, None


def _geocode_order_snapshot(snapshot: Dict[str, Any], db: Optional[Session] = None) -> Tuple[Optional[float], Optional[float]]:
    if not GEOCODE_ENABLED:
        return None, None
    query = _compose_order_maps_query(snapshot)
    if not query:
        return None, None
    return _geocode_lookup(query, db=db)


def _get_route_start_point(db: Optional[Session] = None) -> Tuple[Optional[float], Optional[float]]:
    """Return fixed depot start point for routing, if configured."""
    try:
        cached = _ROUTE_START_CACHE.get('point')
        if isinstance(cached, tuple) and len(cached) == 2:
            lat_c, lon_c = cached
            if _is_mendoza_point(lat_c, lon_c):
                return float(lat_c), float(lon_c)
    except Exception:
        pass
    lat = _coerce_coord(ROUTE_START_LAT_RAW)
    lon = _coerce_coord(ROUTE_START_LON_RAW)
    if _is_mendoza_point(lat, lon):
        _ROUTE_START_CACHE['point'] = (float(lat), float(lon))
        return float(lat), float(lon)
    address = str(ROUTE_START_ADDRESS or '').strip()
    if address:
        query = address
        if 'mendoza' not in query.lower():
            query = f"{query}, Mendoza, Argentina"
        try:
            lat, lon = _geocode_lookup(query, db=db)
        except Exception:
            lat, lon = None, None
        if _is_mendoza_point(lat, lon):
            _ROUTE_START_CACHE['point'] = (float(lat), float(lon))
            return float(lat), float(lon)
    return None, None


def _build_order_maps_url(order_data: Dict[str, Any]) -> str:
    snapshot = _extract_order_address_snapshot(order_data if isinstance(order_data, dict) else {})
    lat = snapshot.get('lat')
    lon = snapshot.get('lon')
    if _is_mendoza_point(lat, lon):
        coords_query = f"{float(lat):.6f},{float(lon):.6f}"
        return f"https://www.google.com/maps/search/?api=1&query={quote_plus(coords_query)}"
    query = _compose_order_maps_query(snapshot)
    postal_code = _normalize_postal_code(snapshot.get('postal_code'))
    department = str(snapshot.get('department') or '').strip()
    if query and (postal_code or department):
        return f"https://www.google.com/maps/search/?api=1&query={quote_plus(query)}"
    if not query:
        return ''
    return f"https://www.google.com/maps/search/?api=1&query={quote_plus(query)}"


def _attach_maps_url(order_data: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    payload = dict(order_data) if isinstance(order_data, dict) else {}
    existing_maps_url = str(payload.get('maps_url') or '').strip()
    try:
        maps_url = _build_order_maps_url(payload)
    except Exception:
        maps_url = ''
    if maps_url:
        payload['maps_url'] = maps_url
    elif existing_maps_url:
        payload['maps_url'] = existing_maps_url
    return payload


def _build_order_confirmation_subject(order_data: Dict[str, Any]) -> str:
    order_id = order_data.get('id')
    if order_id is None:
        return 'Recibimos tu pedido'
    return f'Recibimos tu pedido #{order_id}'


def _build_order_seen_subject(order_data: Dict[str, Any]) -> str:
    order_id = order_data.get('id')
    if order_id is None:
        return 'Tu pedido ya fue visto por administracion'
    return f'Tu pedido #{order_id} ya fue visto por administracion'


def _build_order_prepared_subject(order_data: Dict[str, Any]) -> str:
    order_id = order_data.get('id')
    if order_id is None:
        return 'Tu pedido ya fue preparado'
    return f'Tu pedido #{order_id} ya fue preparado'


def _order_payment_ui(order_data: Dict[str, Any]) -> Dict[str, str]:
    method_raw = str(order_data.get('payment_method') or '').strip().lower()
    status_raw = str(order_data.get('payment_status') or '').strip().lower()
    reference_raw = str(order_data.get('payment_reference') or '').strip()

    method_map = {
        'mercadopago': 'Mercado Pago',
        'mp': 'Mercado Pago',
        'mercado_pago': 'Mercado Pago',
        'cash': 'Efectivo',
        'efectivo': 'Efectivo',
    }
    status_map = {
        'mp_pending': 'Pendiente',
        'pending': 'Pendiente',
        'in_process': 'Pendiente',
        'inprocess': 'Pendiente',
        'authorized': 'Pendiente',
        'cash_pending': 'A cobrar en entrega',
        'approved': 'Aprobado',
        'accredited': 'Aprobado',
        'rejected': 'Rechazado',
        'cancelled': 'Cancelado',
        'cancelled_by_user': 'Cancelado',
        'refunded': 'Reintegrado',
        'charged_back': 'Reintegrado',
    }

    method_label = method_map.get(method_raw, method_raw.replace('_', ' ').title() if method_raw else 'No especificado')
    status_label = status_map.get(status_raw, status_raw.replace('_', ' ').title() if status_raw else 'Pendiente')

    status_bg = '#fff7ed'
    status_color = '#9a3412'
    if status_raw in ('approved', 'accredited'):
        status_bg = '#ecfdf3'
        status_color = '#166534'
    elif status_raw in ('rejected', 'cancelled', 'cancelled_by_user'):
        status_bg = '#fef2f2'
        status_color = '#991b1b'
    elif status_raw in ('refunded', 'charged_back'):
        status_bg = '#eff6ff'
        status_color = '#1d4ed8'
    elif status_raw in ('cash_pending',):
        status_bg = '#ecfeff'
        status_color = '#0f766e'

    return {
        'method_label': method_label,
        'status_label': status_label,
        'status_bg': status_bg,
        'status_color': status_color,
        'reference': reference_raw,
    }


def _parse_bool_like(value: Any) -> Optional[bool]:
    if isinstance(value, bool):
        return value
    if value is None:
        return None
    raw = str(value).strip().lower()
    if not raw:
        return None
    if raw in ('1', 'true', 'yes', 'si', 'on'):
        return True
    if raw in ('0', 'false', 'no', 'off'):
        return False
    return None


def _normalize_iso_date_key(value: Any) -> str:
    raw = str(value or '').strip()
    if not raw:
        return ''
    candidate = raw[:10]
    try:
        parsed = datetime.date.fromisoformat(candidate)
        return parsed.isoformat()
    except Exception:
        return ''


def _parse_datetime_utc(value: Any) -> Optional[datetime.datetime]:
    if value is None:
        return None
    try:
        parsed = value if isinstance(value, datetime.datetime) else datetime.datetime.fromisoformat(str(value).replace('Z', '+00:00'))
    except Exception:
        return None
    try:
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=datetime.timezone.utc)
        return parsed.astimezone(datetime.timezone.utc)
    except Exception:
        return None


def _resolve_delivery_schedule(order_data: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    data = order_data if isinstance(order_data, dict) else {}
    timezone_name = str(
        data.get('delivery_timezone')
        or os.environ.get('ORDER_DELIVERY_TIMEZONE')
        or os.environ.get('ORDER_EMAIL_TIMEZONE')
        or 'America/Argentina/Buenos_Aires'
    ).strip() or 'America/Argentina/Buenos_Aires'

    raw_cutoff_hour = data.get('delivery_cutoff_hour')
    if raw_cutoff_hour is None or str(raw_cutoff_hour).strip() == '':
        raw_cutoff_hour = os.environ.get('ORDER_CUTOFF_HOUR', 18)
    try:
        cutoff_hour = int(str(raw_cutoff_hour).strip())
    except Exception:
        cutoff_hour = 18
    cutoff_hour = max(0, min(23, cutoff_hour))

    scheduled_delivery_date = _normalize_iso_date_key(data.get('scheduled_delivery_date'))
    delivery_cutoff_applied = _parse_bool_like(data.get('delivery_cutoff_applied'))

    if not scheduled_delivery_date or delivery_cutoff_applied is None:
        base_dt_utc = _parse_datetime_utc(data.get('created_at')) or datetime.datetime.now(datetime.timezone.utc)
        try:
            computed = crud._compute_delivery_schedule_snapshot(
                now_utc=base_dt_utc,
                cutoff_hour=cutoff_hour,
                timezone_name=timezone_name,
            )
        except Exception:
            computed = {}
        if not scheduled_delivery_date:
            scheduled_delivery_date = _normalize_iso_date_key(computed.get('scheduled_delivery_date'))
        if delivery_cutoff_applied is None and 'delivery_cutoff_applied' in computed:
            try:
                delivery_cutoff_applied = bool(computed.get('delivery_cutoff_applied'))
            except Exception:
                delivery_cutoff_applied = None
        try:
            tz_computed = str(computed.get('delivery_timezone') or '').strip()
            if tz_computed:
                timezone_name = tz_computed
        except Exception:
            pass
        try:
            computed_cutoff = int(computed.get('delivery_cutoff_hour'))
            cutoff_hour = max(0, min(23, computed_cutoff))
        except Exception:
            pass

    if delivery_cutoff_applied is None:
        delivery_cutoff_applied = False

    return {
        'scheduled_delivery_date': scheduled_delivery_date,
        'delivery_cutoff_applied': bool(delivery_cutoff_applied),
        'delivery_timezone': timezone_name,
        'delivery_cutoff_hour': int(cutoff_hour),
    }


def _extract_delivery_time_window_text() -> str:
    configured = str(os.environ.get('ORDER_DELIVERY_TIME_WINDOW') or '').strip()
    if configured:
        return configured
    legacy = str(os.environ.get('ORDER_SEEN_DELIVERY_WINDOW') or '').strip()
    if not legacy:
        return ''
    legacy = legacy.replace('manana', 'ma\u00f1ana')
    lower = legacy.lower()
    if 'entre' in lower:
        return legacy[lower.find('entre'):].strip()
    return ''


def _format_delivery_schedule_label(order_data: Optional[Dict[str, Any]]) -> str:
    schedule = _resolve_delivery_schedule(order_data)
    date_key = _normalize_iso_date_key(schedule.get('scheduled_delivery_date'))
    if not date_key:
        legacy = str(os.environ.get('ORDER_SEEN_DELIVERY_WINDOW') or '').strip().replace('manana', 'ma\u00f1ana')
        return legacy or 'A confirmar'

    try:
        parsed_date = datetime.date.fromisoformat(date_key)
        weekday_map = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo']
        weekday = weekday_map[parsed_date.weekday()] if 0 <= parsed_date.weekday() < 7 else ''
        if weekday:
            label = weekday.capitalize() + ' ' + parsed_date.strftime('%d/%m/%Y')
        else:
            label = parsed_date.strftime('%d/%m/%Y')
    except Exception:
        label = date_key

    try:
        cutoff_hour = max(0, min(23, int(schedule.get('delivery_cutoff_hour'))))
    except Exception:
        cutoff_hour = 18

    if schedule.get('delivery_cutoff_applied'):
        label = label + f" (pedido despues de las {cutoff_hour:02d}:00)"

    delivery_window = _extract_delivery_time_window_text()
    if delivery_window:
        if delivery_window.lower().startswith('entre'):
            label = label + ' ' + delivery_window
        else:
            label = label + ' (' + delivery_window + ')'

    return label


def _format_ars_money(value: Any) -> str:
    """Format ARS with thousands '.' and decimals ',' (email-safe)."""
    try:
        num = float(value or 0)
    except Exception:
        return f"${html_escape(str(value or '0'))}"
    # Python uses 10,400.00 with this format; swap separators for es-AR style.
    txt = f"{num:,.2f}"
    txt = txt.replace(',', '_').replace('.', ',').replace('_', '.')
    return f"${txt}"


def _resolve_order_tracking_url(order_data: Optional[Dict[str, Any]]) -> Optional[str]:
    try:
        oid = order_data.get('id') if isinstance(order_data, dict) else None
    except Exception:
        oid = None
    if oid is None or str(oid).strip() == '':
        return None

    template = str(os.environ.get('ORDER_TRACKING_URL_TEMPLATE') or '').strip()
    if template:
        try:
            return template.replace('{order_id}', quote_plus(str(oid)))
        except Exception:
            return template

    base = str(os.environ.get('CATALOG_PUBLIC_URL') or os.environ.get('FRONTEND_PUBLIC_URL') or '').strip()
    if not base:
        return None
    url = base.rstrip('/')
    if not url.lower().endswith('.html'):
        url = url + '/catalogo.html'
    sep = '&' if '?' in url else '?'
    return url + sep + 'order=' + quote_plus(str(oid))


def _build_tracking_button_html(order_data: Optional[Dict[str, Any]]) -> str:
    url = _resolve_order_tracking_url(order_data)
    if not url:
        return ''
    safe_url = html_escape(str(url))
    return (
        "<div style='margin:16px 0 0 0;'>"
        f"<a href='{safe_url}' "
        "style='display:inline-block;background:#f26b38;color:#ffffff;text-decoration:none;"
        "font-weight:800;padding:12px 16px;border-radius:12px;'>"
        "Ver estado del pedido</a>"
        "</div>"
    )


def _build_brand_footer_text(brand_name: str) -> str:
    parts: List[str] = []
    try:
        address = str(os.environ.get('RESEND_BRAND_ADDRESS') or '').strip()
        if address:
            parts.append(address)
    except Exception:
        pass
    try:
        phone = str(os.environ.get('RESEND_BRAND_PHONE') or '').strip()
        if phone:
            parts.append(phone)
    except Exception:
        pass
    try:
        website = str(os.environ.get('RESEND_BRAND_WEBSITE') or '').strip()
        if website:
            parts.append(website)
    except Exception:
        pass
    if parts:
        return brand_name + " · " + " · ".join(parts)
    return brand_name


def _build_order_seen_html(order_data: Dict[str, Any]) -> str:
    brand_name = html_escape(str(os.environ.get('RESEND_BRAND_NAME') or 'DistriAr'))
    logo_url = (os.environ.get('RESEND_BRAND_LOGO_URL') or '').strip()
    order_id = html_escape(str(order_data.get('id') or 'sin-id'))
    customer_name = html_escape(str(order_data.get('user_full_name') or 'cliente'))
    delivery_window = html_escape(_format_delivery_schedule_label(order_data))

    created_at_raw = order_data.get('created_at')
    created_at_text = ''
    if created_at_raw:
        try:
            parsed_dt = datetime.datetime.fromisoformat(str(created_at_raw).replace('Z', '+00:00'))
            created_at_text = parsed_dt.strftime('%d/%m/%Y %H:%M')
        except Exception:
            created_at_text = str(created_at_raw)
    created_at_text = html_escape(created_at_text) if created_at_text else '-'

    seen_at_text = ''
    try:
        timezone_name = str(os.environ.get('ORDER_EMAIL_TIMEZONE') or 'America/Argentina/Buenos_Aires').strip()
        now_utc = datetime.datetime.now(datetime.timezone.utc)
        seen_dt = now_utc
        if timezone_name:
            try:
                from zoneinfo import ZoneInfo
                seen_dt = now_utc.astimezone(ZoneInfo(timezone_name))
            except Exception:
                seen_dt = now_utc
        seen_at_text = seen_dt.strftime('%d/%m/%Y %H:%M')
    except Exception:
        try:
            seen_at_text = datetime.datetime.utcnow().strftime('%d/%m/%Y %H:%M')
        except Exception:
            seen_at_text = ''
    seen_at_text = html_escape(seen_at_text) if seen_at_text else '-'

    total_raw = order_data.get('total')
    total_text = _format_ars_money(total_raw)

    email_val = _normalize_email(order_data.get('user_email'))
    email_text = html_escape(email_val) if email_val else '-'
    payment_ui = _order_payment_ui(order_data)
    payment_method_text = html_escape(payment_ui.get('method_label') or 'No especificado')
    payment_status_text = html_escape(payment_ui.get('status_label') or 'Pendiente')
    payment_status_badge = (
        "<span style='display:inline-block;padding:4px 10px;border-radius:999px;"
        f"background:{payment_ui.get('status_bg') or '#fff7ed'};"
        f"color:{payment_ui.get('status_color') or '#9a3412'};"
        "font-weight:700;font-size:12px;'>"
        f"{payment_status_text}</span>"
    )
    payment_ref_text = html_escape(payment_ui.get('reference') or '-') if payment_ui.get('reference') else '-'

    address_parts: List[str] = []
    calle = order_data.get('user_calle')
    numeracion = order_data.get('user_numeracion')
    barrio = order_data.get('user_barrio')
    if calle:
        address_parts.append(str(calle).strip())
    if numeracion:
        address_parts.append(str(numeracion).strip())
    if barrio:
        address_parts.append(f"Barrio {str(barrio).strip()}")
    delivery_address = html_escape(', '.join([p for p in address_parts if p])) if address_parts else '-'

    items = order_data.get('items') if isinstance(order_data.get('items'), list) else []
    rows: List[str] = []
    for idx, item in enumerate(items[:25], start=1):
        if isinstance(item, dict):
            prod_id_raw = item.get('id', '?')
            qty_raw = item.get('qty', '?')
            meta = item.get('meta') if isinstance(item.get('meta'), dict) else {}
            item_name = meta.get('name') or meta.get('title') or meta.get('product_name')
            qty_label = meta.get('qty_label')
        else:
            prod_id_raw = getattr(item, 'id', '?')
            qty_raw = getattr(item, 'qty', '?')
            meta = getattr(item, 'meta', None)
            meta = meta if isinstance(meta, dict) else {}
            item_name = meta.get('name') or meta.get('title') or meta.get('product_name')
            qty_label = meta.get('qty_label')

        try:
            qty_text = f"{float(qty_raw):g}"
        except Exception:
            qty_text = str(qty_raw)
        if qty_label:
            qty_text = f"{qty_text} ({str(qty_label)})"

        if not item_name:
            item_name = f"Producto #{prod_id_raw}"

        rows.append(
            "<tr>"
            f"<td style='padding:10px 12px;border-bottom:1px solid #edf1f5;color:#1f2937'>{idx}</td>"
            f"<td style='padding:10px 12px;border-bottom:1px solid #edf1f5;color:#1f2937'>{html_escape(str(item_name))}</td>"
            f"<td style='padding:10px 12px;border-bottom:1px solid #edf1f5;color:#334155;text-align:right'>{html_escape(qty_text)}</td>"
            "</tr>"
        )
    if not rows:
        rows.append(
            "<tr>"
            "<td colspan='3' style='padding:12px;color:#64748b;border-bottom:1px solid #edf1f5'>Sin detalle de productos</td>"
            "</tr>"
        )
    if len(items) > 25:
        rows.append(
            "<tr>"
            "<td colspan='3' style='padding:10px 12px;color:#64748b'>... y mas productos</td>"
            "</tr>"
        )

    support_email = html_escape(str(os.environ.get('RESEND_REPLY_TO') or ''))
    support_line = (
        f"<p style='margin:18px 0 0 0;font-size:13px;color:#64748b;'>"
        f"Si tenes dudas, respondenos a {support_email}."
        f"</p>"
        if support_email else
        "<p style='margin:18px 0 0 0;font-size:13px;color:#64748b;'>"
        "Si tenes dudas, respondenos a este correo."
        "</p>"
    )

    logo_block = (
        f"<img src='{html_escape(logo_url)}' alt='{brand_name}' style='height:42px;display:block;margin:0 0 12px 0;'/>"
        if logo_url else ""
    )
    tracking_button = _build_tracking_button_html(order_data)
    footer_text = html_escape(_build_brand_footer_text(str(os.environ.get('RESEND_BRAND_NAME') or 'DistriAr')))

    return (
        "<!doctype html>"
        "<html><body style='margin:0;padding:0;background:#f3f5f7;font-family:Arial,sans-serif;color:#0f172a;'>"
        "<table role='presentation' width='100%' cellspacing='0' cellpadding='0' style='background:#f3f5f7;'>"
        "<tr><td align='center' style='padding:24px 12px;'>"
        "<table role='presentation' width='640' cellspacing='0' cellpadding='0' "
        "style='max-width:640px;width:100%;background:#ffffff;border:1px solid #e5e7eb;border-radius:14px;overflow:hidden;'>"
        "<tr><td style='padding:22px 26px;background:#0f172a;color:#f8fafc;'>"
        f"{logo_block}"
        "<div style='font-size:20px;font-weight:700;'>Pedido visto por administracion</div>"
        "<div style='margin-top:6px;font-size:14px;color:#cbd5e1;'>Tu pedido ya esta en preparacion.</div>"
        "</td></tr>"
        "<tr><td style='padding:24px 26px;'>"
        f"<p style='margin:0 0 12px 0;font-size:15px;'>Hola {customer_name},</p>"
        "<p style='margin:0 0 12px 0;font-size:15px;'>"
        "tu pedido ya fue visto por el panel de administracion."
        "</p>"
        "<table role='presentation' width='100%' cellspacing='0' cellpadding='0' "
        "style='border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;margin-bottom:18px;'>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Pedido</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:600;'>#{order_id}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Fecha del pedido</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:600;'>{created_at_text}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Fecha de revision</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:600;'>{seen_at_text}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Total</td><td style='padding:10px 12px;text-align:right;font-size:15px;color:#0f172a;font-weight:700;'>{total_text}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Email</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:600;'>{email_text}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Entrega</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:600;'>{delivery_address}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Forma de pago</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:700;'>{payment_method_text}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Estado del pago</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:700;'>{payment_status_badge}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Referencia</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:600;'>{payment_ref_text}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Entrega programada</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:700;'>{delivery_window}</td></tr>"
        "</table>"
        "<div style='font-size:14px;font-weight:700;color:#0f172a;margin-bottom:8px;'>Productos solicitados</div>"
        "<table role='presentation' width='100%' cellspacing='0' cellpadding='0' "
        "style='border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;'>"
        "<thead><tr>"
        "<th align='left' style='padding:10px 12px;background:#f8fafc;font-size:12px;color:#475569;text-transform:uppercase;'>#</th>"
        "<th align='left' style='padding:10px 12px;background:#f8fafc;font-size:12px;color:#475569;text-transform:uppercase;'>Producto</th>"
        "<th align='right' style='padding:10px 12px;background:#f8fafc;font-size:12px;color:#475569;text-transform:uppercase;'>Cantidad</th>"
        "</tr></thead>"
        f"<tbody>{''.join(rows)}</tbody>"
        "</table>"
        "<p style='margin:18px 0 12px 0;font-size:15px;'>Gracias por tu compra.</p>"
        f"{tracking_button}"
        f"{support_line}"
        "</td></tr>"
        "<tr><td style='padding:12px 26px;border-top:1px solid #e5e7eb;color:#64748b;font-size:12px;'>"
        f"{footer_text} - Actualizacion automatica de pedido"
        "</td></tr>"
        "</table>"
        "</td></tr></table>"
        "</body></html>"
    )


def _build_order_prepared_html(order_data: Dict[str, Any]) -> str:
    brand_name = html_escape(str(os.environ.get('RESEND_BRAND_NAME') or 'DistriAr'))
    logo_url = (os.environ.get('RESEND_BRAND_LOGO_URL') or '').strip()
    order_id = html_escape(str(order_data.get('id') or 'sin-id'))
    customer_name = html_escape(str(order_data.get('user_full_name') or 'cliente'))
    delivery_window = html_escape(_format_delivery_schedule_label(order_data))

    prepared_at_text = ''
    try:
        timezone_name = str(os.environ.get('ORDER_EMAIL_TIMEZONE') or 'America/Argentina/Buenos_Aires').strip()
        now_utc = datetime.datetime.now(datetime.timezone.utc)
        prepared_dt = now_utc
        if timezone_name:
            try:
                from zoneinfo import ZoneInfo
                prepared_dt = now_utc.astimezone(ZoneInfo(timezone_name))
            except Exception:
                prepared_dt = now_utc
        prepared_at_text = prepared_dt.strftime('%d/%m/%Y %H:%M')
    except Exception:
        try:
            prepared_at_text = datetime.datetime.utcnow().strftime('%d/%m/%Y %H:%M')
        except Exception:
            prepared_at_text = ''
    prepared_at_text = html_escape(prepared_at_text) if prepared_at_text else '-'

    total_raw = order_data.get('total')
    total_text = _format_ars_money(total_raw)

    items = order_data.get('items') if isinstance(order_data.get('items'), list) else []
    rows: List[str] = []
    for idx, item in enumerate(items[:25], start=1):
        if isinstance(item, dict):
            prod_id_raw = item.get('id', '?')
            qty_raw = item.get('qty', '?')
            meta = item.get('meta') if isinstance(item.get('meta'), dict) else {}
            item_name = meta.get('name') or meta.get('title') or meta.get('product_name')
            qty_label = meta.get('qty_label')
        else:
            prod_id_raw = getattr(item, 'id', '?')
            qty_raw = getattr(item, 'qty', '?')
            meta = getattr(item, 'meta', None)
            meta = meta if isinstance(meta, dict) else {}
            item_name = meta.get('name') or meta.get('title') or meta.get('product_name')
            qty_label = meta.get('qty_label')

        try:
            qty_text = f"{float(qty_raw):g}"
        except Exception:
            qty_text = str(qty_raw)
        if qty_label:
            qty_text = f"{qty_text} ({str(qty_label)})"

        if not item_name:
            item_name = f"Producto #{prod_id_raw}"

        rows.append(
            "<tr>"
            f"<td style='padding:10px 12px;border-bottom:1px solid #edf1f5;color:#1f2937'>{idx}</td>"
            f"<td style='padding:10px 12px;border-bottom:1px solid #edf1f5;color:#1f2937'>{html_escape(str(item_name))}</td>"
            f"<td style='padding:10px 12px;border-bottom:1px solid #edf1f5;color:#334155;text-align:right'>{html_escape(qty_text)}</td>"
            "</tr>"
        )
    if not rows:
        rows.append(
            "<tr>"
            "<td colspan='3' style='padding:12px;color:#64748b;border-bottom:1px solid #edf1f5'>Sin detalle de productos</td>"
            "</tr>"
        )
    if len(items) > 25:
        rows.append(
            "<tr>"
            "<td colspan='3' style='padding:10px 12px;color:#64748b'>... y mas productos</td>"
            "</tr>"
        )

    support_email = html_escape(str(os.environ.get('RESEND_REPLY_TO') or ''))
    support_line = (
        f"<p style='margin:16px 0 0 0;font-size:13px;color:#64748b;'>"
        f"Si tenes dudas, respondenos a {support_email}."
        f"</p>"
        if support_email else
        "<p style='margin:16px 0 0 0;font-size:13px;color:#64748b;'>"
        "Si tenes dudas, respondenos a este correo."
        "</p>"
    )

    logo_block = (
        f"<img src='{html_escape(logo_url)}' alt='{brand_name}' style='height:42px;display:block;margin:0 0 12px 0;'/>"
        if logo_url else ""
    )
    tracking_button = _build_tracking_button_html(order_data)
    footer_text = html_escape(_build_brand_footer_text(str(os.environ.get('RESEND_BRAND_NAME') or 'DistriAr')))

    return (
        "<!doctype html>"
        "<html><body style='margin:0;padding:0;background:#f3f5f7;font-family:Arial,sans-serif;color:#0f172a;'>"
        "<table role='presentation' width='100%' cellspacing='0' cellpadding='0' style='background:#f3f5f7;'>"
        "<tr><td align='center' style='padding:24px 12px;'>"
        "<table role='presentation' width='640' cellspacing='0' cellpadding='0' "
        "style='max-width:640px;width:100%;background:#ffffff;border:1px solid #e5e7eb;border-radius:14px;overflow:hidden;'>"
        "<tr><td style='padding:22px 26px;background:#0f172a;color:#f8fafc;'>"
        f"{logo_block}"
        "<div style='font-size:20px;font-weight:700;'>Pedido preparado</div>"
        "<div style='margin-top:6px;font-size:14px;color:#cbd5e1;'>Tu pedido ya fue preparado por administracion.</div>"
        "</td></tr>"
        "<tr><td style='padding:24px 26px;'>"
        f"<p style='margin:0 0 12px 0;font-size:15px;'>Hola {customer_name},</p>"
        "<p style='margin:0 0 14px 0;font-size:15px;'>"
        "tu pedido ya fue preparado y llegara segun esta programacion:"
        "</p>"
        "<table role='presentation' width='100%' cellspacing='0' cellpadding='0' "
        "style='border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;margin-bottom:16px;'>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Pedido</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:600;'>#{order_id}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Preparado el</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:600;'>{prepared_at_text}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Entrega programada</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:700;'>{delivery_window}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Total</td><td style='padding:10px 12px;text-align:right;font-size:15px;color:#0f172a;font-weight:700;'>{total_text}</td></tr>"
        "</table>"
        "<div style='font-size:14px;font-weight:700;color:#0f172a;margin:8px 0 8px 0;'>Detalle del pedido</div>"
        "<table role='presentation' width='100%' cellspacing='0' cellpadding='0' "
        "style='border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;'>"
        "<thead><tr>"
        "<th align='left' style='padding:10px 12px;background:#f8fafc;font-size:12px;color:#475569;text-transform:uppercase;'>#</th>"
        "<th align='left' style='padding:10px 12px;background:#f8fafc;font-size:12px;color:#475569;text-transform:uppercase;'>Producto</th>"
        "<th align='right' style='padding:10px 12px;background:#f8fafc;font-size:12px;color:#475569;text-transform:uppercase;'>Cantidad</th>"
        "</tr></thead>"
        f"<tbody>{''.join(rows)}</tbody>"
        "</table>"
        "<p style='margin:18px 0 0 0;font-size:15px;'>Gracias por tu compra.</p>"
        f"{tracking_button}"
        f"{support_line}"
        "</td></tr>"
        "<tr><td style='padding:12px 26px;border-top:1px solid #e5e7eb;color:#64748b;font-size:12px;'>"
        f"{footer_text} - Actualizacion automatica de pedido"
        "</td></tr>"
        "</table>"
        "</td></tr></table>"
        "</body></html>"
    )


def _build_order_confirmation_html(order_data: Dict[str, Any]) -> str:
    brand_name = html_escape(str(os.environ.get('RESEND_BRAND_NAME') or 'DistriAr'))
    logo_url = (os.environ.get('RESEND_BRAND_LOGO_URL') or '').strip()

    order_id_raw = order_data.get('id')
    order_id = html_escape(str(order_id_raw if order_id_raw is not None else 'sin-id'))

    created_at_raw = order_data.get('created_at')
    created_at_text = ''
    if created_at_raw:
        try:
            parsed_dt = datetime.datetime.fromisoformat(str(created_at_raw).replace('Z', '+00:00'))
            created_at_text = parsed_dt.strftime('%d/%m/%Y %H:%M')
        except Exception:
            created_at_text = str(created_at_raw)
    created_at_text = html_escape(created_at_text) if created_at_text else ''

    total_raw = order_data.get('total')
    total_text = _format_ars_money(total_raw)

    customer_name = order_data.get('user_full_name')
    if not customer_name:
        token_preview = order_data.get('_token_preview')
        if isinstance(token_preview, dict):
            customer_name = token_preview.get('name')
    customer_name = html_escape(str(customer_name or 'cliente'))

    email_val = _normalize_email(order_data.get('user_email'))
    email_text = html_escape(email_val) if email_val else '-'
    payment_ui = _order_payment_ui(order_data)
    payment_method_text = html_escape(payment_ui.get('method_label') or 'No especificado')
    payment_status_text = html_escape(payment_ui.get('status_label') or 'Pendiente')
    payment_status_badge = (
        "<span style='display:inline-block;padding:4px 10px;border-radius:999px;"
        f"background:{payment_ui.get('status_bg') or '#fff7ed'};"
        f"color:{payment_ui.get('status_color') or '#9a3412'};"
        "font-weight:700;font-size:12px;'>"
        f"{payment_status_text}</span>"
    )
    payment_ref_text = html_escape(payment_ui.get('reference') or '-') if payment_ui.get('reference') else '-'

    address_parts: List[str] = []
    calle = order_data.get('user_calle')
    numeracion = order_data.get('user_numeracion')
    barrio = order_data.get('user_barrio')
    if calle:
        address_parts.append(str(calle).strip())
    if numeracion:
        address_parts.append(str(numeracion).strip())
    if barrio:
        address_parts.append(f"Barrio {str(barrio).strip()}")
    delivery_address = html_escape(', '.join([p for p in address_parts if p])) if address_parts else '-'
    delivery_window = html_escape(_format_delivery_schedule_label(order_data))

    items = order_data.get('items') if isinstance(order_data.get('items'), list) else []
    rows: List[str] = []
    for idx, item in enumerate(items[:25], start=1):
        if isinstance(item, dict):
            prod_id_raw = item.get('id', '?')
            qty_raw = item.get('qty', '?')
            meta = item.get('meta') if isinstance(item.get('meta'), dict) else {}
            item_name = meta.get('name') or meta.get('title') or meta.get('product_name')
            qty_label = meta.get('qty_label')
        else:
            prod_id_raw = getattr(item, 'id', '?')
            qty_raw = getattr(item, 'qty', '?')
            meta = getattr(item, 'meta', None)
            meta = meta if isinstance(meta, dict) else {}
            item_name = meta.get('name') or meta.get('title') or meta.get('product_name')
            qty_label = meta.get('qty_label')

        try:
            qty_text = f"{float(qty_raw):g}"
        except Exception:
            qty_text = str(qty_raw)
        if qty_label:
            qty_text = f"{qty_text} ({str(qty_label)})"

        if not item_name:
            item_name = f"Producto #{prod_id_raw}"

        rows.append(
            "<tr>"
            f"<td style='padding:10px 12px;border-bottom:1px solid #edf1f5;color:#1f2937'>{idx}</td>"
            f"<td style='padding:10px 12px;border-bottom:1px solid #edf1f5;color:#1f2937'>{html_escape(str(item_name))}</td>"
            f"<td style='padding:10px 12px;border-bottom:1px solid #edf1f5;color:#334155;text-align:right'>{html_escape(qty_text)}</td>"
            "</tr>"
        )
    if not rows:
        rows.append(
            "<tr>"
            "<td colspan='3' style='padding:12px;color:#64748b;border-bottom:1px solid #edf1f5'>Sin detalle de productos</td>"
            "</tr>"
        )
    if len(items) > 25:
        rows.append(
            "<tr>"
            "<td colspan='3' style='padding:10px 12px;color:#64748b'>... y mas productos</td>"
            "</tr>"
        )

    logo_block = (
        f"<img src='{html_escape(logo_url)}' alt='{brand_name}' style='height:42px;display:block;margin:0 0 12px 0;'/>"
        if logo_url else ""
    )
    tracking_button = _build_tracking_button_html(order_data)
    footer_text = html_escape(_build_brand_footer_text(str(os.environ.get('RESEND_BRAND_NAME') or 'DistriAr')))

    return (
        "<!doctype html>"
        "<html><body style='margin:0;padding:0;background:#f3f5f7;font-family:Arial,sans-serif;color:#0f172a;'>"
        "<table role='presentation' width='100%' cellspacing='0' cellpadding='0' style='background:#f3f5f7;'>"
        "<tr><td align='center' style='padding:24px 12px;'>"
        "<table role='presentation' width='640' cellspacing='0' cellpadding='0' "
        "style='max-width:640px;width:100%;background:#ffffff;border:1px solid #e5e7eb;border-radius:14px;overflow:hidden;'>"
        "<tr><td style='padding:24px 28px;background:#0f172a;color:#f8fafc;'>"
        f"{logo_block}"
        "<div style='font-size:22px;font-weight:700;line-height:1.2;'>Pedido recibido</div>"
        "<div style='margin-top:6px;font-size:14px;color:#cbd5e1;'>Estamos preparando tu pedido.</div>"
        "</td></tr>"
        "<tr><td style='padding:24px 28px;'>"
        f"<p style='margin:0 0 14px 0;font-size:15px;'>Hola {customer_name}, gracias por tu compra.</p>"
        "<table role='presentation' width='100%' cellspacing='0' cellpadding='0' "
        "style='border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;margin-bottom:18px;'>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Pedido</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:600;'>#{order_id}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Fecha</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:600;'>{created_at_text or '-'}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Total</td><td style='padding:10px 12px;text-align:right;font-size:15px;color:#0f172a;font-weight:700;'>{total_text}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Email</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:600;'>{email_text}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Entrega</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:600;'>{delivery_address}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Entrega programada</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:700;'>{delivery_window}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Forma de pago</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:700;'>{payment_method_text}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Estado del pago</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:700;'>{payment_status_badge}</td></tr>"
        f"<tr><td style='padding:10px 12px;background:#f8fafc;color:#334155;font-size:13px;'>Referencia</td><td style='padding:10px 12px;text-align:right;font-size:13px;color:#0f172a;font-weight:600;'>{payment_ref_text}</td></tr>"
        "</table>"
        "<div style='font-size:14px;font-weight:700;color:#0f172a;margin-bottom:8px;'>Detalle del pedido</div>"
        "<table role='presentation' width='100%' cellspacing='0' cellpadding='0' "
        "style='border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;'>"
        "<thead><tr>"
        "<th align='left' style='padding:10px 12px;background:#f8fafc;font-size:12px;color:#475569;text-transform:uppercase;'>#</th>"
        "<th align='left' style='padding:10px 12px;background:#f8fafc;font-size:12px;color:#475569;text-transform:uppercase;'>Producto</th>"
        "<th align='right' style='padding:10px 12px;background:#f8fafc;font-size:12px;color:#475569;text-transform:uppercase;'>Cantidad</th>"
        "</tr></thead>"
        f"<tbody>{''.join(rows)}</tbody>"
        "</table>"
        "<p style='margin:18px 0 0 0;font-size:12px;color:#64748b;'>Si necesitas ayuda, responde este correo y te asistimos.</p>"
        f"{tracking_button}"
        "</td></tr>"
        "<tr><td style='padding:14px 28px;border-top:1px solid #e5e7eb;color:#64748b;font-size:12px;'>"
        f"{footer_text} - Confirmacion automatica de pedido"
        "</td></tr>"
        "</table>"
        "</td></tr></table>"
        "</body></html>"
    )


async def _send_order_confirmation_email(
    order_data: Optional[Dict[str, Any]],
    request_data: Optional[Dict[str, Any]] = None,
    token_payload: Optional[Dict[str, Any]] = None,
) -> bool:
    def _masked_key(v: str) -> str:
        if not v:
            return ''
        if len(v) <= 8:
            return '*' * len(v)
        return f"{v[:4]}...{v[-4:]}"

    # Configure in environment:
    # RESEND_API_KEY=re_xxxxxxxxx  (replace re_xxxxxxxxx with your real API key)
    api_key = (os.environ.get('RESEND_API_KEY') or '').strip()
    if not api_key:
        logger.warning('RESEND_API_KEY is not configured. Skipping confirmation email.')
        return False
    if api_key == 're_xxxxxxxxx':
        logger.warning("RESEND_API_KEY is still 're_xxxxxxxxx'. Replace it with your real API key.")
        return False

    enabled_raw = (os.environ.get('RESEND_ORDER_CONFIRMATION_ENABLED') or 'true').strip().lower()
    if enabled_raw in ('0', 'false', 'no', 'off'):
        logger.info('RESEND_ORDER_CONFIRMATION_ENABLED disabled. Skipping confirmation email.')
        return False

    order_payload = _enrich_order_contact_fields(order_data if isinstance(order_data, dict) else {})
    to_email = _extract_customer_email_for_order(order_payload, request_data, token_payload)
    force_to_email = _normalize_email(os.environ.get('RESEND_FORCE_TO_EMAIL'))
    if force_to_email:
        logger.info(
            'RESEND_FORCE_TO_EMAIL active: overriding recipient %s -> %s for order id=%s',
            to_email,
            force_to_email,
            order_payload.get('id'),
        )
        to_email = force_to_email
    if not to_email:
        logger.info(
            'Skipping Resend confirmation: no customer email found for order id=%s',
            order_payload.get('id'),
        )
        return False

    from_email = (os.environ.get('RESEND_FROM_EMAIL') or 'onboarding@resend.dev').strip()
    logger.info(
        'Resend send attempt order id=%s to=%s from=%s key=%s',
        order_payload.get('id'),
        to_email,
        from_email,
        _masked_key(api_key),
    )
    send_payload = {
        'from': from_email,
        'to': [to_email],
        'subject': _build_order_confirmation_subject(order_payload),
        'html': _build_order_confirmation_html(order_payload),
    }
    reply_to = _normalize_email(os.environ.get('RESEND_REPLY_TO'))
    if reply_to:
        send_payload['reply_to'] = reply_to

    try:
        timeout_seconds = float(os.environ.get('RESEND_TIMEOUT_SECONDS') or '12')
    except Exception:
        timeout_seconds = 12.0

    try:
        async with httpx.AsyncClient(timeout=timeout_seconds) as client:
            response = await client.post(
                'https://api.resend.com/emails',
                headers={
                    'Authorization': f'Bearer {api_key}',
                    'Content-Type': 'application/json',
                },
                json=send_payload,
            )
    except Exception as e:
        logger.warning('Resend request failed for order id=%s: %s', order_payload.get('id'), e)
        return False

    if response.status_code >= 400:
        try:
            body_preview = response.text[:300]
        except Exception:
            body_preview = '<unavailable>'
        logger.warning(
            'Resend send failed for order id=%s status=%s body=%s',
            order_payload.get('id'),
            response.status_code,
            body_preview,
        )
        return False

    resend_id = None
    try:
        resend_id = (response.json() or {}).get('id')
    except Exception:
        resend_id = None

    logger.info(
        'Resend confirmation sent for order id=%s to=%s resend_id=%s',
        order_payload.get('id'),
        to_email,
        resend_id,
    )
    return True


async def _send_order_seen_notification_email(order_data: Optional[Dict[str, Any]]) -> bool:
    def _masked_key(v: str) -> str:
        if not v:
            return ''
        if len(v) <= 8:
            return '*' * len(v)
        return f"{v[:4]}...{v[-4:]}"

    enabled_raw = (os.environ.get('RESEND_ORDER_SEEN_EMAIL_ENABLED') or 'true').strip().lower()
    if enabled_raw in ('0', 'false', 'no', 'off'):
        logger.info('RESEND_ORDER_SEEN_EMAIL_ENABLED disabled. Skipping seen-notification email.')
        return False

    api_key = (os.environ.get('RESEND_API_KEY') or '').strip()
    if not api_key:
        logger.warning('RESEND_API_KEY is not configured. Skipping seen-notification email.')
        return False
    if api_key == 're_xxxxxxxxx':
        logger.warning("RESEND_API_KEY is still 're_xxxxxxxxx'. Replace it with your real API key.")
        return False

    order_payload = _enrich_order_contact_fields(order_data if isinstance(order_data, dict) else {})
    to_email = _extract_customer_email_for_order(order_payload, order_payload, None)
    force_to_email = _normalize_email(os.environ.get('RESEND_FORCE_TO_EMAIL'))
    if force_to_email:
        logger.info(
            'RESEND_FORCE_TO_EMAIL active: overriding recipient %s -> %s for seen-notification order id=%s',
            to_email,
            force_to_email,
            order_payload.get('id'),
        )
        to_email = force_to_email
    if not to_email:
        logger.info(
            'Skipping seen-notification email: no customer email found for order id=%s',
            order_payload.get('id'),
        )
        return False

    from_email = (os.environ.get('RESEND_FROM_EMAIL') or 'onboarding@resend.dev').strip()
    logger.info(
        'Resend seen-notification send attempt order id=%s to=%s from=%s key=%s',
        order_payload.get('id'),
        to_email,
        from_email,
        _masked_key(api_key),
    )

    send_payload = {
        'from': from_email,
        'to': [to_email],
        'subject': _build_order_seen_subject(order_payload),
        'html': _build_order_seen_html(order_payload),
    }
    reply_to = _normalize_email(os.environ.get('RESEND_REPLY_TO'))
    if reply_to:
        send_payload['reply_to'] = reply_to

    try:
        timeout_seconds = float(os.environ.get('RESEND_TIMEOUT_SECONDS') or '12')
    except Exception:
        timeout_seconds = 12.0

    try:
        async with httpx.AsyncClient(timeout=timeout_seconds) as client:
            response = await client.post(
                'https://api.resend.com/emails',
                headers={
                    'Authorization': f'Bearer {api_key}',
                    'Content-Type': 'application/json',
                },
                json=send_payload,
            )
    except Exception as e:
        logger.warning('Resend request failed for seen-notification order id=%s: %s', order_payload.get('id'), e)
        return False

    if response.status_code >= 400:
        try:
            body_preview = response.text[:300]
        except Exception:
            body_preview = '<unavailable>'
        logger.warning(
            'Resend seen-notification failed for order id=%s status=%s body=%s',
            order_payload.get('id'),
            response.status_code,
            body_preview,
        )
        return False

    resend_id = None
    try:
        resend_id = (response.json() or {}).get('id')
    except Exception:
        resend_id = None

    logger.info(
        'Resend seen-notification sent for order id=%s to=%s resend_id=%s',
        order_payload.get('id'),
        to_email,
        resend_id,
    )
    return True


async def _send_order_prepared_notification_email(order_data: Optional[Dict[str, Any]]) -> bool:
    def _masked_key(v: str) -> str:
        if not v:
            return ''
        if len(v) <= 8:
            return '*' * len(v)
        return f"{v[:4]}...{v[-4:]}"

    enabled_raw = (os.environ.get('RESEND_ORDER_PREPARED_EMAIL_ENABLED') or 'true').strip().lower()
    if enabled_raw in ('0', 'false', 'no', 'off'):
        logger.info('RESEND_ORDER_PREPARED_EMAIL_ENABLED disabled. Skipping prepared-notification email.')
        return False

    api_key = (os.environ.get('RESEND_API_KEY') or '').strip()
    if not api_key:
        logger.warning('RESEND_API_KEY is not configured. Skipping prepared-notification email.')
        return False
    if api_key == 're_xxxxxxxxx':
        logger.warning("RESEND_API_KEY is still 're_xxxxxxxxx'. Replace it with your real API key.")
        return False

    order_payload = _enrich_order_contact_fields(order_data if isinstance(order_data, dict) else {})
    to_email = _extract_customer_email_for_order(order_payload, order_payload, None)
    force_to_email = _normalize_email(os.environ.get('RESEND_FORCE_TO_EMAIL'))
    if force_to_email:
        logger.info(
            'RESEND_FORCE_TO_EMAIL active: overriding recipient %s -> %s for prepared-notification order id=%s',
            to_email,
            force_to_email,
            order_payload.get('id'),
        )
        to_email = force_to_email
    if not to_email:
        logger.info(
            'Skipping prepared-notification email: no customer email found for order id=%s',
            order_payload.get('id'),
        )
        return False

    from_email = (os.environ.get('RESEND_FROM_EMAIL') or 'onboarding@resend.dev').strip()
    logger.info(
        'Resend prepared-notification send attempt order id=%s to=%s from=%s key=%s',
        order_payload.get('id'),
        to_email,
        from_email,
        _masked_key(api_key),
    )

    send_payload = {
        'from': from_email,
        'to': [to_email],
        'subject': _build_order_prepared_subject(order_payload),
        'html': _build_order_prepared_html(order_payload),
    }
    reply_to = _normalize_email(os.environ.get('RESEND_REPLY_TO'))
    if reply_to:
        send_payload['reply_to'] = reply_to

    try:
        timeout_seconds = float(os.environ.get('RESEND_TIMEOUT_SECONDS') or '12')
    except Exception:
        timeout_seconds = 12.0

    try:
        async with httpx.AsyncClient(timeout=timeout_seconds) as client:
            response = await client.post(
                'https://api.resend.com/emails',
                headers={
                    'Authorization': f'Bearer {api_key}',
                    'Content-Type': 'application/json',
                },
                json=send_payload,
            )
    except Exception as e:
        logger.warning('Resend request failed for prepared-notification order id=%s: %s', order_payload.get('id'), e)
        return False

    if response.status_code >= 400:
        try:
            body_preview = response.text[:300]
        except Exception:
            body_preview = '<unavailable>'
        logger.warning(
            'Resend prepared-notification failed for order id=%s status=%s body=%s',
            order_payload.get('id'),
            response.status_code,
            body_preview,
        )
        return False

    resend_id = None
    try:
        resend_id = (response.json() or {}).get('id')
    except Exception:
        resend_id = None

    logger.info(
        'Resend prepared-notification sent for order id=%s to=%s resend_id=%s',
        order_payload.get('id'),
        to_email,
        resend_id,
    )
    return True


def _resend_status_snapshot() -> Dict[str, Any]:
    api_key = (os.environ.get('RESEND_API_KEY') or '').strip()
    from_email = (os.environ.get('RESEND_FROM_EMAIL') or 'onboarding@resend.dev').strip()
    reply_to = _normalize_email(os.environ.get('RESEND_REPLY_TO'))
    enabled_raw = (os.environ.get('RESEND_ORDER_CONFIRMATION_ENABLED') or 'true').strip().lower()
    enabled = enabled_raw not in ('0', 'false', 'no', 'off')
    seen_enabled_raw = (os.environ.get('RESEND_ORDER_SEEN_EMAIL_ENABLED') or 'true').strip().lower()
    seen_enabled = seen_enabled_raw not in ('0', 'false', 'no', 'off')
    prepared_enabled_raw = (os.environ.get('RESEND_ORDER_PREPARED_EMAIL_ENABLED') or 'true').strip().lower()
    prepared_enabled = prepared_enabled_raw not in ('0', 'false', 'no', 'off')
    return {
        'enabled': enabled,
        'seen_enabled': seen_enabled,
        'prepared_enabled': prepared_enabled,
        'api_key_present': bool(api_key),
        'api_key_placeholder': api_key == 're_xxxxxxxxx',
        'from_email': from_email,
        'reply_to': reply_to,
        'force_to_email': _normalize_email(os.environ.get('RESEND_FORCE_TO_EMAIL')),
        'require_customer_email': (os.environ.get('ORDER_REQUIRE_CUSTOMER_EMAIL') or 'true').strip().lower() not in ('0', 'false', 'no', 'off'),
    }

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/auth/token")
admin_oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/admin/auth/token")

# -------------------------------------------------------------------
# MIDDLEWARES
# -------------------------------------------------------------------
# Configure CORS origins via env var CORS_ALLOWED_ORIGINS (comma-separated). When credentials are required
# we cannot use wildcard '*' for Access-Control-Allow-Origin, so prefer explicit origins.
CORS_ALLOWED_ORIGINS = [o.strip() for o in (os.environ.get('CORS_ALLOWED_ORIGINS') or '').split(',') if o.strip()]
if CORS_ALLOWED_ORIGINS:
    cors_allow_origins = CORS_ALLOWED_ORIGINS
    cors_allow_credentials = True
else:
    # No explicit allowed origins configured: allow all origins but disable credentials to avoid
    # browser rejection when Access-Control-Allow-Origin == '*'. This is a safe default.
    cors_allow_origins = ['*']
    cors_allow_credentials = False

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_allow_origins,
    allow_credentials=cors_allow_credentials,
    allow_methods=["*"],
    # Explicitly allow Authorization and common headers instead of wildcard so browsers' preflights
    # reliably accept the Authorization header across origins.
    allow_headers=["Authorization", "Content-Type", "Accept", "Origin", "X-Requested-With"],
)


def _cors_headers_for_request(request: Request) -> Dict[str, str]:
    """Return CORS response headers appropriate for the incoming request.
    If the request includes an Origin header, echo it back (required when credentials are included).
    If the browser made a preflight (OPTIONS) request with Access-Control-Request-Headers, echo
    that back so custom headers like Authorization are explicitly permitted.
    """
    origin = request.headers.get('origin')
    headers: Dict[str, str] = {
        'Access-Control-Allow-Credentials': 'true' if cors_allow_credentials else 'false',
        'Access-Control-Allow-Methods': '*',
    }
    # Respect the browser's requested headers if provided (preflight); otherwise advertise common safe set
    acrh = request.headers.get('access-control-request-headers')
    if acrh:
        headers['Access-Control-Allow-Headers'] = acrh
    else:
        headers['Access-Control-Allow-Headers'] = 'Authorization, Content-Type, Accept, Origin, X-Requested-With'

    if origin:
        headers['Access-Control-Allow-Origin'] = origin
    else:
        headers['Access-Control-Allow-Origin'] = '*'
    return headers


@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    # Log validation errors and request body to server_log.txt to help diagnose 422 issues in production
    try:
        try:
            body_bytes = await request.body()
            try:
                body = body_bytes.decode('utf-8')
            except Exception:
                body = str(body_bytes)
        except Exception:
            body = '<could not read body>'
        tb = traceback.format_exc()
        try:
            base = os.path.dirname(os.path.dirname(__file__))
            logpath = os.path.join(base, 'server_log.txt')
            with open(logpath, 'a', encoding='utf-8') as f:
                f.write(f"{datetime.datetime.utcnow().isoformat()} - RequestValidationError path={request.url.path} body={body[:2000]} error={str(exc)[:500]}\n")
                f.write(tb + "\n\n")
        except Exception:
            pass
            logger.exception('RequestValidationError for %s: %s body=%s', request.url.path, exc, (body[:500] if body else ''))
    except Exception:
        logger.exception('validation_exception_handler failed')

    # Fallback: for admin product create/update requests, try to coerce payload and perform the DB operation directly
    try:
        m = (request.method or '').upper()
        p = request.url.path or ''
        if p.startswith('/products') and m in ('POST', 'PUT'):
            try:
                body_json = None
                try:
                    body_json = await request.json()
                except Exception:
                    body_json = None
                if body_json and isinstance(body_json, dict):
                    # Coerce common product fields
                    name = str(body_json.get('name') or '').strip()
                    price_raw = body_json.get('price')
                    try:
                        price = float(price_raw)
                    except Exception:
                        price = None
                    try:
                        price_retail = float(body_json.get('price_retail')) if body_json.get('price_retail') is not None else None
                    except Exception:
                        price_retail = None
                    if m == 'POST':
                        if name and price is not None:
                            payload_obj = SimpleNamespace(
                                code=(str(body_json.get('code') or body_json.get('codigo') or '').strip() or None),
                                name=name,
                                price=price,
                                price_retail=price_retail,
                                description=body_json.get('description') or '',
                                category=body_json.get('category') or '',
                                image_url=body_json.get('image_url') or '',
                                active=bool(body_json.get('active', True)),
                                stock=int(body_json.get('stock') or 0),
                                stock_kg=float(body_json.get('stock_kg') or body_json.get('stock') or 0.0),
                                kg_per_unit=float(body_json.get('kg_per_unit') or 1.0),
                                discount=float(body_json.get('discount') or 0.0),
                                sale_unit=str(body_json.get('sale_unit') or 'unit')
                            )
                            db = SessionLocal()
                            try:
                                created = crud.create_product(db, payload_obj)
                                if created:
                                    # Normalize to dict
                                    if isinstance(created, dict):
                                        res = created
                                    else:
                                        res = {k: getattr(created, k) for k in ('id','code','name','price','price_retail','description','category','image_url','active','stock','stock_kg','kg_per_unit','discount','sale_unit') if hasattr(created, k)}
                                    # log the fallback usage
                                    try:
                                        base = os.path.dirname(os.path.dirname(__file__))
                                        with open(os.path.join(base, 'server_log.txt'), 'a', encoding='utf-8') as f:
                                            f.write(f"{datetime.datetime.utcnow().isoformat()} - Validation fallback POST /products used for body: {str(body)[:1000]} created_id={res.get('id')}\n")
                                    except Exception:
                                        pass
                                    headers = _cors_headers_for_request(request)
                                    return JSONResponse(status_code=200, content=res, headers=headers)
                            finally:
                                try:
                                    db.close()
                                except Exception:
                                    pass
                    else:
                        # PUT fallback: extract id from path and perform permissive update
                        try:
                            parts = [x for x in p.split('/') if x]
                            # path may be /products or /products/123; ensure last part is id
                            if len(parts) >= 2 and parts[-2] == 'products':
                                pid = parts[-1]
                            elif len(parts) >= 1 and parts[0] == 'products' and len(parts) == 2:
                                pid = parts[1]
                            else:
                                pid = None
                        except Exception:
                            pid = None
                        if pid is not None:
                            # Build permissive updates dict
                            updates = {}
                            for k in ('code','name','price','price_retail','description','category','image_url','active','stock','stock_kg','kg_per_unit','discount','sale_unit'):
                                if k in body_json:
                                    updates[k] = body_json[k]
                            if 'price' in updates:
                                try:
                                    updates['price'] = float(updates['price'])
                                except Exception:
                                    updates.pop('price', None)
                            if 'price_retail' in updates:
                                try:
                                    updates['price_retail'] = float(updates['price_retail']) if updates['price_retail'] is not None else None
                                except Exception:
                                    updates.pop('price_retail', None)
                            if 'stock' in updates:
                                try:
                                    updates['stock'] = int(updates['stock'])
                                except Exception:
                                    updates.pop('stock', None)
                            if 'discount' in updates:
                                try:
                                    updates['discount'] = float(updates['discount'])
                                except Exception:
                                    updates.pop('discount', None)
                            if 'stock_kg' in updates:
                                try:
                                    updates['stock_kg'] = float(updates['stock_kg'])
                                except Exception:
                                    updates.pop('stock_kg', None)
                            if 'kg_per_unit' in updates:
                                try:
                                    updates['kg_per_unit'] = float(updates['kg_per_unit'])
                                    if updates['kg_per_unit'] <= 0:
                                        updates['kg_per_unit'] = 1.0
                                except Exception:
                                    updates.pop('kg_per_unit', None)
                            if 'sale_unit' in updates:
                                try:
                                    updates['sale_unit'] = str(updates['sale_unit'] or 'unit')
                                except Exception:
                                    updates.pop('sale_unit', None)
                            if 'code' in updates:
                                try:
                                    updates['code'] = str(updates['code']).strip() if updates['code'] is not None else None
                                    if updates['code'] == '':
                                        updates['code'] = None
                                except Exception:
                                    updates.pop('code', None)
                            if updates:
                                # Create a small payload object with dict() method used by crud.update_product
                                upd_dict = updates.copy()
                                class _UpdObj:
                                    def __init__(self, d):
                                        self.__dict__.update(d)
                                    def dict(self, exclude_unset=True):
                                        return d
                                d = upd_dict
                                upd_obj = _UpdObj(d)
                                db = SessionLocal()
                                try:
                                    updated = crud.update_product(db, int(pid), upd_obj)
                                    if updated:
                                        # Normalize to dict
                                        res = updated if isinstance(updated, dict) else {k: getattr(updated, k) for k in ('id','code','name','price','price_retail','description','category','image_url','active','stock','stock_kg','kg_per_unit','discount','sale_unit') if hasattr(updated,k)}
                                        try:
                                            base = os.path.dirname(os.path.dirname(__file__))
                                            with open(os.path.join(base, 'server_log.txt'), 'a', encoding='utf-8') as f:
                                                f.write(f"{datetime.datetime.utcnow().isoformat()} - Validation fallback PUT /products/{pid} used for body: {str(body)[:1000]}\n")
                                        except Exception:
                                            pass
                                        headers = _cors_headers_for_request(request)
                                        return JSONResponse(status_code=200, content=res, headers=headers)
                                finally:
                                    try:
                                        db.close()
                                    except Exception:
                                        pass
            except Exception:
                pass
    except Exception:
        pass

    headers = _cors_headers_for_request(request)
    return JSONResponse(status_code=422, content={"detail": exc.errors()}, headers=headers)


@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    logger.exception('Unhandled exception: %s', exc)
    headers = _cors_headers_for_request(request)
    return JSONResponse(status_code=500, content={"detail": "Internal Server Error"}, headers=headers)


# Lightweight health and debug endpoints
@app.get('/health')
def health():
    mp_configured = bool((os.environ.get('MERCADOPAGO_ACCESS_TOKEN') or '').strip())
    return {'status': 'ok', 'mercadopago_configured': mp_configured}

@app.get('/debug/version')
def debug_version():
    """Return deploy identification info to help confirm the running code version."""
    commit = os.environ.get('DEPLOY_COMMIT') or os.environ.get('RENDER_GIT_COMMIT') or os.environ.get('HEROKU_SLUG_COMMIT')
    return {
        'commit': commit or 'unknown',
        'python': sys.version.splitlines()[0]
    }


@app.post('/debug/echo')
async def debug_echo(request: Request):
    try:
        data = await request.json()
    except Exception:
        data = None
    logger.info('debug/echo called, origin=%s, data=%s', request.headers.get('origin'), data)
    headers = _cors_headers_for_request(request)
    return JSONResponse(status_code=200, content={"echo": data, "headers": dict(request.headers)}, headers=headers)


@app.post('/debug/clear')
async def debug_clear(request: Request):
    headers = _cors_headers_for_request(request)
    try:
        try:
            body = await request.json()
        except Exception:
            body = {}
        raw_target = body.get('target') if isinstance(body, dict) else None
        if raw_target is None and isinstance(body, dict):
            raw_target = body.get('code')
        target_key = str(raw_target or '').strip().lower()
        if target_key.startswith('@'):
            target_key = target_key[1:]

        aliases = {
            '1': 'catalog',
            'catalog': 'catalog',
            'catalogo': 'catalog',
            'productos': 'catalog',
            'products': 'catalog',
            '2': 'filters',
            'filters': 'filters',
            'filtros': 'filters',
            '3': 'orders',
            'orders': 'orders',
            'pedidos': 'orders',
            '4': 'customers',
            'customers': 'customers',
            'clientes': 'customers',
            '5': 'preparations',
            'preparations': 'preparations',
            'preparaciones': 'preparations',
        }
        target = aliases.get(target_key)
        if not target:
            return JSONResponse(status_code=400, content={'detail': 'invalid_target', 'target': target_key}, headers=headers)

        def _table_exists(name: str) -> bool:
            try:
                insp_local = inspect(engine)
                return name in (insp_local.get_table_names() or [])
            except Exception:
                return False

        def _delete_all(name: str) -> int:
            if not _table_exists(name):
                return 0
            try:
                with engine.begin() as conn:
                    res = conn.execute(text(f"DELETE FROM {name}"))
                    try:
                        return int(res.rowcount or 0)
                    except Exception:
                        return 0
            except Exception as e_del:
                logger.exception('debug/clear failed to delete %s: %s', name, e_del)
                return 0

        result = {'ok': True, 'target': target, 'deleted': {}, 'updated': {}}

        if target == 'catalog':
            result['deleted']['products'] = _delete_all('products')
            result['deleted']['product_changes'] = _delete_all('product_changes')
            try:
                PRODUCTS_CACHE.clear()
                PRODUCTS_COUNT_CACHE.clear()
            except Exception:
                pass

        elif target == 'filters':
            db = SessionLocal()
            try:
                try:
                    crud.set_setting(db, 'filters', [])
                except Exception:
                    logger.exception('debug/clear filters: failed to persist DB setting')
                try:
                    os.makedirs(CATALOG_DIR, exist_ok=True)
                    path = os.path.join(CATALOG_DIR, 'filters.json')
                    with open(path, 'w', encoding='utf-8') as f:
                        f.write(json.dumps([], ensure_ascii=False, indent=2))
                    result['updated']['filters_snapshot'] = 1
                except Exception:
                    logger.exception('debug/clear filters: failed to write snapshot file')
                try:
                    await push_event({'type': 'filters-updated', 'count': 0})
                except Exception:
                    pass
            finally:
                try:
                    db.close()
                except Exception:
                    pass

        elif target == 'orders':
            result['deleted']['orders'] = _delete_all('orders')
            result['deleted']['order_token_previews'] = _delete_all('order_token_previews')
            db = SessionLocal()
            try:
                try:
                    crud.set_setting(db, 'order_status_overrides', {})
                except Exception:
                    logger.exception('debug/clear orders: failed to reset status overrides')
            finally:
                try:
                    db.close()
                except Exception:
                    pass
            try:
                ORDER_PAYLOAD_CACHE.clear()
                ORDER_STATUS_CACHE.clear()
            except Exception:
                pass

        elif target == 'customers':
            result['deleted']['user_addresses'] = _delete_all('user_addresses')
            result['deleted']['users'] = _delete_all('users')
            result['deleted']['order_token_previews'] = _delete_all('order_token_previews')
            # Anonymize orders so customer lists derived from orders are cleared.
            try:
                if _table_exists('orders'):
                    insp_local = inspect(engine)
                    cols = {c['name'] for c in (insp_local.get_columns('orders') or [])}
                    to_null = [
                        'user_id', 'user_full_name', 'user_email', 'user_barrio', 'user_calle',
                        'user_numeracion', 'user_postal_code', 'user_department',
                        '_token_preview', '_token_received',
                    ]
                    existing = [c for c in to_null if c in cols]
                    if existing:
                        set_sql = ", ".join([f"{c} = NULL" for c in existing])
                        with engine.begin() as conn:
                            res = conn.execute(text(f"UPDATE orders SET {set_sql}"))
                            try:
                                result['updated']['orders_anonymized'] = int(res.rowcount or 0)
                            except Exception:
                                result['updated']['orders_anonymized'] = 0
            except Exception:
                logger.exception('debug/clear customers: failed to anonymize orders')
            try:
                ORDER_PAYLOAD_CACHE.clear()
                ORDER_STATUS_CACHE.clear()
            except Exception:
                pass

        elif target == 'preparations':
            try:
                if _table_exists('orders'):
                    insp_local = inspect(engine)
                    cols = {c['name'] for c in (insp_local.get_columns('orders') or [])}
                    if 'status' in cols:
                        with engine.begin() as conn:
                            res = conn.execute(
                                text("UPDATE orders SET status = :status WHERE LOWER(status) IN ('visto','preparado','preparando','preparing','prepared','seen','viewed')"),
                                {'status': 'entregado'}
                            )
                            try:
                                result['updated']['orders_status_updated'] = int(res.rowcount or 0)
                            except Exception:
                                result['updated']['orders_status_updated'] = 0
                db = SessionLocal()
                try:
                    try:
                        crud.set_setting(db, 'order_status_overrides', {})
                    except Exception:
                        logger.exception('debug/clear preparations: failed to reset status overrides')
                finally:
                    try:
                        db.close()
                    except Exception:
                        pass
                try:
                    ORDER_STATUS_CACHE.clear()
                except Exception:
                    pass
            except Exception:
                logger.exception('debug/clear preparations failed')

        return JSONResponse(status_code=200, content=result, headers=headers)
    except Exception as e:
        logger.exception('debug/clear failed: %s', e)
        return JSONResponse(status_code=500, content={'detail': 'clear_failed'}, headers=headers)


# -------------------------------------------------------------------
# Filters endpoint: serve and persist admin-managed filters so public catalog can fetch them
# -------------------------------------------------------------------
@app.get('/filters.json')
def get_filters(request: Request, db: Session = Depends(get_db)):
    """Return admin-managed filters. Prefer the DB-backed setting when available
    so filters survive deploys; fall back to the on-disk snapshot for compatibility."""
    data = None
    try:
        # Try DB first
        dbv = crud.get_setting(db, 'filters')
        if isinstance(dbv, list):
            data = dbv
        else:
            # fallback to file if DB has no value or it's malformed
            path = os.path.join(CATALOG_DIR, 'filters.json')
            if os.path.exists(path):
                with open(path, 'r', encoding='utf-8') as f:
                    loaded = json.load(f)
                data = loaded if isinstance(loaded, list) else []
            else:
                data = []
    except Exception as e:
        logger.exception('get_filters failed: %s', e)
        data = []
    headers = _cors_headers_for_request(request)
    return JSONResponse(status_code=200, content=data, headers=headers)


@app.post('/filters')
async def set_filters(request: Request, db: Session = Depends(get_db)):
    try:
        body = await request.json()
        if not isinstance(body, list):
            raise HTTPException(status_code=400, detail='filters must be an array')
        # Persist to DB (best-effort)
        try:
            ok = crud.set_setting(db, 'filters', body)
            if not ok:
                logger.warning('set_filters: failed to persist filters to DB')
        except Exception:
            logger.exception('set_filters: DB persist failed')
        # Also write file snapshot for backward compatibility
        try:
            os.makedirs(CATALOG_DIR, exist_ok=True)
            path = os.path.join(CATALOG_DIR, 'filters.json')
            with open(path, 'w', encoding='utf-8') as f:
                f.write(json.dumps(body, ensure_ascii=False, indent=2))
        except Exception:
            logger.exception('set_filters: failed to write snapshot file')
        headers = _cors_headers_for_request(request)
        # notify websocket listeners that filters changed
        try:
            await push_event({'type':'filters-updated','count': len(body)})
        except Exception:
            pass
        logger.info('Saved filters.json with %s entries', len(body))
        return JSONResponse(status_code=200, content={'ok': True, 'saved': len(body)}, headers=headers)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception('set_filters failed: %s', e)
        raise HTTPException(status_code=500, detail='Failed to save filters')


# -------------------------------------------------------------------
# Product categories: mappings productKey -> [filterValues]
# Stored in catalogo/product_categories.json for the public catalog to consume
# -------------------------------------------------------------------
@app.get('/product-categories.json')
def get_product_categories(request: Request, db: Session = Depends(get_db)):
    """Return mapping productKey -> [categoryValues]. Prefer DB-stored mapping so
    the configuration survives deploys; fall back to the file snapshot otherwise."""
    data = None
    try:
        dbv = crud.get_setting(db, 'product_categories')
        if isinstance(dbv, dict):
            data = dbv
        else:
            path = os.path.join(CATALOG_DIR, 'product_categories.json')
            if os.path.exists(path):
                with open(path, 'r', encoding='utf-8') as f:
                    loaded = json.load(f)
                data = loaded if isinstance(loaded, dict) else {}
            else:
                data = {}
    except Exception as e:
        logger.exception('get_product_categories failed: %s', e)
        data = {}
    headers = _cors_headers_for_request(request)
    return JSONResponse(status_code=200, content=data, headers=headers)


@app.post('/product-categories')
async def set_product_categories(request: Request, db: Session = Depends(get_db)):
    try:
        body = await request.json()
        if not isinstance(body, dict):
            raise HTTPException(status_code=400, detail='body must be an object mapping productKey to array')
        # Persist into DB (best-effort)
        try:
            ok = crud.set_setting(db, 'product_categories', body)
            if not ok:
                logger.warning('set_product_categories: failed to persist mapping to DB')
        except Exception:
            logger.exception('set_product_categories: DB persist failed')
        # Also write local snapshot file for compatibility
        try:
            os.makedirs(CATALOG_DIR, exist_ok=True)
            path = os.path.join(CATALOG_DIR, 'product_categories.json')
            with open(path, 'w', encoding='utf-8') as f:
                f.write(json.dumps(body, ensure_ascii=False, indent=2))
        except Exception:
            logger.exception('set_product_categories: failed to write snapshot file')
        headers = _cors_headers_for_request(request)
        try:
            await push_event({'type':'product-categories-updated','count': len(body)})
        except Exception:
            pass
        logger.info('Saved product_categories.json with %s keys', len(body))
        return JSONResponse(status_code=200, content={'ok': True, 'saved': len(body)}, headers=headers)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception('set_product_categories failed: %s', e)
        raise HTTPException(status_code=500, detail='Failed to save product categories')


def _run_add_user_columns() -> dict:
    """Run the same migration logic as `add_user_columns.py` and return a report dict."""
    results = { 'added': [], 'skipped': [], 'failed': [] }
    needed = [
        ('status', "VARCHAR(50) DEFAULT 'recibido'"),
        ('customer_type', "VARCHAR(50) DEFAULT 'mayorista'"),
        ('user_id', 'INTEGER'),
        ('user_full_name', 'VARCHAR(200)'),
        ('user_email', 'VARCHAR(320)'),
        ('user_barrio', 'VARCHAR(200)'),
        ('user_calle', 'VARCHAR(200)'),
        ('user_numeracion', 'VARCHAR(100)'),
        ('user_postal_code', 'VARCHAR(20)'),
        ('user_department', 'VARCHAR(120)'),
        ('_token_received', 'BOOLEAN'),
        ('_token_preview', 'TEXT'),
        ('source', "VARCHAR(50) DEFAULT 'web'"),
        ('payment_method', 'VARCHAR(50)'),
        ('payment_status', 'VARCHAR(50)'),
        ('payment_reference', 'VARCHAR(200)'),
        ('scheduled_delivery_date', 'VARCHAR(10)'),
        ('delivery_cutoff_applied', 'BOOLEAN'),
        ('delivery_timezone', 'VARCHAR(80)'),
        ('delivery_cutoff_hour', 'INTEGER'),
        ('assigned_driver_id', 'INTEGER'),
        ('assigned_driver_username', 'VARCHAR(80)'),
        ('assigned_driver_name', 'VARCHAR(200)'),
        ('assigned_driver_zone', 'VARCHAR(120)'),
        ('assigned_at', 'TIMESTAMP'),
        ('delivery_lat', 'FLOAT'),
        ('delivery_lon', 'FLOAT'),
        ('route_id', 'VARCHAR(120)'),
        ('route_order', 'INTEGER'),
        ('route_generated_at', 'TIMESTAMP'),
        ('delivered_at', 'TIMESTAMP'),
        ('delivered_by_id', 'INTEGER'),
        ('delivered_by_username', 'VARCHAR(80)'),
    ]
    dialect = engine.dialect.name if engine and getattr(engine, 'dialect', None) else ''
    try:
        insp = inspect(engine)
        has_orders = 'orders' in insp.get_table_names()
        if not has_orders:
            return {'error': 'no_orders_table'}
        existing = {c['name'] for c in insp.get_columns('orders')}
        with engine.begin() as conn:
            for name, coltype in needed:
                if name in existing:
                    results['skipped'].append(name)
                    continue
                try:
                    if 'postgres' in dialect:
                        sql = f"ALTER TABLE orders ADD COLUMN IF NOT EXISTS {name} {coltype};"
                    else:
                        sql = f"ALTER TABLE orders ADD COLUMN {name} {coltype};"
                    conn.execute(text(sql))
                    results['added'].append(name)
                except Exception as e:
                    try:
                        _invalidate_conn(conn)
                    except Exception:
                        pass
                    results['failed'].append({ 'name': name, 'error': str(e) })
        return results
    except Exception as e:
        logger.exception('Migration helper failed: %s', e)
        return {'error': str(e)}


@app.get('/debug/db-columns')
async def debug_db_columns(request: Request):
    """Return column names from `orders` table for diagnosis. Requires MIGRATION_SECRET header."""
    secret = os.environ.get('MIGRATION_SECRET')
    provided = request.headers.get('x-migrate-secret')
    if secret and provided != secret:
        return JSONResponse(status_code=403, content={'error': 'forbidden'})
    try:
        insp = inspect(engine)
        if 'orders' not in insp.get_table_names():
            return JSONResponse(status_code=200, content={'columns': []})
        cols = [c['name'] for c in insp.get_columns('orders')]
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=200, content={'columns': cols}, headers=headers)
    except Exception as e:
        logger.exception('debug_db_columns failed: %s', e)
        return JSONResponse(status_code=500, content={'error': str(e)})


@app.get('/debug/db-info')
async def debug_db_info(request: Request):
    """Return basic DB engine info (dialect and whether using sqlite). Requires MIGRATION_SECRET header."""
    secret = os.environ.get('MIGRATION_SECRET')
    provided = request.headers.get('x-migrate-secret')
    if secret and provided != secret:
        return JSONResponse(status_code=403, content={'error': 'forbidden'})
    try:
        dialect = getattr(engine, 'dialect', None)
        dialect_name = getattr(dialect, 'name', '') if dialect else ''
        using_sqlite = 'sqlite' in dialect_name
        # Masked DB url
        db_env = os.environ.get('DATABASE_URL')
        masked = 'sqlite (local file)'
        if db_env:
            try:
                from urllib.parse import urlparse
                u = urlparse(db_env)
                user = u.username or ''
                host = u.hostname or ''
                port = u.port or ''
                path = u.path or ''
                masked = f"{u.scheme}://{user + ':****@' if user else ''}{host}{(':'+str(port)) if port else ''}{path}"
            except Exception:
                masked = 'postgres (masked)'
        headers = _cors_headers_for_request(request)
        # quick connectivity test
        ok = False
        try:
            with engine.connect() as conn:
                conn.execute(text('SELECT 1'))
                ok = True
        except Exception:
            ok = False
        return JSONResponse(status_code=200, content={'dialect': dialect_name, 'using_sqlite': using_sqlite, 'database_url_masked': masked, 'connection_ok': ok}, headers=headers)
    except Exception as e:
        logger.exception('debug_db_info failed: %s', e)
        return JSONResponse(status_code=500, content={'error': str(e)})


@app.post('/debug/migrate')
async def debug_migrate(request: Request):
    """Run migration to add missing `user_*` columns to `orders` table. Requires MIGRATION_SECRET header."""
    secret = os.environ.get('MIGRATION_SECRET')
    provided = request.headers.get('x-migrate-secret')
    if secret and provided != secret:
        return JSONResponse(status_code=403, content={'error': 'forbidden'})
    # Run migration (blocking but quick)
    result = _run_add_user_columns()
    headers = _cors_headers_for_request(request)
    return JSONResponse(status_code=200, content={'result': result}, headers=headers)


@app.post('/debug/push-order')
async def debug_push_order(request: Request):
    """Debug-only: push a synthetic order event to connected admin clients. Requires MIGRATION_SECRET if set."""
    secret = os.environ.get('MIGRATION_SECRET')
    provided = request.headers.get('x-migrate-secret')
    if secret and provided != secret:
        return JSONResponse(status_code=403, content={'error': 'forbidden'})
    try:
        body = await request.json()
    except Exception:
        body = {}
    # ensure we include at least an id
    if not body.get('id'):
        import time
        body['id'] = f"debug-{int(time.time())}"
        body['items'] = body.get('items', [])
        body['total'] = body.get('total', 0)
        body['created_at'] = body.get('created_at')
    await push_event({"action": "order_created", "order": body})
    headers = _cors_headers_for_request(request)
    return JSONResponse(status_code=200, content={'ok': True, 'order': body}, headers=headers)


@app.post('/debug/backfill-token-previews')
async def debug_backfill_token_previews(request: Request):
    """Attempt to persist cached token previews from ORDER_PAYLOAD_CACHE into the DB.
    Only for debugging; requires MIGRATION_SECRET header if set."""
    secret = os.environ.get('MIGRATION_SECRET')
    provided = request.headers.get('x-migrate-secret')
    if secret and provided != secret:
        return JSONResponse(status_code=403, content={'error': 'forbidden'})

    results = {'updated': [], 'skipped': [], 'failed': []}
    try:
        insp = inspect(engine)
        existing = {c['name'] for c in insp.get_columns('orders')}
    except Exception:
        existing = set()

    if '_token_preview' not in existing and '_token_received' not in existing:
        return JSONResponse(status_code=400, content={'error': 'token_columns_missing'})

    import json as _json
    now = time.time()
    for oid, rec in list(ORDER_PAYLOAD_CACHE.items()):
        try:
            p = rec.get('payload') if rec and isinstance(rec, dict) else None
            if not p or not p.get('_token_preview'):
                results['skipped'].append({'id': oid, 'reason': 'no_preview'})
                continue
            tp = p.get('_token_preview')
            # Serialize
            tp_json = _json.dumps(tp, ensure_ascii=False)
            # Attempt update via transactional context
            try:
                with engine.begin() as conn:
                    # Try numeric id first
                    try:
                        iid = int(oid)
                        conn.execute(text('UPDATE orders SET _token_preview = :tp, _token_received = :tr WHERE id = :id'), {'tp': tp_json, 'tr': True, 'id': iid})
                    except Exception:
                        conn.execute(text('UPDATE orders SET _token_preview = :tp, _token_received = :tr WHERE CAST(id AS TEXT) = :id'), {'tp': tp_json, 'tr': True, 'id': str(oid)})
                results['updated'].append(oid)
            except Exception as e:
                results['failed'].append({'id': oid, 'error': str(e)})
        except Exception as e:
            results['failed'].append({'id': oid, 'error': str(e)})

    headers = _cors_headers_for_request(request)
    return JSONResponse(status_code=200, content={'result': results}, headers=headers)


@app.get('/debug/whoami')
async def debug_whoami(request: Request):
    """Return decoded bearer token payload for debugging (safe to call from browser)."""
    auth = request.headers.get('authorization') or request.headers.get('Authorization')
    headers = _cors_headers_for_request(request)
    if not auth or not isinstance(auth, str) or not auth.lower().startswith('bearer '):
        return JSONResponse(status_code=200, content={'ok': False, 'error': 'no_authorization'}, headers=headers)
    token = auth.split(' ', 1)[1]
    try:
        payload = utils.decode_access_token(token)
        # Mask token in logs
        try: logger.info('debug_whoami called; token=***%s payload=%s', token[-12:], payload)  # last chars for identification
        except Exception: pass
        if not payload:
            return JSONResponse(status_code=200, content={'ok': False, 'error': 'invalid_token'}, headers=headers)
        return JSONResponse(status_code=200, content={'ok': True, 'payload': payload}, headers=headers)
    except Exception as e:
        logger.exception('debug_whoami failed: %s', e)
        return JSONResponse(status_code=500, content={'ok': False, 'error': str(e)}, headers=headers)


@app.post('/debug/test-email')
async def debug_test_email(request: Request):
    """Send a test order-confirmation email to validate Resend setup.
    Requires MIGRATION_SECRET header when configured.
    """
    secret = os.environ.get('MIGRATION_SECRET')
    provided = request.headers.get('x-migrate-secret')
    if secret and provided != secret:
        return JSONResponse(status_code=403, content={'error': 'forbidden'})

    try:
        body = await request.json()
    except Exception:
        body = {}

    to_email = _normalize_email((body or {}).get('to'))
    if not to_email:
        return JSONResponse(
            status_code=400,
            content={'error': 'missing_to', 'detail': 'Provide a valid `to` email in JSON body.'},
        )

    test_order = {
        'id': body.get('order_id') or f"test-{int(time.time())}",
        'total': body.get('total') if body.get('total') is not None else 0,
        'items': body.get('items') if isinstance(body.get('items'), list) else [{'id': 'TEST', 'qty': 1}],
        'user_email': to_email,
        'user_full_name': body.get('name') or 'Cliente test',
        'created_at': datetime.datetime.utcnow().isoformat(),
    }

    ok = await _send_order_confirmation_email(
        order_data=test_order,
        request_data={'user_email': to_email},
        token_payload=None,
    )
    return JSONResponse(
        status_code=200,
        content={
            'ok': bool(ok),
            'to': to_email,
            'order_id': test_order['id'],
            'config': _resend_status_snapshot(),
        },
    )



@app.middleware("http")
async def log_requests(request: Request, call_next):
    logger.info(f"{request.method} {request.url}")
    return await call_next(request)


@app.middleware("http")
async def ensure_cors_middleware(request: Request, call_next):
    """Handle preflight OPTIONS and ensure CORS response headers are present on every response.
    This echoes the request Origin when present (required when credentials are used), or falls back to '*'.
    """
    # Fast-path preflight
    if request.method == 'OPTIONS':
        headers = _cors_headers_for_request(request)
        return Response(status_code=204, headers=headers)

    response = await call_next(request)

    # Add CORS headers if not already present
    try:
        headers = _cors_headers_for_request(request)
        for k, v in headers.items():
            if k not in response.headers:
                response.headers[k] = v
    except Exception:
        # be defensive: don't break the response if header handling fails
        pass

    return response


# -------------------------------------------------------------------
# PATHS
# -------------------------------------------------------------------
ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
# Allow overriding upload directory via env var so a Render persistent disk can be mounted there.
# If not provided, fall back to the local `uploads` folder inside the project (ephemeral on some hosts).
UPLOAD_DIR = os.environ.get('UPLOAD_DIR') or os.path.join(ROOT_DIR, "uploads")
PROMO_DIR = os.path.join(UPLOAD_DIR, 'promos')
FRONTEND_DIR = os.path.join(ROOT_DIR, "admin")
CATALOG_DIR = os.environ.get('CATALOG_DIR') or os.path.join(ROOT_DIR, "catalogo")

# ensure catalog dir exists when possible (if mounted as persistent volume this will create it)
try:
    os.makedirs(CATALOG_DIR, exist_ok=True)
except Exception:
    pass

utils.ensure_upload_folder(UPLOAD_DIR)
try:
    os.makedirs(PROMO_DIR, exist_ok=True)
except Exception:
    pass

# --- Promotional images API (admin/frontend) ---
@app.get('/api/promos')
def list_promos(request: Request, db: Session = Depends(get_db)):
    """Return only the images selected for the public promo carousel.
    The selection is stored in `promo_list.json` under the catalog directory.
    """
    try:
        # Prefer DB-backed selection of promo images. Fall back to scanned uploads when DB empty.
        selected = []
        try:
            promos = db.query(models.PromoImage).filter(models.PromoImage.selected == True).order_by(models.PromoImage.created_at.desc()).all()
            for p in promos:
                # prefer stored URL, but make it absolute so frontend/admin can load across origins
                try:
                    base = str(request.base_url).rstrip('/')
                except Exception:
                    base = ''
                stored = (p.url or '')
                use_filename = False
                if stored:
                    s = stored.replace('\\', '/')
                    if s.startswith('http://') or s.startswith('https://'):
                        url = s
                    elif s.startswith('/'):
                        url = (base + s) if base else s
                    else:
                        # treat suspicious filesystem-like values as filesystem
                        if (':' in stored and ('\\' in stored or '/' in stored)) or '\\' in stored:
                            use_filename = True
                        else:
                            if s.startswith('uploads/'):
                                s = '/' + s
                            url = (base + s) if base else s
                else:
                    use_filename = True

                if use_filename:
                    url = f'/uploads/promos/{p.filename}'
                    if base:
                        url = base + url

                if isinstance(url, str):
                    url = url.replace('\\', '/')

                selected.append({ 'url': url, 'alt': p.alt or p.filename, 'name': p.filename, 'ts': int(p.created_at.timestamp()) if p.created_at else None })
            # If DB has none, fall back to scanning PROMO_DIR for compatibility
            if not selected and os.path.isdir(PROMO_DIR):
                for fname in sorted(os.listdir(PROMO_DIR)):
                    if fname.startswith('.'):
                        continue
                    path = os.path.join(PROMO_DIR, fname)
                    if not os.path.isfile(path):
                        continue
                    try:
                        base = str(request.base_url).rstrip('/')
                    except Exception:
                        base = ''
                    url = f'/uploads/promos/{fname}'
                    if base:
                        url = base + url
                    selected.append({ 'url': url, 'alt': fname, 'name': fname, 'ts': int(os.path.getmtime(path)) if os.path.exists(path) else None })
        except Exception:
            selected = []
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=200, content=selected, headers=headers)
    except Exception as e:
        logger.exception('list_promos failed: %s', e)
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=500, content={'detail': 'failed'}, headers=headers)


@app.post('/api/promos')
async def upload_promo(file: UploadFile = File(...), request: Request = None, db: Session = Depends(get_db)):
    """Save promo image either to S3 (if configured) or local uploads and persist metadata.

    Uses `utils.save_upload_file` so deployments on hosts like Render can be configured
    to use a persistent `UPLOAD_DIR` or S3 via env vars.
    """
    try:
        # Ensure upload folder exists
        utils.ensure_upload_folder(PROMO_DIR)

        # Read file content so we can both persist to external storage and to DB
        contents = await file.read()
        orig_name = os.path.basename(file.filename or 'promo')
        s3_bucket = os.environ.get('S3_BUCKET') or os.environ.get('AWS_S3_BUCKET')

        # Persist image bytes into DB `images` table first so promo images are durable
        img_id = None
        try:
            img = models.Image(data=contents, mime=getattr(file, 'content_type', None), filename=orig_name)
            db.add(img)
            db.commit()
            db.refresh(img)
            img_id = img.id
        except Exception:
            try:
                db.rollback()
            except Exception:
                pass
            img_id = None

        # Best-effort: also save to S3 or disk for compatibility, but do not rely on it
        try:
            _ = await utils.save_bytes_upload(contents, orig_name, PROMO_DIR, s3_bucket=s3_bucket, content_type=getattr(file, 'content_type', None))
        except Exception:
            # ignore storage errors; DB-backed image is authoritative
            pass

        # Build a durable URL that points to our DB-backed image endpoint when possible
        if img_id:
            try:
                base = str(request.base_url).rstrip('/')
            except Exception:
                base = ''
            url = (base + f'/images/{img_id}') if base else f'/images/{img_id}'
            fname = orig_name
        else:
            # fallback: try the saved path or public uploads path
            dest_path = None
            try:
                dest_path = await utils.save_bytes_upload(contents, orig_name, PROMO_DIR, s3_bucket=s3_bucket, content_type=getattr(file, 'content_type', None))
            except Exception:
                dest_path = None
            fname = os.path.basename(str(dest_path)) if dest_path else orig_name
            try:
                base = str(request.base_url).rstrip('/')
            except Exception:
                base = ''
            url_path = f'/uploads/promos/{fname}'
            url = (base + url_path) if base else url_path

        # persist metadata in DB (store absolute URL when possible)
        try:
            # Try to avoid duplicate PromoImage rows: match by exact filename or by suffix of original filename
            existing_pi = db.query(models.PromoImage).filter(
                or_(models.PromoImage.filename == fname, models.PromoImage.filename.like(f'%{orig_name}'))
            ).first()
            if existing_pi:
                existing_pi.url = url
                existing_pi.alt = fname
                existing_pi.selected = False
                db.add(existing_pi)
                db.commit()
            else:
                pi = models.PromoImage(filename=fname, url=url, alt=fname, selected=False)
                db.add(pi)
                db.commit()
        except Exception:
            try:
                db.rollback()
            except Exception:
                pass

        headers = _cors_headers_for_request(request or Request)
        return JSONResponse(status_code=201, content={'url': url, 'name': fname}, headers=headers)
    except Exception as e:
        logger.exception('upload_promo failed: %s', e)
        headers = _cors_headers_for_request(request or Request)
        return JSONResponse(status_code=500, content={'detail': 'upload failed'}, headers=headers)


@app.get('/api/uploads')
def list_uploads(request: Request, db: Session = Depends(get_db)):
    """Return all uploaded image files (for admin browsing)."""
    try:
        items = []
        try:
            # prefer DB records
            rows = db.query(models.PromoImage).order_by(models.PromoImage.created_at.desc()).all()
            for r in rows:
                try:
                    base = str(request.base_url).rstrip('/')
                except Exception:
                    base = ''
                # Normalize stored URL ? if it's an absolute HTTP URL, keep it.
                stored = (r.url or '')
                # If stored URL looks like a filesystem path (backslashes or drive letter), ignore it and build public path
                use_filename = False
                if stored:
                    s = stored.replace('\\', '/')
                    if s.startswith('http://') or s.startswith('https://'):
                        url = s
                    elif s.startswith('/'):
                        url = (base + s) if base else s
                    else:
                        # If it contains a colon (windows drive C:) or a backslash originally, treat as filesystem
                        if (':' in stored and ('\\' in stored or '/' in stored)) or '\\' in stored:
                            use_filename = True
                        else:
                            # relative path like 'uploads/promos/...' ? ensure leading slash
                            if s.startswith('uploads/'):
                                s = '/' + s
                            url = (base + s) if base else s
                else:
                    use_filename = True

                if use_filename:
                    url = f'/uploads/promos/{r.filename}'
                    if base:
                        url = base + url

                # final normalize
                if isinstance(url, str):
                    url = url.replace('\\', '/')

                items.append({ 'url': url, 'alt': r.alt or r.filename, 'name': r.filename, 'ts': int(r.created_at.timestamp()) if r.created_at else None, 'selected': bool(r.selected) })
            # If DB empty, fallback to scanning disk
            if not items and os.path.isdir(PROMO_DIR):
                for fname in sorted(os.listdir(PROMO_DIR)):
                    if fname.startswith('.'):
                        continue
                    path = os.path.join(PROMO_DIR, fname)
                    if not os.path.isfile(path):
                        continue
                    lower = fname.lower()
                    if not (lower.endswith('.png') or lower.endswith('.jpg') or lower.endswith('.jpeg') or lower.endswith('.webp') or lower.endswith('.gif')):
                        continue
                    try:
                        mtime = int(os.path.getmtime(path))
                    except Exception:
                        mtime = None
                    try:
                        base = str(request.base_url).rstrip('/')
                    except Exception:
                        base = ''
                    url = f'/uploads/promos/{fname}'
                    if base:
                        url = base + url
                    url = url.replace('\\', '/')
                    items.append({ 'url': url, 'alt': fname, 'name': fname, 'ts': mtime, 'selected': False })
        except Exception:
            items = []
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=200, content=items, headers=headers)
    except Exception as e:
        logger.exception('list_uploads failed: %s', e)
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=500, content={'detail': 'failed'}, headers=headers)


@app.delete('/api/promos/{filename}')
def delete_promo(filename: str, request: Request, db: Session = Depends(get_db)):
    try:
        # prevent traversal
        fname = os.path.basename(filename)
        path = os.path.join(PROMO_DIR, fname)

        deleted_any = False

        # If the file exists on disk, remove it
        try:
            if os.path.exists(path) and os.path.isfile(path):
                os.remove(path)
                deleted_any = True
        except Exception as e:
            logger.exception('Failed removing file %s: %s', path, e)

        # Remove associated promo DB record(s)
        try:
            rec = db.query(models.PromoImage).filter(models.PromoImage.filename == fname).first()
            if rec:
                # If the promo image points to a DB-backed image URL like /images/{id}, remove that image row too
                try:
                    if rec.url and isinstance(rec.url, str) and '/images/' in rec.url:
                        # extract numeric id if possible
                        try:
                            img_id = int(rec.url.rstrip('/').split('/')[-1])
                            db.query(models.Image).filter(models.Image.id == img_id).delete()
                            db.commit()
                            deleted_any = True
                        except Exception:
                            db.rollback()
                except Exception:
                    db.rollback()
                # delete the promo record
                try:
                    db.query(models.PromoImage).filter(models.PromoImage.filename == fname).delete()
                    db.commit()
                    deleted_any = True
                except Exception:
                    db.rollback()
        except Exception:
            db.rollback()

        headers = _cors_headers_for_request(request)
        if deleted_any:
            return JSONResponse(status_code=200, content={'detail': 'deleted'}, headers=headers)
        else:
            return JSONResponse(status_code=404, content={'detail': 'not found or not deletable'}, headers=headers)
    except Exception as e:
        logger.exception('delete_promo failed: %s', e)
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=500, content={'detail': 'delete failed'}, headers=headers)


# Fallback route: serve promos from disk when present, otherwise from DB `images` table.
# This lets the existing `/uploads/promos/<file>` URLs keep working for admins without
# requiring an immediate migration of DB rows; if the file is missing, try to locate
# the binary in the `images` table by filename or via a linked `promo_images` record.
@app.get('/uploads/promos/{filename}')
def serve_promo_fallback(filename: str, request: Request, db: Session = Depends(get_db)):
    try:
        fname = os.path.basename(filename)
        path = os.path.join(PROMO_DIR, fname)

        # If file exists on disk, let FastAPI serve it directly for efficiency
        if os.path.exists(path) and os.path.isfile(path):
            return FileResponse(path)

        # Try to resolve a DB-backed image via the PromoImage record
        try:
            rec = db.query(models.PromoImage).filter(models.PromoImage.filename == fname).first()
            if rec and rec.url and isinstance(rec.url, str) and '/images/' in rec.url:
                try:
                    img_id = int(rec.url.rstrip('/').split('/')[-1])
                    img = db.query(models.Image).filter(models.Image.id == img_id).first()
                    if img and getattr(img, 'data', None):
                        headers = _cors_headers_for_request(request)
                        return Response(content=img.data, media_type=(img.mime or 'application/octet-stream'), headers=headers)
                except Exception:
                    pass
        except Exception:
            pass

        # As a last resort, try to find an Image row by filename
        try:
            img = db.query(models.Image).filter(models.Image.filename == fname).first()
            if img and getattr(img, 'data', None):
                headers = _cors_headers_for_request(request)
                return Response(content=img.data, media_type=(img.mime or 'application/octet-stream'), headers=headers)
        except Exception:
            pass

        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=404, content={'detail': 'not found'}, headers=headers)
    except Exception as e:
        logger.exception('serve_promo_fallback failed: %s', e)
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=500, content={'detail': 'internal error'}, headers=headers)


@app.post('/api/promos/select')
def select_promo(request: Request, db: Session = Depends(get_db)):
    """Add a filename to the promo_list (body: { "name": "filename" })."""
    try:
        body = request.json() if hasattr(request, 'json') else None
    except Exception:
        body = None
    try:
        data = None
        try:
            data = request.json()
        except Exception:
            try:
                # fallback to reading body bytes
                import asyncio
                data = None
            except Exception:
                data = None
        # simpler: parse raw body
        try:
            raw = request._body if hasattr(request, '_body') else None
        except Exception:
            raw = None
        # We'll instead read from the request stream synchronously (FastAPI gives us Request instance)
        # But to keep compatibility, accept query param 'name'
        name = None
        try:
            name = request.query_params.get('name')
        except Exception:
            name = None
        if not name:
            # try json load from body
            try:
                import json as _json
                raw_body = request._body if hasattr(request, '_body') else None
                if not raw_body:
                    # attempt to read via awaitable (not possible here), return error
                    raise HTTPException(status_code=400, detail='name required')
                pd = _json.loads(raw_body)
                name = pd.get('name')
            except Exception:
                raise HTTPException(status_code=400, detail='name required')

        fname = os.path.basename(name)
        path = os.path.join(PROMO_DIR, fname)
        if not os.path.isfile(path):
            # If the file is not present on disk, allow selection if a DB record exists (tolerant behavior)
            rec = db.query(models.PromoImage).filter(models.PromoImage.filename == fname).first()
            if not rec:
                headers = _cors_headers_for_request(request)
                return JSONResponse(status_code=404, content={'detail': 'file not found'}, headers=headers)
        # mark DB record selected (and deselect others if necessary)
        try:
            # ensure record exists
            rec = db.query(models.PromoImage).filter(models.PromoImage.filename == fname).first()
            if not rec:
                # create record with absolute URL when possible
                try:
                    base = str(request.base_url).rstrip('/')
                except Exception:
                    base = ''
                url_path = f'/uploads/promos/{fname}'
                url = (base + url_path) if base else url_path
                rec = models.PromoImage(filename=fname, url=url, alt=fname, selected=True)
                db.add(rec)
            else:
                rec.selected = True
            # commit selection
            db.commit()
        except Exception:
            db.rollback()
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=200, content={'detail': 'selected', 'name': fname}, headers=headers)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception('select_promo failed: %s', e)
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=500, content={'detail': 'failed'}, headers=headers)


@app.delete('/api/promos/select/{filename}')
def deselect_promo(filename: str, request: Request, db: Session = Depends(get_db)):
    try:
        fname = os.path.basename(filename)
        try:
            rec = db.query(models.PromoImage).filter(models.PromoImage.filename == fname).first()
            if not rec:
                headers = _cors_headers_for_request(request)
                return JSONResponse(status_code=404, content={'detail': 'not found'}, headers=headers)
            rec.selected = False
            db.commit()
            headers = _cors_headers_for_request(request)
            return JSONResponse(status_code=200, content={'detail': 'deselected', 'name': fname}, headers=headers)
        except Exception:
            try:
                db.rollback()
            except Exception:
                pass
            headers = _cors_headers_for_request(request)
            return JSONResponse(status_code=500, content={'detail': 'failed to deselect'}, headers=headers)
    except Exception as e:
        logger.exception('deselect_promo failed: %s', e)
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=500, content={'detail': 'failed'}, headers=headers)


# --- Consumición inmediata API (admin) ---
@app.get('/api/consumos')
def list_consumos(request: Request, db: Session = Depends(get_db)):
    """Return consumos config: list of { id: product_id, discount: percent }.
    Prefer DB-backed settings when available; fall back to consumos.json.
    """
    try:
        items = []
        # Prefer DB-backed setting if present
        try:
            rec = db.query(models.Setting).filter(models.Setting.key == 'consumos').first()
            if rec and rec.value:
                try:
                    items = json.loads(rec.value) or []
                except Exception:
                    items = []
        except Exception:
            items = []
        # Fallback to file if DB empty
        if not items:
            consumos_path = os.path.join(CATALOG_DIR, 'consumos.json')
            if os.path.exists(consumos_path):
                try:
                    with open(consumos_path, 'r', encoding='utf-8') as f:
                        items = json.load(f) or []
                except Exception:
                    items = []
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=200, content=items, headers=headers)
    except Exception as e:
        logger.exception('list_consumos failed: %s', e)
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=500, content={'detail': 'failed'}, headers=headers)


@app.post('/api/consumos')
async def save_consumos(request: Request, db: Session = Depends(get_db)):
    """Replace the consumos list. Accepts JSON array in body."""
    try:
        body = await request.body()
        try:
            data = json.loads(body) if body else []
        except Exception:
            data = []
        if not isinstance(data, list):
            raise HTTPException(status_code=400, detail='expected list')
        # If client attempts to save an empty list, require explicit confirmation to avoid accidental wipes
        if len(data) == 0 and not request.query_params.get('confirm'):
            raise HTTPException(status_code=400, detail='empty-list-requires-confirm')
        # Normalize and validate incoming entries (coerce id and numeric fields). Invalid entries are rejected.
        cleaned = []
        for entry in data:
            if not isinstance(entry, dict):
                continue
            try:
                pid = int(entry.get('id'))
            except Exception:
                continue
            # tolerate strings with comma decimal (e.g. "10,5") and percent signs
            try:
                raw_disc = entry.get('discount') if entry.get('discount') is not None else entry.get('value', 0)
                if isinstance(raw_disc, str):
                    raw_disc = raw_disc.replace('%', '').strip().replace(',', '.')
                discount = float(raw_disc) if raw_disc is not None else 0.0
            except Exception:
                discount = 0.0
            try:
                raw_qty = entry.get('qty') if entry.get('qty') is not None else entry.get('cantidad', 0)
                if isinstance(raw_qty, str):
                    raw_qty = raw_qty.strip().replace(',', '.')
                qty = int(float(raw_qty)) if raw_qty is not None else 0
            except Exception:
                qty = 0
            # keep entries with a positive discount and non-negative qty
            if discount <= 0 or qty < 0:
                continue
            # Default consumos to percent-type discounts unless caller specifies otherwise
            ctype = entry.get('type') if isinstance(entry, dict) else None
            if not ctype and discount > 0:
                ctype = 'percent'
            cleaned.append({'id': pid, 'discount': float(discount), 'qty': int(qty), 'type': ctype})
        # If after cleaning there is nothing to save but the original data wasn't empty, reject to avoid accidental clears
        if len(cleaned) == 0 and len(data) > 0:
            raise HTTPException(status_code=400, detail='no-valid-consumos')
        data = cleaned
        consumos_path = os.path.join(CATALOG_DIR, 'consumos.json')
        # write and verify success; if writing fails return 500 so client knows
        try:
            os.makedirs(CATALOG_DIR, exist_ok=True)
            with open(consumos_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            # Persist to DB settings so consumos survive deploys
            try:
                rec = db.query(models.Setting).filter(models.Setting.key == 'consumos').first()
                if rec:
                    rec.value = json.dumps(data, ensure_ascii=False)
                else:
                    rec = models.Setting(key='consumos', value=json.dumps(data, ensure_ascii=False))
                    db.add(rec)
                db.commit()
            except Exception as db_err:
                try:
                    db.rollback()
                except Exception:
                    pass
                logger.exception('save_consumos db write failed: %s', db_err)
            # Update static snapshot with consumos reflected (best-effort) and notify WS clients
            try:
                await anyio.to_thread.run_sync(write_catalog_snapshot)
            except Exception:
                logger.exception('write_catalog_snapshot after save_consumos failed')
            try:
                await push_event({"action": "consumos-updated", "consumos": data})
            except Exception:
                logger.exception('push_event consumos-updated failed')
            headers = _cors_headers_for_request(request)
            return JSONResponse(status_code=200, content={'detail': 'saved'}, headers=headers)
        except Exception as write_err:
            logger.exception('save_consumos write failed: %s', write_err)
            headers = _cors_headers_for_request(request)
            return JSONResponse(status_code=500, content={'detail': 'failed to save consumos'}, headers=headers)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception('save_consumos failed: %s', e)
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=500, content={'detail': 'failed'}, headers=headers)


@app.delete('/api/consumos/{product_id}')
def delete_consumo(product_id: str, request: Request):
    try:
        pid = int(product_id)
    except Exception:
        pid = product_id
    try:
        consumos_path = os.path.join(CATALOG_DIR, 'consumos.json')
        items = []
        if os.path.exists(consumos_path):
            try:
                with open(consumos_path, 'r', encoding='utf-8') as f:
                    items = json.load(f) or []
            except Exception:
                items = []
            new_items = [it for it in items if it.get('id') != pid and str(it.get('id')) != str(pid)]
            try:
                with open(consumos_path, 'w', encoding='utf-8') as f:
                    json.dump(new_items, f, ensure_ascii=False, indent=2)
                headers = _cors_headers_for_request(request)
                return JSONResponse(status_code=200, content={'detail': 'deleted'}, headers=headers)
            except Exception as write_err:
                logger.exception('delete_consumo write failed: %s', write_err)
                headers = _cors_headers_for_request(request)
                return JSONResponse(status_code=500, content={'detail': 'failed to delete consumos'}, headers=headers)
    except Exception as e:
        logger.exception('delete_consumo failed: %s', e)
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=500, content={'detail': 'failed'}, headers=headers)

# -------------------------------------------------------------------
# WEBSOCKETS
# -------------------------------------------------------------------
connections: List[WebSocket] = []

async def push_event(data: dict):
    """Broadcast JSON `data` to all connected WebSocket clients and log diagnostics."""
    try:
        logger.info('push_event sending to %s connections: %s', len(connections), data)
    except Exception:
        pass
    tasks = []
    for ws in list(connections):
        try:
            tasks.append(ws.send_json(data))
        except Exception:
            try:
                connections.remove(ws)
            except Exception:
                pass
    if tasks:
        await asyncio.gather(*tasks, return_exceptions=True)
    try:
        logger.info('push_event completed')
    except Exception:
        pass

@app.websocket("/ws/products")
async def ws_products(ws: WebSocket):
    await ws.accept()
    connections.append(ws)
    try:
        while True:
            await asyncio.sleep(1)
    except WebSocketDisconnect:
        pass
    finally:
        if ws in connections:
            connections.remove(ws)

# -------------------------------------------------------------------
# ENDPOINTS
# -------------------------------------------------------------------
@app.post("/upload-image")
async def upload_image(file: UploadFile = File(...)):
    # Read content
    content = await file.read()
    mime = getattr(file, 'content_type', None) or 'application/octet-stream'
    filename = getattr(file, 'filename', None)

    # Save into DB Image table (run blocking DB write in thread)
    def task():
        db = SessionLocal()
        try:
            img = models.Image(data=content, mime=mime, filename=filename)
            db.add(img)
            db.commit()
            db.refresh(img)
            return img.id
        finally:
            db.close()

    img_id = await anyio.to_thread.run_sync(task)
    return {"image_url": f"/images/{img_id}"}


@app.post("/upload-image-url")
async def upload_image_url(payload: Dict[str, Any] = Body(default=None)):
    payload = payload or {}
    url_raw = payload.get('url') or payload.get('image_url') or payload.get('image') or payload.get('link')
    normalized = _normalize_import_image_url(url_raw)
    if not normalized:
        raise HTTPException(status_code=400, detail='URL inválida')
    if normalized.startswith('/'):
        # Accept local paths as-is
        return {"image_url": normalized, "source_url": normalized, "stored": False}
    if not re.match(r'^https?://', normalized, re.IGNORECASE):
        raise HTTPException(status_code=400, detail='URL inválida')

    max_bytes = int(os.environ.get('IMAGE_FETCH_MAX_BYTES') or 4_000_000)
    downloaded = _download_image_bytes(normalized, max_bytes=max_bytes)
    if not downloaded:
        raise HTTPException(status_code=400, detail='No se pudo descargar la imagen')
    content, mime, filename = downloaded

    def task():
        db = SessionLocal()
        try:
            img = models.Image(data=content, mime=mime, filename=filename)
            db.add(img)
            db.commit()
            db.refresh(img)
            return img.id
        finally:
            db.close()

    img_id = await anyio.to_thread.run_sync(task)
    return {"image_url": f"/images/{img_id}", "source_url": normalized, "stored": True}


@app.get('/images/{image_id}')
def get_image(image_id: int, db: Session = Depends(get_db)):
    img = db.query(models.Image).filter(models.Image.id == image_id).first()
    if not img:
        raise HTTPException(404, 'Image not found')
    return Response(img.data, media_type=img.mime)


# --- Auth endpoints ---
@app.post('/auth/register', response_model=schemas.UserResponse)
async def register(user: schemas.UserCreate):
    # validate unique email and create user with hashed password
    def task():
        db = SessionLocal()
        try:
            if crud.get_user_by_email(db, user.email):
                # signal duplicate via exception for the outer handler
                raise IntegrityError('duplicate', params=None, orig=None)
            hashed = utils.hash_password(user.password)
            new = crud.create_user(db, user, hashed)
            return new
        finally:
            db.close()

    try:
        new_user = await anyio.to_thread.run_sync(task)
    except IntegrityError:
        raise HTTPException(status_code=400, detail='Email already registered')
    except Exception as e:
        logger.exception('Unexpected error in /auth/register: %s', e)
        raise HTTPException(status_code=500, detail='Server error')
    return new_user


@app.post('/auth/token', response_model=schemas.Token)
async def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends()):
    def task():
        db = SessionLocal()
        try:
            user = crud.authenticate_user(db, form_data.username, form_data.password)
            return user
        finally:
            db.close()

    try:
        user = await anyio.to_thread.run_sync(task)
    except Exception as e:
        logger.exception('Unexpected error in /auth/token: %s', e)
        raise HTTPException(status_code=500, detail='Server error')

    if not user:
        raise HTTPException(status_code=401, detail='Incorrect credentials')
    access_token = utils.create_access_token({"sub": user.email, "id": user.id, "full_name": user.full_name})
    logger.info('login_for_access_token: issued token for user id=%s email=%s', user.id, user.email)
    return {"access_token": access_token, "token_type": "bearer"}


def get_current_user(token: str = Depends(oauth2_scheme)):
    try:
        payload = utils.decode_access_token(token)
    except Exception as e:
        logger.exception('get_current_user: decode_access_token failed: %s', e)
        raise HTTPException(status_code=401, detail='Invalid token')
    logger.info('get_current_user: token payload=%s', payload)
    if not payload:
        raise HTTPException(status_code=401, detail='Invalid token')
    email = payload.get('sub')
    if not email:
        raise HTTPException(status_code=401, detail='Invalid token payload')
    db = SessionLocal()
    try:
        user = crud.get_user_by_email(db, email)
        logger.info('get_current_user: user lookup for email %s -> %s', email, bool(user))
        if not user:
            raise HTTPException(status_code=401, detail='User not found')
        return user
    finally:
        db.close()


@app.get('/auth/me', response_model=schemas.UserResponse)
def auth_me(current_user = Depends(get_current_user)):
    return current_user


def _to_epoch_ms(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        dt = value
        if not isinstance(dt, datetime.datetime):
            return None
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=datetime.timezone.utc)
        return float(int(dt.timestamp() * 1000))
    except Exception:
        return None


def _serialize_user_address_row(row: Any) -> Dict[str, Any]:
    raw_id = str(getattr(row, 'id', '') or '').strip()
    try:
        user_id_int = int(getattr(row, 'user_id', 0) or 0)
    except Exception:
        user_id_int = 0
    prefix = f"{user_id_int}:"
    client_id = raw_id[len(prefix):] if (raw_id and prefix and raw_id.startswith(prefix)) else raw_id
    payload = {
        'id': client_id,
        'user_id': user_id_int,
        'label': str(getattr(row, 'label', '') or '').strip() or None,
        'notes': str(getattr(row, 'notes', '') or '').strip() or None,
        'barrio': str(getattr(row, 'barrio', '') or '').strip(),
        'calle': str(getattr(row, 'calle', '') or '').strip(),
        'numeracion': str(getattr(row, 'numeracion', '') or '').strip(),
        'postal_code': str(getattr(row, 'postal_code', '') or '').strip() or None,
        'department': str(getattr(row, 'department', '') or '').strip() or None,
        'query_hint': str(getattr(row, 'query_hint', '') or '').strip() or None,
        'full_text': str(getattr(row, 'full_text', '') or '').strip() or None,
        'lat': None,
        'lon': None,
        'is_default': bool(getattr(row, 'is_default', False)),
        # Keep frontend-compatible millisecond timestamp field.
        'created_at': _to_epoch_ms(getattr(row, 'created_at', None)),
        # Extra DB timestamps for diagnostics/admin compatibility.
        'created_at_db': getattr(row, 'created_at', None),
        'updated_at_db': getattr(row, 'updated_at', None),
    }
    try:
        lat_val = getattr(row, 'lat', None)
        payload['lat'] = round(float(lat_val), 6) if lat_val is not None else None
    except Exception:
        payload['lat'] = None
    try:
        lon_val = getattr(row, 'lon', None)
        payload['lon'] = round(float(lon_val), 6) if lon_val is not None else None
    except Exception:
        payload['lon'] = None
    return payload


@app.get('/auth/addresses', response_model=schemas.UserAddressBookResponse)
def auth_list_addresses(current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    try:
        uid = int(getattr(current_user, 'id'))
    except Exception:
        raise HTTPException(status_code=401, detail='Invalid user')
    try:
        try:
            crud.ensure_user_primary_address(db, current_user)
        except Exception:
            logger.exception('auth_list_addresses: ensure_user_primary_address failed')
        rows = crud.list_user_addresses(db, uid)
        serialized = [_serialize_user_address_row(row) for row in rows]
        default_id = ''
        for row in serialized:
            if row.get('is_default'):
                default_id = str(row.get('id') or '').strip()
                break
        if not default_id and serialized:
            default_id = str(serialized[0].get('id') or '').strip()
        return {
            'default_id': default_id or None,
            'addresses': serialized,
        }
    except Exception as e:
        logger.exception('auth_list_addresses failed: %s', e)
        raise HTTPException(status_code=500, detail='Could not load addresses')


@app.put('/auth/addresses', response_model=schemas.UserAddressBookResponse)
def auth_replace_addresses(
    payload: schemas.UserAddressBookSync,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    try:
        uid = int(getattr(current_user, 'id'))
    except Exception:
        raise HTTPException(status_code=401, detail='Invalid user')
    try:
        rows = crud.replace_user_addresses(db, uid, payload)
        serialized = [_serialize_user_address_row(row) for row in rows]
        default_id = ''
        for row in serialized:
            if row.get('is_default'):
                default_id = str(row.get('id') or '').strip()
                break
        if not default_id and serialized:
            default_id = str(serialized[0].get('id') or '').strip()
        return {
            'default_id': default_id or None,
            'addresses': serialized,
        }
    except Exception as e:
        logger.exception('auth_replace_addresses failed: %s', e)
        raise HTTPException(status_code=500, detail='Could not save addresses')

# --- Admin users / staff auth ---
def _normalize_admin_role(value: Optional[str]) -> str:
    role = str(value or '').strip().lower()
    if role not in ('owner', 'admin', 'repartidor'):
        role = 'admin'
    return role


def _normalize_admin_zone(value: Optional[str]) -> Optional[str]:
    try:
        zone = str(value or '').strip()
    except Exception:
        zone = ''
    if not zone:
        return None
    return zone


def _parse_admin_zones(value: Optional[str]) -> List[str]:
    if value is None:
        return []
    text = str(value or '').strip()
    if not text:
        return []
    token = _normalize_region_token(text)
    if token in ('todos', 'todas', 'all', '*'):
        return []
    parts = re.split(r'[;,/|]+', text)
    zones = []
    for part in parts:
        t = _normalize_region_token(part)
        if not t:
            continue
        if t in ('todos', 'todas', 'all', '*'):
            return []
        zones.append(t)
    return zones


def _canonicalize_zones(zone_tokens: List[str]) -> List[str]:
    if not zone_tokens:
        return []
    canonical = []
    seen = set()
    for token in zone_tokens:
        if not token or token in seen:
            continue
        seen.add(token)
        match = None
        for dep in _MENDOZA_DEPARTMENTS:
            if _normalize_region_token(dep) == token:
                match = dep
                break
        canonical.append(match or token)
    return canonical


def _order_matches_zone(order_data: Dict[str, Any], zone_tokens: List[str]) -> bool:
    if not zone_tokens:
        return True
    try:
        snapshot = _extract_order_address_snapshot(order_data if isinstance(order_data, dict) else {})
    except Exception:
        snapshot = {}
    barrio = _normalize_region_token(snapshot.get('barrio'))
    dept = _normalize_region_token(snapshot.get('department'))
    raw_address = _normalize_region_token(snapshot.get('raw_address'))
    postal = _normalize_postal_code(snapshot.get('postal_code'))
    deps_for_postal = _departments_for_postal(postal)
    deps_tokens = { _normalize_region_token(d) for d in deps_for_postal }
    for z in zone_tokens:
        if not z:
            continue
        if dept and dept == z:
            return True
        if z in deps_tokens:
            return True
        if barrio and z in barrio:
            return True
        if raw_address and z in raw_address:
            return True
    return False


def _haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    try:
        r = 6371.0
        phi1 = math.radians(lat1)
        phi2 = math.radians(lat2)
        dphi = math.radians(lat2 - lat1)
        dlambda = math.radians(lon2 - lon1)
        a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
        c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
        return r * c
    except Exception:
        return 0.0


def _optimize_route_order(nodes: List[dict]) -> List[dict]:
    """Nearest neighbor + 2-opt optimization. Nodes must include lat/lon."""
    if len(nodes) <= 2:
        return nodes
    remaining = nodes[:]
    start_idx = 0
    try:
        best_score = float('inf')
        for idx, cand in enumerate(remaining):
            score = 0.0
            for other in remaining:
                if other is cand:
                    continue
                score += _haversine_km(cand['lat'], cand['lon'], other['lat'], other['lon'])
            if score < best_score:
                best_score = score
                start_idx = idx
    except Exception:
        start_idx = 0
    path = [remaining.pop(start_idx)]
    while remaining:
        last = path[-1]
        best_idx = 0
        best_dist = float('inf')
        for idx, cand in enumerate(remaining):
            d = _haversine_km(last['lat'], last['lon'], cand['lat'], cand['lon'])
            if d < best_dist:
                best_dist = d
                best_idx = idx
        path.append(remaining.pop(best_idx))
    # 2-opt improvement
    def path_distance(p):
        total = 0.0
        for i in range(len(p) - 1):
            total += _haversine_km(p[i]['lat'], p[i]['lon'], p[i+1]['lat'], p[i+1]['lon'])
        return total
    improved = True
    iterations = 0
    max_iter = min(600, max(120, len(path) * 6))
    while improved and iterations < max_iter:
        improved = False
        iterations += 1
        for i in range(1, len(path) - 2):
            for j in range(i + 1, len(path) - 1):
                a = path[i - 1]; b = path[i]
                c = path[j]; d = path[j + 1]
                dist_before = _haversine_km(a['lat'], a['lon'], b['lat'], b['lon']) + _haversine_km(c['lat'], c['lon'], d['lat'], d['lon'])
                dist_after = _haversine_km(a['lat'], a['lon'], c['lat'], c['lon']) + _haversine_km(b['lat'], b['lon'], d['lat'], d['lon'])
                if dist_after + 1e-6 < dist_before:
                    path[i:j+1] = reversed(path[i:j+1])
                    improved = True
        if not improved:
            break
    return path


def _optimize_route_order_with_start(nodes: List[dict], start_lat: float, start_lon: float) -> List[dict]:
    """Optimize route order starting from a fixed depot point."""
    if len(nodes) <= 1:
        return nodes
    def build_path(start_idx: int) -> List[dict]:
        remaining = nodes[:]
        start_idx = max(0, min(start_idx, len(remaining) - 1))
        path_local = [remaining.pop(start_idx)]
        while remaining:
            last = path_local[-1]
            best_idx = 0
            best_dist = float('inf')
            for idx, cand in enumerate(remaining):
                d = _haversine_km(last['lat'], last['lon'], cand['lat'], cand['lon'])
                if d < best_dist:
                    best_dist = d
                    best_idx = idx
            path_local.append(remaining.pop(best_idx))
        return path_local

    def path_distance(p):
        if not p:
            return 0.0
        total = _haversine_km(start_lat, start_lon, p[0]['lat'], p[0]['lon'])
        for i in range(len(p) - 1):
            total += _haversine_km(p[i]['lat'], p[i]['lon'], p[i+1]['lat'], p[i+1]['lon'])
        return total

    def improve_path(path_local: List[dict]) -> List[dict]:
        improved = True
        iterations = 0
        max_iter = min(800, max(160, len(path_local) * 8))
        while improved and iterations < max_iter:
            improved = False
            iterations += 1
            n = len(path_local)
            for i in range(0, n - 1):
                for j in range(i + 1, n):
                    a_lat = start_lat if i == 0 else path_local[i - 1]['lat']
                    a_lon = start_lon if i == 0 else path_local[i - 1]['lon']
                    b = path_local[i]
                    c = path_local[j]
                    if j + 1 < n:
                        d = path_local[j + 1]
                        dist_before = _haversine_km(a_lat, a_lon, b['lat'], b['lon']) + _haversine_km(c['lat'], c['lon'], d['lat'], d['lon'])
                        dist_after = _haversine_km(a_lat, a_lon, c['lat'], c['lon']) + _haversine_km(b['lat'], b['lon'], d['lat'], d['lon'])
                    else:
                        dist_before = _haversine_km(a_lat, a_lon, b['lat'], b['lon'])
                        dist_after = _haversine_km(a_lat, a_lon, c['lat'], c['lon'])
                    if dist_after + 1e-6 < dist_before:
                        path_local[i:j + 1] = reversed(path_local[i:j + 1])
                        improved = True
            if not improved:
                break
        return path_local

    # Candidate starts: nearest, farthest, and bearing-sorted
    try:
        nearest_idx = min(
            range(len(nodes)),
            key=lambda i: _haversine_km(start_lat, start_lon, nodes[i]['lat'], nodes[i]['lon'])
        )
    except Exception:
        nearest_idx = 0
    try:
        farthest_idx = max(
            range(len(nodes)),
            key=lambda i: _haversine_km(start_lat, start_lon, nodes[i]['lat'], nodes[i]['lon'])
        )
    except Exception:
        farthest_idx = 0

    candidates: List[List[dict]] = []
    for idx in {nearest_idx, farthest_idx}:
        try:
            candidates.append(improve_path(build_path(idx)))
        except Exception:
            continue

    try:
        bearing_sorted = sorted(
            nodes,
            key=lambda n: math.atan2(n['lon'] - start_lon, n['lat'] - start_lat)
        )
        candidates.append(improve_path(bearing_sorted))
    except Exception:
        pass

    if not candidates:
        return build_path(nearest_idx)
    return min(candidates, key=path_distance)


def _extract_admin_actor_from_request(request: Request) -> Tuple[Optional[int], Optional[str]]:
    try:
        auth = request.headers.get('authorization') or request.headers.get('Authorization')
    except Exception:
        auth = None
    if not auth or not isinstance(auth, str) or not auth.lower().startswith('bearer '):
        return None, None
    token = auth.split(' ', 1)[1]
    try:
        payload = utils.decode_access_token(token)
    except Exception:
        return None, None
    if not payload or payload.get('kind') != 'admin':
        return None, None
    try:
        uid = payload.get('id')
        uid = int(uid) if uid is not None else None
    except Exception:
        uid = None
    uname = payload.get('sub') or payload.get('username')
    return uid, (str(uname).strip() if uname else None)


def _parse_epoch_to_datetime(value: Any) -> Optional[datetime.datetime]:
    try:
        if value is None:
            return None
        ts = float(value)
    except Exception:
        return None
    if ts > 1e12:
        ts = ts / 1000.0
    if ts <= 0:
        return None
    try:
        return datetime.datetime.fromtimestamp(ts, tz=datetime.timezone.utc)
    except Exception:
        return None


def _resolve_order_coords(order: Dict[str, Any], db: Optional[Session] = None) -> Tuple[Optional[float], Optional[float]]:
    """Return (lat, lon) for the order, using stored coords or user address fallback."""
    try:
        snapshot = _extract_order_address_snapshot(order if isinstance(order, dict) else {})
    except Exception:
        snapshot = {}
    lat = snapshot.get('lat')
    lon = snapshot.get('lon')
    if _is_mendoza_point(lat, lon):
        return float(lat), float(lon)

    # Try user default address
    uid = None
    try:
        uid = order.get('user_id')
    except Exception:
        uid = None
    if uid is not None and db is not None:
        try:
            addr = db.query(models.UserAddress).filter(models.UserAddress.user_id == int(uid)).order_by(
                models.UserAddress.is_default.desc(),
                models.UserAddress.updated_at.desc().nullslast(),
                models.UserAddress.created_at.desc(),
            ).first()
            if addr and addr.lat is not None and addr.lon is not None:
                lat = float(addr.lat)
                lon = float(addr.lon)
        except Exception:
            lat = None
            lon = None

    if not _is_mendoza_point(lat, lon):
        try:
            lat, lon = _geocode_order_snapshot(snapshot, db=db)
        except Exception:
            lat = None
            lon = None

    if _is_mendoza_point(lat, lon):
        # Persist to order for future routing
        try:
            oid = order.get('id')
            if oid is not None and db is not None:
                with engine.begin() as conn:
                    if str(oid).isdigit():
                        conn.execute(
                            text("UPDATE orders SET delivery_lat = :lat, delivery_lon = :lon WHERE id = :id"),
                            {'lat': lat, 'lon': lon, 'id': int(oid)},
                        )
                    else:
                        conn.execute(
                            text("UPDATE orders SET delivery_lat = :lat, delivery_lon = :lon WHERE CAST(id AS TEXT) = :id"),
                            {'lat': lat, 'lon': lon, 'id': str(oid)},
                        )
                order['delivery_lat'] = lat
                order['delivery_lon'] = lon
        except Exception:
            pass
        return float(lat), float(lon)
    return None, None


def get_current_admin_user(token: str = Depends(admin_oauth2_scheme)):
    try:
        payload = utils.decode_access_token(token)
    except Exception as e:
        logger.exception('get_current_admin_user: decode_access_token failed: %s', e)
        raise HTTPException(status_code=401, detail='Invalid token')
    if not payload or payload.get('kind') != 'admin':
        raise HTTPException(status_code=401, detail='Invalid token payload')
    username = payload.get('sub') or payload.get('username')
    user_id = payload.get('id')
    if not username and not user_id:
        raise HTTPException(status_code=401, detail='Invalid token payload')
    db = SessionLocal()
    try:
        user = None
        if user_id is not None:
            try:
                user = crud.get_admin_user_by_id(db, int(user_id))
            except Exception:
                user = None
        if not user and username:
            user = crud.get_admin_user_by_username(db, username)
        if not user:
            raise HTTPException(status_code=401, detail='Admin user not found')
        if not getattr(user, 'is_active', True):
            raise HTTPException(status_code=403, detail='Admin user disabled')
        return user
    finally:
        db.close()


def require_admin_owner(current_admin=Depends(get_current_admin_user)):
    if not current_admin or getattr(current_admin, 'role', None) != 'owner':
        raise HTTPException(status_code=403, detail='Owner access required')
    return current_admin


@app.post('/admin/auth/token', response_model=schemas.Token)
async def admin_login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends()):
    def task():
        db = SessionLocal()
        try:
            user = crud.authenticate_admin_user(db, form_data.username, form_data.password)
            return user
        finally:
            db.close()

    try:
        user = await anyio.to_thread.run_sync(task)
    except Exception as e:
        logger.exception('Unexpected error in /admin/auth/token: %s', e)
        raise HTTPException(status_code=500, detail='Server error')

    if not user:
        raise HTTPException(status_code=401, detail='Incorrect credentials')
    access_token = utils.create_access_token({
        "sub": user.username,
        "id": user.id,
        "role": user.role,
        "zone": getattr(user, 'zone', None),
        "kind": "admin",
    })
    logger.info('admin_login_for_access_token: issued token for admin id=%s user=%s', user.id, user.username)
    return {"access_token": access_token, "token_type": "bearer"}


@app.get('/admin/auth/me', response_model=schemas.AdminUserResponse)
def admin_auth_me(current_admin=Depends(get_current_admin_user)):
    return current_admin


@app.get('/admin/users', response_model=List[schemas.AdminUserResponse])
def admin_list_users(current_admin=Depends(get_current_admin_user), db: Session = Depends(get_db)):
    role = str(getattr(current_admin, 'role', '') or '').strip().lower()
    users = crud.list_admin_users(db)
    if role == 'owner':
        return users
    # Admins can see only repartidores to assign zones.
    return [u for u in (users or []) if str(getattr(u, 'role', '') or '').strip().lower() == 'repartidor']


@app.post('/admin/users', response_model=schemas.AdminUserResponse)
def admin_create_user(payload: schemas.AdminUserCreate, current_owner=Depends(require_admin_owner), db: Session = Depends(get_db)):
    username = str(payload.username or '').strip()
    if not username:
        raise HTTPException(status_code=400, detail='Username requerido')
    role = _normalize_admin_role(payload.role)
    if role == 'owner':
        raise HTTPException(status_code=400, detail='No se pueden crear owners adicionales')
    zone = _normalize_admin_zone(getattr(payload, 'zone', None))
    if role == 'repartidor':
        zone_tokens = _parse_admin_zones(zone)
        if not zone_tokens:
            raise HTTPException(status_code=400, detail='Zona requerida para repartidor')
        allowed_tokens = {_normalize_region_token(dep) for dep in _MENDOZA_DEPARTMENTS}
        invalid = [z for z in zone_tokens if z not in allowed_tokens]
        if invalid:
            raise HTTPException(status_code=400, detail='Zona inválida')
        canonical = _canonicalize_zones(zone_tokens)
        zone = ', '.join(canonical)
    try:
        if crud.get_admin_user_by_username(db, username):
            raise HTTPException(status_code=400, detail='Ese usuario ya existe')
    except HTTPException:
        raise
    except Exception:
        pass
    payload.zone = zone
    hashed = utils.hash_password(payload.password)
    try:
        new_user = crud.create_admin_user(db, payload, hashed_password=hashed, created_by=current_owner.username)
        return new_user
    except IntegrityError:
        raise HTTPException(status_code=400, detail='Ese usuario ya existe')
    except Exception as e:
        logger.exception('admin_create_user failed: %s', e)
        raise HTTPException(status_code=500, detail='No se pudo crear el usuario')


@app.delete('/admin/users/{user_key}')
def admin_delete_user(user_key: str, current_owner=Depends(require_admin_owner), db: Session = Depends(get_db)):
    user = None
    try:
        if str(user_key).isdigit():
            user = crud.get_admin_user_by_id(db, int(user_key))
    except Exception:
        user = None
    if not user:
        user = crud.get_admin_user_by_username(db, user_key)
    if not user:
        raise HTTPException(status_code=404, detail='Usuario no encontrado')
    if getattr(user, 'role', None) == 'owner':
        raise HTTPException(status_code=400, detail='No se puede eliminar el owner')
    try:
        ok = crud.delete_admin_user(db, int(getattr(user, 'id')))
        if not ok:
            raise HTTPException(status_code=404, detail='Usuario no encontrado')
        return {'ok': True}
    except HTTPException:
        raise
    except Exception as e:
        logger.exception('admin_delete_user failed: %s', e)
        raise HTTPException(status_code=500, detail='No se pudo eliminar el usuario')


@app.patch('/admin/users/{user_key}', response_model=schemas.AdminUserResponse)
def admin_update_user(
    user_key: str,
    payload: schemas.AdminUserUpdate,
    current_admin=Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    role = str(getattr(current_admin, 'role', '') or '').strip().lower()
    if role not in ('owner', 'admin'):
        raise HTTPException(status_code=403, detail='No autorizado')
    user = None
    try:
        if str(user_key).isdigit():
            user = crud.get_admin_user_by_id(db, int(user_key))
    except Exception:
        user = None
    if not user:
        user = crud.get_admin_user_by_username(db, user_key)
    if not user:
        raise HTTPException(status_code=404, detail='Usuario no encontrado')
    target_role = str(getattr(user, 'role', '') or '').strip().lower()
    if target_role != 'repartidor':
        raise HTTPException(status_code=403, detail='Solo se puede asignar zona a repartidores')
    zone = _normalize_admin_zone(getattr(payload, 'zone', None))
    zone_tokens = _parse_admin_zones(zone)
    if not zone_tokens:
        raise HTTPException(status_code=400, detail='Zona requerida')
    allowed_tokens = {_normalize_region_token(dep) for dep in _MENDOZA_DEPARTMENTS}
    invalid = [z for z in zone_tokens if z not in allowed_tokens]
    if invalid:
        raise HTTPException(status_code=400, detail='Zona inválida')
    canonical = _canonicalize_zones(zone_tokens)
    zone = ', '.join(canonical)
    try:
        updated = crud.update_admin_user_zone(db, int(getattr(user, 'id')), zone)
        if not updated:
            raise HTTPException(status_code=404, detail='Usuario no encontrado')
        return updated
    except HTTPException:
        raise
    except Exception as e:
        logger.exception('admin_update_user failed: %s', e)
        raise HTTPException(status_code=500, detail='No se pudo actualizar el usuario')

@app.post("/products")
async def create_product(payload: schemas.ProductCreate, request: Request, background_tasks: BackgroundTasks):
    """Create a product - no response validation, just raw JSON."""
    # Log incoming payload (helps debug server vs validation failures)
    try:
        logger.info('POST /products called with payload: %s', jsonable_encoder(payload))
    except Exception:
        logger.exception('Failed to log payload for POST /products')

    actor = None
    try:
        actor = (request.headers.get('x-actor') or request.headers.get('X-Actor') or '').strip() or None
    except Exception:
        actor = None

    def task():
        db = SessionLocal()
        try:
            prod = crud.create_product(db, payload, actor=actor)
            if not prod:
                return None
            # Convert safely to dict whether prod is a dict or object
            if isinstance(prod, dict):
                getv = prod.get
            else:
                getv = lambda k, d=None: getattr(prod, k, d)
            result = {
                'id': getv('id', None),
                'code': getv('code', None),
                'name': getv('name', ''),
                'brand': getv('brand', None),
                'price': getv('price', 0),
                'price_retail': getv('price_retail', None),
                'cost': getv('cost', None),
                'description': getv('description', ''),
                'category': getv('category', ''),
                'image_url': getv('image_url', ''),
                'active': bool(getv('active', True)),
                'created_at': getv('created_at', None),
                'updated_at': getv('updated_at', None),
                'stock': int(getv('stock', 0) or 0),
                'min_stock': int(getv('min_stock', 0) or 0),
                'stock_kg': float(getv('stock_kg', getv('stock', 0.0)) or 0.0),
                'kg_per_unit': float(getv('kg_per_unit', 1.0) or 1.0),
                'discount': float(getv('discount', 0.0) or 0.0),
                'sale_unit': str(getv('sale_unit', 'unit') or 'unit')
            }
            return result
        finally:
            try:
                db.rollback()
            except:
                pass
            try:
                db.close()
            except:
                pass
                pass

    try:
        result = await anyio.to_thread.run_sync(task)
        if not result:
            raise HTTPException(status_code=500, detail='Product creation failed')
        # Normalize result to plain dict
        if not isinstance(result, dict):
            try:
                result = {k: getattr(result, k) for k in ('id','code','name','price','price_retail','description','category','image_url','active','stock','stock_kg','kg_per_unit','discount','sale_unit') if hasattr(result, k)}
            except Exception:
                result = dict(result.__dict__) if hasattr(result, '__dict__') else dict(result)
        try:
            await push_event({"action": "created", "product": {"id": result.get('id')}})
        except:
            pass
        
        try:
            await anyio.to_thread.run_sync(write_catalog_snapshot)
        except:
            pass

        try:
            if _auto_image_enabled():
                image_url = str(result.get('image_url') or '').strip()
                if not image_url and result.get('id'):
                    background_tasks.add_task(
                        _auto_fetch_and_set_image,
                        result.get('id'),
                        result.get('name'),
                        result.get('code'),
                        True,
                    )
        except Exception:
            pass

        _invalidate_products_cache()
        
        return result
    except HTTPException:
        raise
    except Exception as e:
        try:
            tb = traceback.format_exc()
            base = os.path.dirname(os.path.dirname(__file__))
            logpath = os.path.join(base, 'server_log.txt')
            with open(logpath, 'a', encoding='utf-8') as f:
                f.write(f"{datetime.datetime.utcnow().isoformat()} - POST /products exception: {str(e)[:400]}\n")
                f.write(tb + "\n\n")
        except Exception:
            pass
        logger.exception('POST /products: %s', e)
        raise HTTPException(status_code=500, detail=str(e)[:100])

@app.get("/products", response_model=List[schemas.ProductResponse])
def list_products(
    skip: int = 0,
    limit: int = 100,
    q: Optional[str] = None,
    category: Optional[str] = None,
    active: Optional[bool] = None,
    sort: Optional[str] = None,
    db: Session = Depends(get_db),
):
    limit_max = int(os.environ.get('PRODUCTS_LIST_LIMIT_MAX') or 5000)
    limit = max(1, min(int(limit or 100), limit_max))
    skip = max(0, int(skip or 0))
    cache_key = _products_cache_key(skip, limit, q, category, active, sort)
    cached = _cache_get(PRODUCTS_CACHE, cache_key, PRODUCTS_CACHE_TTL)
    if cached is not None:
        return cached
    try:
        items = crud.get_products(db, skip, limit, q, category, active, sort)
        serialized = _serialize_products(items)
        if limit <= PRODUCTS_CACHE_LIMIT_MAX:
            _cache_set(PRODUCTS_CACHE, cache_key, serialized, PRODUCTS_CACHE_MAX)
        return serialized
    except Exception:
        logger.exception('list_products ORM failed, attempting raw fallback')
        # Try raw fallback select to avoid failing when DB lacks mapped columns
        try:
            try:
                bind = db.get_bind()
                insp = inspect(bind)
                existing = {c['name'] for c in insp.get_columns('products')}
            except Exception:
                existing = set()
            cols = ['id','name','price','description','category','image_url','created_at','updated_at','active']
            if 'code' in existing: cols.append('code')
            if 'brand' in existing: cols.append('brand')
            if 'price_retail' in existing: cols.append('price_retail')
            if 'cost' in existing: cols.append('cost')
            if 'stock' in existing: cols.append('stock')
            if 'min_stock' in existing: cols.append('min_stock')
            if 'stock_kg' in existing: cols.append('stock_kg')
            if 'kg_per_unit' in existing: cols.append('kg_per_unit')
            if 'discount' in existing: cols.append('discount')
            if 'sale_unit' in existing: cols.append('sale_unit')
            where = []
            params = {'skip': skip, 'limit': limit}
            if q:
                match_parts = ["LOWER(COALESCE(name, '')) LIKE :q", "LOWER(COALESCE(description, '')) LIKE :q"]
                if 'code' in existing:
                    match_parts.append("LOWER(COALESCE(code, '')) LIKE :q")
                where.append('(' + ' OR '.join(match_parts) + ')'); params['q'] = f"%{q.lower()}%"
            if category:
                where.append('category = :category'); params['category'] = category
            if active is not None and 'active' in existing:
                where.append('active = :active'); params['active'] = bool(active)
            where_clause = (' WHERE ' + ' AND '.join(where)) if where else ''
            order_clause = ''
            if sort == 'price_asc': order_clause = ' ORDER BY price ASC'
            elif sort == 'price_desc': order_clause = ' ORDER BY price DESC'
            elif sort == 'name_asc': order_clause = ' ORDER BY name ASC'
            elif sort == 'name_desc': order_clause = ' ORDER BY name DESC'
            cols_sql = ', '.join(cols)
            sql = f"SELECT {cols_sql} FROM products{where_clause}{order_clause} LIMIT :limit OFFSET :skip"
            rows = crud._safe_execute_fetchall(db, sql, params)
            result = []
            for row in rows:
                objd = {cols[i]: row[i] for i in range(len(cols))}
                result.append(objd)
            if limit <= PRODUCTS_CACHE_LIMIT_MAX:
                _cache_set(PRODUCTS_CACHE, cache_key, result, PRODUCTS_CACHE_MAX)
            return result
        except Exception:
            logger.exception('list_products raw fallback failed')
            return []

@app.get("/products/paged", response_model=schemas.PagedProductsResponse)
def list_products_paged(
    skip: int = 0,
    limit: int = 50,
    q: Optional[str] = None,
    category: Optional[str] = None,
    active: Optional[bool] = None,
    sort: Optional[str] = None,
    db: Session = Depends(get_db),
):
    limit_max = int(os.environ.get('PRODUCTS_LIST_LIMIT_MAX') or 5000)
    limit = max(1, min(int(limit or 50), min(200, limit_max)))
    skip = max(0, int(skip or 0))
    count_key = (_normalize_cache_text(q), _normalize_cache_text(category), None if active is None else bool(active))
    total = _cache_get(PRODUCTS_COUNT_CACHE, count_key, PRODUCTS_COUNT_CACHE_TTL)
    if total is None:
        total = crud.count_products(db, q=q, category=category, active=active)
        _cache_set(PRODUCTS_COUNT_CACHE, count_key, int(total or 0), PRODUCTS_CACHE_MAX)
    items_key = _products_cache_key(skip, limit, q, category, active, sort)
    cached_items = _cache_get(PRODUCTS_CACHE, items_key, PRODUCTS_CACHE_TTL)
    if cached_items is None:
        items_raw = crud.get_products(db, skip=skip, limit=limit, q=q, category=category, active=active, sort=sort)
        cached_items = _serialize_products(items_raw)
        _cache_set(PRODUCTS_CACHE, items_key, cached_items, PRODUCTS_CACHE_MAX)
    items = cached_items
    return { 'total': int(total or 0), 'skip': skip, 'limit': limit, 'items': items }


def _normalize_import_header(value: Any) -> str:
    try:
        raw = str(value or '').strip().lower()
    except Exception:
        return ''
    if not raw:
        return ''
    try:
        raw = unicodedata.normalize('NFKD', raw)
        raw = ''.join(ch for ch in raw if not unicodedata.combining(ch))
    except Exception:
        pass
    raw = re.sub(r'[^a-z0-9]+', '_', raw)
    raw = raw.strip('_')
    return raw


_IMPORT_HEADER_ALIASES = {
    'code': {'code', 'codigo', 'código', 'sku', 'cod', 'codigo_sku', 'codigo_de_producto'},
    'name': {'name', 'nombre', 'producto', 'producto_nombre', 'descripcion_producto'},
    'brand': {'brand', 'marca'},
    'description': {'description', 'descripcion', 'detalle', 'desc'},
    'category': {'category', 'categoria', 'rubro', 'familia', 'linea'},
    'price': {'price', 'precio', 'precio_mayorista', 'mayorista', 'precio_may', 'precio_wholesale', 'precio_mayoreo'},
    'price_retail': {'price_retail', 'precio_minorista', 'minorista', 'precio_minor', 'precio_retail', 'precio_venta', 'precio_menudeo'},
    'cost': {'cost', 'costo', 'coste', 'precio_costo', 'costo_unitario'},
    'discount': {'discount', 'descuento', 'descuento_pct', 'descuento_%', 'dto'},
    'stock': {'stock', 'cantidad', 'existencias', 'stock_actual'},
    'min_stock': {'min_stock', 'stock_min', 'stock_minimo', 'stock_mínimo', 'minimo', 'mínimo'},
    'stock_kg': {'stock_kg', 'stock_kilos', 'kg_stock', 'stock_kilogramos'},
    'kg_per_unit': {'kg_per_unit', 'kg_unidad', 'kg_por_unidad', 'kg_x_unidad', 'unidad_kg', 'peso_unidad'},
    'sale_unit': {'sale_unit', 'unidad', 'unidad_venta', 'tipo_venta', 'unidad_de_venta'},
    'active': {'active', 'activo'},
    'image_url': {'image', 'imagen', 'image_url', 'imagen_url', 'url', 'foto', 'picture', 'img', 'url_imagen', 'url_de_imagen', 'imagen_link', 'image_link', 'foto_url', 'img_url', 'link_imagen'},
}

_IMPORT_HEADER_MAP = {alias: key for key, aliases in _IMPORT_HEADER_ALIASES.items() for alias in aliases}


def _auto_image_enabled() -> bool:
    return _is_truthy(os.environ.get('AUTO_IMAGE_FETCH') or os.environ.get('AUTO_IMAGE_SEARCH'))


def _auto_image_now_iso() -> str:
    try:
        return datetime.datetime.utcnow().replace(tzinfo=datetime.timezone.utc).isoformat().replace('+00:00', 'Z')
    except Exception:
        return datetime.datetime.utcnow().isoformat()


def _queue_auto_image_progress(total: int, source: Optional[str] = None) -> None:
    now = _auto_image_now_iso()
    with AUTO_IMAGE_PROGRESS_LOCK:
        AUTO_IMAGE_PROGRESS.update({
            'status': 'queued',
            'total': int(total or 0),
            'processed': 0,
            'attached': 0,
            'started_at': None,
            'finished_at': None,
            'last_update': now,
            'source': source,
            'last_error': None,
        })


def _start_auto_image_progress(total: int, source: Optional[str] = None) -> None:
    now = _auto_image_now_iso()
    with AUTO_IMAGE_PROGRESS_LOCK:
        AUTO_IMAGE_PROGRESS.update({
            'status': 'running',
            'total': int(total or 0),
            'processed': 0,
            'attached': 0,
            'started_at': now,
            'finished_at': None,
            'last_update': now,
            'source': source,
            'last_error': None,
        })


def _advance_auto_image_progress(processed_inc: int = 0, attached_inc: int = 0) -> None:
    now = _auto_image_now_iso()
    with AUTO_IMAGE_PROGRESS_LOCK:
        AUTO_IMAGE_PROGRESS['processed'] = int(AUTO_IMAGE_PROGRESS.get('processed') or 0) + int(processed_inc or 0)
        AUTO_IMAGE_PROGRESS['attached'] = int(AUTO_IMAGE_PROGRESS.get('attached') or 0) + int(attached_inc or 0)
        AUTO_IMAGE_PROGRESS['last_update'] = now


def _finish_auto_image_progress(status: str = 'done', error: Optional[str] = None) -> None:
    now = _auto_image_now_iso()
    with AUTO_IMAGE_PROGRESS_LOCK:
        AUTO_IMAGE_PROGRESS['status'] = status
        AUTO_IMAGE_PROGRESS['finished_at'] = now
        AUTO_IMAGE_PROGRESS['last_update'] = now
        if error:
            AUTO_IMAGE_PROGRESS['last_error'] = error


def _get_auto_image_progress_snapshot() -> Dict[str, Any]:
    with AUTO_IMAGE_PROGRESS_LOCK:
        return dict(AUTO_IMAGE_PROGRESS)


def _fetch_pexels_photo(query: str) -> Optional[Dict[str, Any]]:
    api_key = str(os.environ.get('PEXELS_API_KEY') or '').strip()
    if not api_key:
        return None
    try:
        params = {'query': query, 'per_page': 1}
        resp = httpx.get(
            'https://api.pexels.com/v1/search',
            params=params,
            headers={'Authorization': api_key},
            timeout=10,
        )
        if resp.status_code >= 400:
            logger.warning('Pexels search failed: %s %s', resp.status_code, resp.text[:200])
            return None
        payload = resp.json() or {}
        photos = payload.get('photos') or []
        if not photos:
            return None
        return photos[0]
    except Exception as e:
        logger.warning('Pexels search error: %s', e)
        return None


def _download_image_bytes(url: str, max_bytes: int) -> Optional[Tuple[bytes, str, str]]:
    try:
        with httpx.stream('GET', url, timeout=15, follow_redirects=True) as resp:
            if resp.status_code >= 400:
                return None
            content_type = str(resp.headers.get('content-type') or 'image/jpeg')
            if not content_type.startswith('image/'):
                return None
            data = bytearray()
            for chunk in resp.iter_bytes():
                if not chunk:
                    continue
                data.extend(chunk)
                if len(data) > max_bytes:
                    logger.warning('Image download exceeded max bytes: %s', max_bytes)
                    return None
            filename = os.path.basename(str(url).split('?')[0] or '')
            return (bytes(data), content_type, filename or 'image')
    except Exception as e:
        logger.warning('Image download failed: %s', e)
        return None


def _auto_fetch_and_set_image(
    product_id: int,
    name: Optional[str] = None,
    code: Optional[str] = None,
    write_snapshot: bool = True,
) -> Optional[Dict[str, Any]]:
    if not _auto_image_enabled():
        return None
    query = str(name or code or '').strip()
    if not query:
        return None

    provider = str(os.environ.get('IMAGE_SEARCH_PROVIDER') or 'pexels').strip().lower()
    photo = None
    if provider == 'pexels':
        photo = _fetch_pexels_photo(query)
    else:
        logger.warning('Unsupported image provider: %s', provider)
        return None
    if not photo:
        return None

    src = photo.get('src') or {}
    size_pref = str(os.environ.get('IMAGE_FETCH_SIZE') or 'medium').strip().lower()
    candidate = src.get(size_pref) or src.get('large') or src.get('medium') or src.get('portrait') or src.get('original') or src.get('small')
    if not candidate:
        return None

    max_bytes = int(os.environ.get('IMAGE_FETCH_MAX_BYTES') or 4_000_000)
    downloaded = _download_image_bytes(candidate, max_bytes=max_bytes)
    if not downloaded:
        return None
    content, mime, filename = downloaded

    db = SessionLocal()
    try:
        prod = db.query(models.Product).filter(models.Product.id == int(product_id)).first()
        if not prod:
            return None
        existing_url = str(getattr(prod, 'image_url', '') or '').strip()
        if existing_url:
            return None
        img = models.Image(data=content, mime=mime, filename=filename)
        db.add(img)
        db.commit()
        db.refresh(img)
        prod.image_url = f"/images/{img.id}"
        db.add(prod)
        db.commit()
        try:
            _invalidate_products_cache()
        except Exception:
            pass
        if write_snapshot:
            try:
                write_catalog_snapshot()
            except Exception:
                pass
        return {
            'image_url': prod.image_url,
            'provider': provider,
            'photographer': photo.get('photographer'),
            'photographer_url': photo.get('photographer_url'),
            'source_url': photo.get('url'),
        }
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        logger.warning('Auto image attach failed: %s', e)
        return None
    finally:
        db.close()


def _auto_fetch_images_for_products(
    items: List[Dict[str, Any]],
    write_snapshot: bool = True,
    source: Optional[str] = None,
) -> Dict[str, Any]:
    if not _auto_image_enabled():
        return {'requested': 0, 'processed': 0, 'attached': 0, 'results': []}
    if not items:
        return {'requested': 0, 'processed': 0, 'attached': 0, 'results': []}

    try:
        max_items = int(os.environ.get('IMAGE_FETCH_BATCH_MAX') or 50)
    except Exception:
        max_items = 50
    if max_items <= 0:
        max_items = len(items)
    try:
        sleep_ms = float(os.environ.get('IMAGE_FETCH_SLEEP_MS') or 0)
    except Exception:
        sleep_ms = 0.0

    candidates: List[Dict[str, Any]] = []
    for item in items:
        try:
            pid = item.get('id') if isinstance(item, dict) else getattr(item, 'id', None)
        except Exception:
            pid = None
        if not pid:
            continue
        try:
            existing_url = str(item.get('image_url') or '').strip()
        except Exception:
            existing_url = ''
        if existing_url:
            continue
        candidates.append(item)
    if max_items and max_items > 0:
        candidates = candidates[:max_items]

    total = len(candidates)
    _start_auto_image_progress(total, source=source)
    if total == 0:
        _finish_auto_image_progress('done')
        return {'requested': len(items), 'processed': 0, 'attached': 0, 'results': []}

    results: List[Dict[str, Any]] = []
    processed = 0
    try:
        for item in candidates:
            try:
                pid = item.get('id') if isinstance(item, dict) else getattr(item, 'id', None)
            except Exception:
                pid = None
            if not pid:
                continue
            try:
                name = item.get('name') if isinstance(item, dict) else getattr(item, 'name', None)
            except Exception:
                name = None
            try:
                code = item.get('code') if isinstance(item, dict) else getattr(item, 'code', None)
            except Exception:
                code = None
            res = _auto_fetch_and_set_image(pid, name=name, code=code, write_snapshot=False)
            processed += 1
            if res:
                results.append({'id': pid, **res})
            _advance_auto_image_progress(1, 1 if res else 0)
            if sleep_ms > 0:
                try:
                    time.sleep(sleep_ms / 1000.0)
                except Exception:
                    pass
        _finish_auto_image_progress('done')
    except Exception as e:
        logger.warning('Auto image batch failed: %s', e)
        _finish_auto_image_progress('error', str(e)[:200])

    if write_snapshot and results:
        try:
            write_catalog_snapshot()
        except Exception:
            pass
    if results:
        try:
            _invalidate_products_cache()
        except Exception:
            pass
    return {
        'requested': len(items),
        'processed': processed,
        'attached': len(results),
        'results': results,
    }

def _normalize_number_text(value: Any) -> str:
    try:
        raw = str(value or '').strip()
    except Exception:
        return ''
    if not raw:
        return ''
    try:
        raw = re.sub(r'[^0-9,.\-]+', '', raw)
    except Exception:
        pass
    if not raw:
        return ''
    if ',' in raw and '.' in raw:
        if raw.rfind(',') > raw.rfind('.'):
            raw = raw.replace('.', '').replace(',', '.')
        else:
            raw = raw.replace(',', '')
    elif ',' in raw:
        parts = raw.split(',')
        if len(parts) > 2:
            raw = ''.join(parts)
        else:
            raw = raw.replace(',', '.')
    elif '.' in raw:
        if raw.count('.') > 1:
            raw = raw.replace('.', '')
    return raw


def _coerce_int(value: Any) -> Optional[int]:
    try:
        if value is None:
            return None
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, (int, float)):
            if not (value == value):  # NaN check
                return None
            return int(round(value))
        raw = _normalize_number_text(value)
        if raw == '':
            return None
        num = float(raw)
        if not (num == num):
            return None
        return int(round(num))
    except Exception:
        return None


def _coerce_float(value: Any) -> Optional[float]:
    try:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            if not (value == value):
                return None
            return float(value)
        raw = _normalize_number_text(value)
        if raw == '':
            return None
        num = float(raw)
        if not (num == num):
            return None
        return float(num)
    except Exception:
        return None


def _coerce_bool(value: Any) -> Optional[bool]:
    try:
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        raw = str(value).strip().lower()
        if raw in {'1', 'true', 'si', 'sí', 'yes', 'y', 'on'}:
            return True
        if raw in {'0', 'false', 'no', 'off'}:
            return False
    except Exception:
        pass
    return None


def _normalize_sale_unit(value: Any) -> Optional[str]:
    try:
        raw = str(value or '').strip().lower()
    except Exception:
        return None
    if not raw:
        return None
    if raw in {'kg', 'kilo', 'kilos', 'kilogram', 'kilograms', 'kilogramo', 'kilogramos'}:
        return 'kg'
    return 'unit'


def _parse_import_rows_from_excel(content: bytes, filename: str) -> List[Dict[str, Any]]:
    lower_name = str(filename or '').lower()
    if lower_name.endswith('.csv'):
        text_content = content.decode('utf-8', errors='ignore')
        try:
            sample = text_content[:2048]
            dialect = csv.Sniffer().sniff(sample, delimiters=';,')
        except Exception:
            dialect = csv.excel
        reader = csv.reader(StringIO(text_content), dialect)
        rows = list(reader)
    else:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as e:
            raise RuntimeError('openpyxl_required') from e
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
        ws = wb.active
        rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]

    if not rows:
        return []
    header_row = rows[0]
    headers = [_normalize_import_header(h) for h in header_row]
    mapped = []
    for h in headers:
        mapped.append(_IMPORT_HEADER_MAP.get(h, h))

    out: List[Dict[str, Any]] = []
    for row in rows[1:]:
        if not row:
            continue
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        entry: Dict[str, Any] = {}
        for idx, key in enumerate(mapped):
            if not key:
                continue
            if idx >= len(row):
                continue
            entry[key] = row[idx]
        out.append(entry)
    return out


def _normalize_lookup_token(value: Any) -> str:
    try:
        raw = str(value or '').strip().lower()
    except Exception:
        return ''
    if not raw:
        return ''
    try:
        raw = unicodedata.normalize('NFKD', raw)
        raw = ''.join(ch for ch in raw if not unicodedata.combining(ch))
    except Exception:
        pass
    raw = re.sub(r'[^a-z0-9]+', '', raw)
    return raw


def _normalize_import_image_url(value: Any) -> Optional[str]:
    if value is None:
        return None
    try:
        raw = str(value).strip()
    except Exception:
        return None
    if not raw:
        return None
    # Handle accidental formula-style cells
    if raw.startswith('='):
        raw = raw.lstrip('=').strip()
    if not raw:
        return None
    lower = raw.lower()
    if lower.startswith('http://') or lower.startswith('https://'):
        return raw
    if raw.startswith('//'):
        return 'https:' + raw
    # Normalize uploads paths
    if re.match(r'^/?uploads/', raw, re.IGNORECASE):
        return '/' + raw.lstrip('/')
    # Keep bare filenames (e.g. "producto.jpg") as-is so we can resolve in DB
    if re.match(r'^[^/]+\\.(png|jpe?g|webp|gif|svg)$', lower):
        return raw
    # Treat "www.example.com/..." as https
    if lower.startswith('www.'):
        return 'https://' + raw
    # If it looks like a bare domain, add scheme
    if re.match(r'^[a-z0-9.-]+\\.[a-z]{2,}(/|$)', lower):
        return 'https://' + raw
    return raw


def _find_image_by_filename(db: Session, name_text: Optional[str], allow_contains: bool = False):
    try:
        raw = str(name_text or '').strip()
    except Exception:
        return None
    if not raw:
        return None
    try:
        name_lower = raw.lower()
    except Exception:
        name_lower = raw
    try:
        # Exact filename match (case-insensitive)
        img = db.query(models.Image).filter(func.lower(models.Image.filename) == name_lower).order_by(models.Image.id.desc()).first()
        if img:
            return img
    except Exception:
        img = None
    try:
        if '.' not in raw:
            img = db.query(models.Image).filter(func.lower(models.Image.filename).like(name_lower + '.%')).order_by(models.Image.id.desc()).first()
            if img:
                return img
    except Exception:
        img = None
    if allow_contains:
        try:
            img = db.query(models.Image).filter(func.lower(models.Image.filename).like('%' + name_lower + '%')).order_by(models.Image.id.desc()).first()
            if img:
                return img
        except Exception:
            return None
    return None


def _resolve_import_image_url(db: Session, code: Optional[str], name: Optional[str], provided: Optional[str]) -> Optional[str]:
    if provided is not None:
        normalized = _normalize_import_image_url(provided)
        if normalized:
            # If it's a filename or relative token, try to resolve to stored image
            if not re.match(r'^https?://', normalized, re.IGNORECASE) and not normalized.startswith('/'):
                try:
                    img = _find_image_by_filename(db, normalized, allow_contains=True)
                    if img:
                        return f"/images/{img.id}"
                except Exception:
                    pass
            return normalized

    try:
        code_text = str(code or '').strip()
    except Exception:
        code_text = ''
    try:
        name_text = str(name or '').strip()
    except Exception:
        name_text = ''
    code_lower = code_text.lower()
    name_lower = name_text.lower()
    code_norm = _normalize_lookup_token(code_text)
    name_norm = _normalize_lookup_token(name_text)

    try:
        if code_lower:
            prod = db.query(models.Product).filter(func.lower(func.trim(models.Product.code)) == code_lower).first()
            if prod and getattr(prod, 'image_url', None):
                return prod.image_url
    except Exception:
        pass
    try:
        if name_lower:
            prod = db.query(models.Product).filter(func.lower(func.trim(models.Product.name)) == name_lower).first()
            if prod and getattr(prod, 'image_url', None):
                return prod.image_url
    except Exception:
        pass

    # Try exact/contains filename matches by code or name (case-insensitive, preserving spaces)
    try:
        if code_text:
            img = _find_image_by_filename(db, code_text, allow_contains=True)
            if img:
                return f"/images/{img.id}"
    except Exception:
        pass
    try:
        if name_text:
            img = _find_image_by_filename(db, name_text, allow_contains=True)
            if img:
                return f"/images/{img.id}"
    except Exception:
        pass

    try:
        if code_norm:
            img = db.query(models.Image).filter(models.Image.filename.ilike(f"%{code_norm}%")).order_by(models.Image.id.desc()).first()
            if img:
                return f"/images/{img.id}"
    except Exception:
        pass

    try:
        if name_norm:
            token = name_norm[:18]
            if token:
                img = db.query(models.Image).filter(models.Image.filename.ilike(f"%{token}%")).order_by(models.Image.id.desc()).first()
                if img:
                    return f"/images/{img.id}"
    except Exception:
        pass
    return None


@app.get("/products/duplicates")
def list_product_duplicates(limit: int = 200, db: Session = Depends(get_db)):
    return crud.list_duplicate_product_codes(db, limit=limit)


@app.patch("/products/bulk")
async def bulk_update_products(payload: List[schemas.ProductBulkUpdateItem], request: Request):
    actor = None
    try:
        actor = (request.headers.get('x-actor') or request.headers.get('X-Actor') or '').strip() or None
    except Exception:
        actor = None

    def task():
        db = SessionLocal()
        try:
            return crud.bulk_update_products(db, payload or [], actor=actor)
        finally:
            try:
                db.close()
            except Exception:
                pass

    result = await anyio.to_thread.run_sync(task)
    try:
        await push_event({"action": "bulk_updated", "updated": int((result or {}).get('updated') or 0)})
    except Exception:
        pass
    try:
        await anyio.to_thread.run_sync(write_catalog_snapshot)
    except Exception:
        pass
    _invalidate_products_cache()
    return result or {'updated': 0, 'results': []}


@app.post("/products/import-excel")
async def import_products_excel(
    file: UploadFile = File(...),
    request: Request = None,
    background_tasks: BackgroundTasks = None,
):
    actor = None
    try:
        actor = (request.headers.get('x-actor') or request.headers.get('X-Actor') or '').strip() or None
    except Exception:
        actor = None

    content = await file.read()
    filename = getattr(file, 'filename', '') or ''

    def task():
        db = SessionLocal()
        created = 0
        updated = 0
        skipped = 0
        missing_name = 0
        duplicates = []
        errors = []
        created_ids = []
        updated_ids = []
        created_products = []
        try:
            rows = _parse_import_rows_from_excel(content, filename)
        except RuntimeError as e:
            if str(e) == 'openpyxl_required':
                raise HTTPException(status_code=500, detail='Falta instalar openpyxl para leer Excel')
            raise
        except Exception:
            raise HTTPException(status_code=400, detail='No se pudo leer el archivo Excel')

        seen_codes = set()
        for idx, row in enumerate(rows, start=2):
            try:
                name = row.get('name') or row.get('producto') or row.get('descripcion_producto')
                if name is not None:
                    name = str(name).strip()
                if not name:
                    missing_name += 1
                    skipped += 1
                    continue

                code_val = row.get('code') or row.get('sku') or row.get('codigo')
                code = None
                if code_val is not None:
                    code = str(code_val).strip()
                    if code == '':
                        code = None

                code_key = str(code or '').strip().lower()
                if code_key:
                    if code_key in seen_codes:
                        skipped += 1
                        duplicates.append({'row': idx, 'code': code, 'reason': 'duplicado_en_archivo'})
                        continue
                    seen_codes.add(code_key)

                brand = row.get('brand')
                if brand is not None:
                    brand = str(brand).strip()
                    if brand == '':
                        brand = None
                description = row.get('description')
                if description is not None:
                    description = str(description).strip()
                    if description == '':
                        description = None
                category = row.get('category')
                if category is not None:
                    category = str(category).strip()
                    if category == '':
                        category = None

                stock = _coerce_int(row.get('stock'))
                min_stock = _coerce_int(row.get('min_stock'))
                stock_kg = _coerce_float(row.get('stock_kg'))
                kg_per_unit = _coerce_float(row.get('kg_per_unit'))
                sale_unit = _normalize_sale_unit(row.get('sale_unit'))
                active = _coerce_bool(row.get('active'))

                price = _coerce_float(row.get('price'))
                price_retail = _coerce_float(row.get('price_retail'))
                cost = _coerce_float(row.get('cost'))
                discount = _coerce_float(row.get('discount'))
                image_url = _resolve_import_image_url(db, code, name, row.get('image_url'))

                existing = None
                try:
                    if code:
                        code_lower = str(code).strip().lower()
                        existing = db.query(models.Product).filter(func.lower(func.trim(models.Product.code)) == code_lower).first()
                except Exception:
                    existing = None
                if not existing:
                    try:
                        name_lower = str(name or '').strip().lower()
                    except Exception:
                        name_lower = ''
                    if name_lower:
                        try:
                            matches = db.query(models.Product).filter(func.lower(func.trim(models.Product.name)) == name_lower).limit(2).all()
                            if matches and len(matches) == 1:
                                existing = matches[0]
                            elif matches and len(matches) > 1:
                                skipped += 1
                                errors.append({'row': idx, 'error': 'nombre_duplicado'})
                                continue
                        except Exception:
                            existing = None

                if existing:
                    updates = {}
                    if code:
                        try:
                            current_code = str(getattr(existing, 'code', '') or '').strip()
                        except Exception:
                            current_code = ''
                        if current_code != code:
                            updates['code'] = code
                    if name:
                        try:
                            current_name = str(getattr(existing, 'name', '') or '').strip()
                        except Exception:
                            current_name = ''
                        if current_name != name:
                            updates['name'] = name
                    if brand is not None:
                        updates['brand'] = brand
                    if description is not None:
                        updates['description'] = description
                    if category is not None:
                        updates['category'] = category
                    if price is not None:
                        updates['price'] = float(price)
                    if price_retail is not None:
                        updates['price_retail'] = float(price_retail)
                    if cost is not None:
                        updates['cost'] = float(cost)
                    if discount is not None:
                        updates['discount'] = float(discount)
                    if stock is not None:
                        updates['stock'] = max(0, int(stock))
                    if min_stock is not None:
                        updates['min_stock'] = max(0, int(min_stock))
                    if stock_kg is not None:
                        updates['stock_kg'] = float(stock_kg)
                    if kg_per_unit is not None:
                        updates['kg_per_unit'] = float(kg_per_unit)
                    if sale_unit:
                        updates['sale_unit'] = sale_unit
                    if active is not None:
                        updates['active'] = bool(active)
                    if image_url is not None:
                        try:
                            image_url = str(image_url).strip()
                        except Exception:
                            image_url = image_url
                        if image_url:
                            updates['image_url'] = image_url

                    if not updates:
                        skipped += 1
                        continue
                    try:
                        upd_payload = schemas.ProductUpdate(**updates)
                        prod = crud.update_product(db, int(getattr(existing, 'id', 0) or 0), upd_payload, actor=actor, action='import-excel')
                        updated += 1
                        try:
                            prod_id = int(prod.get('id') if isinstance(prod, dict) else getattr(prod, 'id', None))
                            if prod_id:
                                updated_ids.append(prod_id)
                        except Exception:
                            pass
                        continue
                    except Exception as e_upd:
                        skipped += 1
                        errors.append({'row': idx, 'error': str(e_upd)[:200]})
                        continue

                payload = schemas.ProductCreate(
                    code=code,
                    name=name,
                    price=0.0 if price is None else float(price),
                    price_retail=None if price_retail is None else float(price_retail),
                    cost=None if cost is None else float(cost),
                    description=description or '',
                    category=category or '',
                    brand=brand,
                    image_url=image_url,
                    active=True if active is None else active,
                    stock=0 if stock is None else max(0, int(stock)),
                    min_stock=0 if min_stock is None else max(0, int(min_stock)),
                    stock_kg=stock_kg,
                    kg_per_unit=kg_per_unit if kg_per_unit is not None else 1.0,
                    discount=0.0 if discount is None else float(discount),
                    sale_unit=sale_unit or 'unit',
                )

                try:
                    prod = crud.create_product(db, payload, actor=actor)
                    created += 1
                    try:
                        prod_id = int(prod.get('id') if isinstance(prod, dict) else getattr(prod, 'id', None))
                        created_ids.append(prod_id)
                        created_products.append({
                            'id': prod_id,
                            'name': name,
                            'code': code,
                            'image_url': image_url,
                        })
                    except Exception:
                        pass
                except HTTPException as he:
                    if he.status_code == 409:
                        skipped += 1
                        duplicates.append({'row': idx, 'code': code, 'reason': 'duplicado_en_bd'})
                        continue
                    raise
            except HTTPException:
                raise
            except Exception as e:
                skipped += 1
                errors.append({'row': idx, 'error': str(e)[:200]})
        return {
            'rows': len(rows),
            'created': created,
            'updated': updated,
            'skipped': skipped,
            'missing_name': missing_name,
            'duplicates': duplicates[:50],
            'errors': errors[:50],
            'created_ids': [c for c in created_ids if c],
            'updated_ids': [u for u in updated_ids if u],
        }, created_products
    result, created_products = await anyio.to_thread.run_sync(task)
    try:
        await push_event({"action": "imported", "created": int(result.get('created') or 0)})
    except Exception:
        pass
    try:
        await anyio.to_thread.run_sync(write_catalog_snapshot)
    except Exception:
        pass
    # Auto-image lookup intentionally disabled for Excel imports (manual only).
    _invalidate_products_cache()
    return result


@app.post("/products/auto-image")
async def auto_image_products(
    payload: Dict[str, Any] = Body(default=None),
    background_tasks: BackgroundTasks = None,
):
    if not _auto_image_enabled():
        raise HTTPException(
            status_code=400,
            detail='AUTO_IMAGE_FETCH y PEXELS_API_KEY deben estar configurados',
        )
    payload = payload or {}
    ids = payload.get('ids') or []
    limit_raw = payload.get('limit')
    only_missing = payload.get('only_missing', True)

    def task():
        db = SessionLocal()
        try:
            query = db.query(models.Product)
            if ids:
                try:
                    parsed_ids = [int(x) for x in ids]
                except Exception:
                    parsed_ids = [int(x) for x in ids if str(x).isdigit()]
                if parsed_ids:
                    query = query.filter(models.Product.id.in_(parsed_ids))
            if only_missing:
                query = query.filter(
                    or_(
                        models.Product.image_url.is_(None),
                        models.Product.image_url == '',
                    )
                )
            try:
                limit_val = int(limit_raw) if limit_raw is not None else 25
            except Exception:
                limit_val = 25
            if limit_val and limit_val > 0:
                query = query.order_by(models.Product.created_at.desc()).limit(limit_val)
            rows = query.all()
            return [
                {
                    'id': r.id,
                    'name': r.name,
                    'code': getattr(r, 'code', None),
                    'image_url': getattr(r, 'image_url', None),
                }
                for r in rows
            ]
        finally:
            db.close()

    items = await anyio.to_thread.run_sync(task)
    if not items:
        return {'scheduled': 0, 'detail': 'sin_productos_para_procesar'}

    try:
        if background_tasks:
            try:
                _queue_auto_image_progress(len(items), source='manual')
            except Exception:
                pass
            background_tasks.add_task(_auto_fetch_images_for_products, items, True, 'manual')
    except Exception:
        pass
    return {
        'scheduled': len(items),
        'ids': [it.get('id') for it in items if it.get('id')],
    }


@app.get("/products/auto-image/status")
def auto_image_status():
    data = _get_auto_image_progress_snapshot()
    data['enabled'] = _auto_image_enabled()
    if not data.get('enabled'):
        data['status'] = 'disabled'
    return data


@app.get("/product-changes", response_model=List[schemas.ProductChangeResponse])
def list_product_changes(
    product_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 200,
    db: Session = Depends(get_db),
):
    limit = max(1, min(int(limit or 200), 500))
    skip = max(0, int(skip or 0))
    query = db.query(models.ProductChange)
    if product_id is not None:
        query = query.filter(models.ProductChange.product_id == int(product_id))
    query = query.order_by(models.ProductChange.created_at.desc(), models.ProductChange.id.desc())
    return query.offset(skip).limit(limit).all()


@app.get("/products/{product_id}", response_model=schemas.ProductResponse)
def get_product(product_id: int, db: Session = Depends(get_db)):
    try:
        prod = crud.get_product(db, product_id)
        if not prod:
            raise HTTPException(404, "Product not found")
        return prod
    except Exception:
        logger.exception('get_product ORM failed, attempting raw fallback')
        try:
            bind = db.get_bind()
            insp = inspect(bind)
            existing = {c['name'] for c in insp.get_columns('products')}
        except Exception:
            existing = set()
        cols = ['id','name','price','description','category','image_url','created_at','updated_at','active']
        if 'code' in existing: cols.append('code')
        if 'brand' in existing: cols.append('brand')
        if 'price_retail' in existing: cols.append('price_retail')
        if 'cost' in existing: cols.append('cost')
        if 'stock' in existing: cols.append('stock')
        if 'min_stock' in existing: cols.append('min_stock')
        if 'stock_kg' in existing: cols.append('stock_kg')
        if 'kg_per_unit' in existing: cols.append('kg_per_unit')
        if 'discount' in existing: cols.append('discount')
        if 'sale_unit' in existing: cols.append('sale_unit')
        cols_sql = ', '.join(cols)
        row = crud._safe_execute_fetchone(db, f"SELECT {cols_sql} FROM products WHERE id = :id LIMIT 1", {'id': product_id})
        if not row:
            raise HTTPException(404, 'Product not found')
        objd = {cols[i]: row[i] for i in range(len(cols))}
        return objd

@app.put("/products/{product_id}", response_model=schemas.ProductResponse)
async def update_product(product_id: int, payload: schemas.ProductUpdate, request: Request):
    # Log incoming payload for diagnostics
    try:
        logger.info('PUT /products/%s called with payload: %s', product_id, jsonable_encoder(payload))
    except Exception:
        logger.exception('Failed to log payload for PUT /products/%s', product_id)

    actor = None
    try:
        actor = (request.headers.get('x-actor') or request.headers.get('X-Actor') or '').strip() or None
    except Exception:
        actor = None

    def task():
        db = SessionLocal()
        try:
            prod = crud.update_product(db, product_id, payload, actor=actor, action='update')
            if isinstance(prod, dict):
                return prod
            # Convert ORM instance to plain dict before session closes
            try:
                return {k: getattr(prod, k) for k in ('id','code','name','brand','price','price_retail','cost','description','category','image_url','active','stock','min_stock','stock_kg','kg_per_unit','discount','sale_unit','created_at','updated_at') if hasattr(prod, k)}
            except Exception:
                try:
                    return jsonable_encoder(prod)
                except Exception:
                    try:
                        return {'id': getattr(prod, 'id', None)}
                    except Exception:
                        return {}
        finally:
            db.close()

    try:
        prod = await anyio.to_thread.run_sync(task)
    except Exception as e:
        try:
            tb = traceback.format_exc()
            base = os.path.dirname(os.path.dirname(__file__))
            with open(os.path.join(base, 'server_log.txt'), 'a', encoding='utf-8') as f:
                f.write(f"{datetime.datetime.utcnow().isoformat()} - PUT /products/{product_id} exception: {str(e)[:400]}\n")
                f.write(tb + "\n\n")
        except Exception:
            pass
        logger.exception('PUT /products/%s failed: %s', product_id, e)
        raise HTTPException(status_code=500, detail=str(e)[:200])

    # Normalize prod to obtain id whether it's a dict or object
    prod_id = prod.get('id') if isinstance(prod, dict) else getattr(prod, 'id', None)
    try:
        await push_event({"action": "updated", "product": {"id": prod_id}})
    except Exception:
        pass
    try:
        await anyio.to_thread.run_sync(write_catalog_snapshot)
    except Exception:
        pass
    _invalidate_products_cache()
    return prod

@app.delete("/products/{product_id}")
async def delete_product(product_id: int, request: Request):
    actor = None
    try:
        actor = (request.headers.get('x-actor') or request.headers.get('X-Actor') or '').strip() or None
    except Exception:
        actor = None

    def task():
        db = SessionLocal()
        try:
            crud.delete_product(db, product_id, actor=actor)
        finally:
            db.close()

    await anyio.to_thread.run_sync(task)
    await push_event({"action": "deleted", "product": {"id": product_id}})
    await anyio.to_thread.run_sync(write_catalog_snapshot)
    _invalidate_products_cache()
    return {"detail": "deleted"}


@app.get('/debug/products-info')
def debug_products_info(db: Session = Depends(get_db)):
    """Return current products table columns and a small sample of rows (raw select).
    Use this to diagnose schema drift on deployments.
    """
    try:
        bind = db.get_bind()
        insp = inspect(bind)
        cols = [c['name'] for c in insp.get_columns('products')]
    except Exception as e:
        cols = []
    sample = []
    try:
        if cols:
            cols_sql = ', '.join(cols)
            rows = crud._safe_execute_fetchall(db, f"SELECT {cols_sql} FROM products ORDER BY created_at DESC LIMIT 10")
            for row in rows:
                obj = {cols[i]: row[i] for i in range(len(cols))}
                sample.append(obj)
    except Exception:
        try:
            # fallback: try ORM
            orm_rows = db.query(models.Product).order_by(models.Product.created_at.desc()).limit(10).all()
            for r in orm_rows:
                d = {c: getattr(r, c, None) for c in ('id','code','name','price','price_retail','description','category','image_url','created_at','updated_at','active','stock','stock_kg','kg_per_unit','discount','sale_unit')}
                sample.append(d)
        except Exception:
            pass
    return { 'columns': cols, 'sample': sample }

# -------------------------------------------------------------------
# EXPORT / SNAPSHOT
# -------------------------------------------------------------------
def write_catalog_snapshot():
    # ensure catalog directory exists
    try:
        os.makedirs(CATALOG_DIR, exist_ok=True)
    except Exception:
        # if we cannot create the directory, bail out
        logger.exception('Could not ensure CATALOG_DIR exists')
        return

    db = SessionLocal()
    try:
        products = crud.export_all(db)
    finally:
        db.close()

    # Try to reflect consumos (reserved near-expiry qty) in the snapshot so public catalog shows correct available stock
    consumos_path = os.path.join(CATALOG_DIR, 'consumos.json')
    consumos_map = {}
    try:
        if os.path.exists(consumos_path):
            with open(consumos_path, 'r', encoding='utf-8') as f:
                _items = json.load(f) or []
            for it in _items:
                try:
                    pid = str(it.get('id'))
                    qty = int(it.get('qty') or 0)
                    disc = int(it.get('discount') or it.get('value') or 0)
                    if pid not in consumos_map:
                        consumos_map[pid] = {'qty': 0, 'discount': disc}
                    consumos_map[pid]['qty'] += qty
                except Exception:
                    continue
    except Exception:
        consumos_map = {}

    data = []
    for p in products:
        pid = str(p.id)
        reserved = consumos_map.get(pid, {}).get('qty', 0)
        discount_for_consumo = consumos_map.get(pid, {}).get('discount')
        try:
            orig_stock = int(p.stock) if getattr(p, 'stock', None) is not None else None
        except Exception:
            orig_stock = None
        try:
            raw_stock_kg = getattr(p, 'stock_kg', None)
            if raw_stock_kg is None or float(raw_stock_kg) <= 0:
                raw_stock_kg = p.stock if getattr(p, 'stock', None) is not None else 0.0
            orig_stock_kg = float(raw_stock_kg or 0.0)
        except Exception:
            orig_stock_kg = 0.0
        try:
            kg_per_unit = float(getattr(p, 'kg_per_unit', 1.0) or 1.0)
        except Exception:
            kg_per_unit = 1.0
        adjusted_stock = None
        if orig_stock is not None:
            adjusted_stock = max(0, orig_stock - reserved)
        adjusted_stock_kg = max(0.0, orig_stock_kg)
        data.append({
            "id": p.id,
            "code": (str(getattr(p, 'code', '') or '').strip() or None),
            "name": p.name,
            "brand": (str(getattr(p, 'brand', '') or '').strip() or None),
            "price": float(p.price) if p.price else None,
            "price_retail": float(p.price_retail) if getattr(p, 'price_retail', None) is not None else None,
            "description": p.description,
            "category": p.category,
            "image_url": p.image_url,
            "active": p.active,
            "stock": adjusted_stock,
            "stock_kg": adjusted_stock_kg,
            "kg_per_unit": kg_per_unit,
            "discount": int(p.discount) if getattr(p, 'discount', None) is not None else None,
            "sale_unit": getattr(p, 'sale_unit', 'unit') or 'unit',
            "consumo_qty": int(reserved) if reserved else None,
            "consumo_discount": int(discount_for_consumo) if discount_for_consumo is not None else None,
        })

    with open(os.path.join(CATALOG_DIR, "products.json"), "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    # best-effort remote backup (async-friendly): update configured gist or URL
    try:
        content = json.dumps(data, ensure_ascii=False, indent=2)
        # schedule a background push (use anyio.to_thread to avoid blocking if not awaited)
        try:
            import anyio
            anyio.from_thread.run(lambda: None)
        except Exception:
            pass
        # try to push (synchronously here but guarded)
        if GIST_TOKEN and GIST_ID:
            try:
                httpx.patch(f"https://api.github.com/gists/{GIST_ID}", json={"files": {"products.json": {"content": content}}}, headers={"Authorization": f"token {GIST_TOKEN}"}, timeout=15)
            except Exception:
                logger.warning('Remote gist backup failed (write_catalog_snapshot)')
    except Exception:
        logger.exception('write_catalog_snapshot: remote backup step failed')


def _parse_promo_datetime(value: Any) -> Optional[datetime.datetime]:
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        dt = value
    else:
        try:
            raw = str(value).strip()
        except Exception:
            return None
        if not raw:
            return None
        # Accept common ISO forms, including trailing Z.
        if raw.endswith('Z'):
            raw = raw[:-1] + '+00:00'
        try:
            dt = datetime.datetime.fromisoformat(raw)
        except Exception:
            return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=datetime.timezone.utc)
    else:
        dt = dt.astimezone(datetime.timezone.utc)
    return dt


def _normalize_promotion_entry(entry: Any) -> Optional[Dict[str, Any]]:
    if not isinstance(entry, dict):
        return None
    out = dict(entry)

    # Normalize common aliases and preserve unknown keys.
    product_ids = out.get('productIds')
    if product_ids is None:
        product_ids = out.get('product_ids')
    if isinstance(product_ids, list):
        norm_ids = []
        for pid in product_ids:
            if pid is None:
                continue
            try:
                s = str(pid).strip()
                if not s:
                    continue
                norm_ids.append(int(s) if s.isdigit() else s)
            except Exception:
                continue
        out['productIds'] = norm_ids

    if 'type' in out and out.get('type') is not None:
        try:
            out['type'] = str(out.get('type')).strip().lower()
        except Exception:
            pass

    if 'value' in out and out.get('value') is not None:
        try:
            out['value'] = float(out.get('value'))
        except Exception:
            pass

    valid_until_raw = (
        out.get('valid_until')
        if out.get('valid_until') is not None
        else out.get('validUntil')
    )
    valid_until_dt = _parse_promo_datetime(valid_until_raw)
    if valid_until_dt:
        out['valid_until'] = valid_until_dt.isoformat().replace('+00:00', 'Z')
    elif valid_until_raw is not None:
        out['valid_until'] = None

    return out


def _normalize_promotions_payload(promos: Any) -> List[Dict[str, Any]]:
    if not isinstance(promos, list):
        return []
    out: List[Dict[str, Any]] = []
    for item in promos:
        normalized = _normalize_promotion_entry(item)
        if normalized is not None:
            out.append(normalized)
    return out


def write_promotions_snapshot(promos):
    """Write promotions snapshot to catalog directory so frontend/admin can read/write a canonical file."""
    try:
        normalized_promos = _normalize_promotions_payload(promos or [])
        if not os.path.exists(CATALOG_DIR):
            os.makedirs(CATALOG_DIR, exist_ok=True)
        path = os.path.join(CATALOG_DIR, "promotions.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(normalized_promos, f, indent=2, ensure_ascii=False)
        logger.info('promotions snapshot written to %s', path)
        # best-effort: also push promotions to configured gist backup
        try:
            if GIST_TOKEN and GIST_ID:
                url = f"https://api.github.com/gists/{GIST_ID}"
                payload = {"files": {"promotions.json": {"content": json.dumps(normalized_promos, ensure_ascii=False, indent=2)}}}
                try:
                    resp = httpx.patch(url, json=payload, headers={"Authorization": f"token {GIST_TOKEN}", "Accept": "application/vnd.github.v3+json"}, timeout=15)
                    if resp.status_code >= 200 and resp.status_code < 300:
                        logger.info('promotions pushed to gist successfully')
                except Exception as e:
                    logger.warning('promotions push to gist failed: %s', e)
        except Exception:
            logger.exception('promotions push step failed')
        return True
    except Exception as e:
        logger.exception('write_promotions_snapshot failed: %s', e)
        return False


@app.get('/promotions')
def list_promotions():
    """Return persisted promotions snapshot if present, otherwise return empty list."""
    try:
        return _get_promotions_cached()
    except Exception as e:
        logger.exception('list_promotions failed: %s', e)
    return []


@app.post('/promotions')
def save_promotions(promos: List[Dict[str, Any]]):
    """Persist promotions snapshot (admin may send the full array)."""
    try:
        normalized_promos = _normalize_promotions_payload(promos or [])
        logger.info('Received /promotions POST, count=%s', len(normalized_promos))
        ok = write_promotions_snapshot(normalized_promos)
        if not ok:
            raise HTTPException(status_code=500, detail='failed to write promotions')
        _invalidate_promotions_cache()
        return { 'detail': 'ok', 'count': len(normalized_promos) }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception('save_promotions error: %s', e)
        raise HTTPException(status_code=500, detail='save failed')

@app.get("/export")
def export(format: str = "json", db: Session = Depends(get_db)):
    products = crud.export_all(db)
    field_order = [
        'id',
        'code',
        'name',
        'brand',
        'category',
        'description',
        'price',
        'price_retail',
        'cost',
        'discount',
        'sale_unit',
        'kg_per_unit',
        'stock',
        'stock_kg',
        'min_stock',
        'image_url',
        'active',
        'created_at',
        'updated_at',
    ]

    rows = []
    all_keys = set()
    for p in (products or []):
        raw = p if isinstance(p, dict) else (getattr(p, '__dict__', None) or {})
        cleaned = {k: v for (k, v) in (raw or {}).items() if k and not str(k).startswith('_')}
        rows.append(cleaned)
        all_keys.update(cleaned.keys())

    # Ensure stable column order for CSV output.
    fieldnames = [k for k in field_order if k in all_keys] + [k for k in sorted(all_keys) if k not in field_order]

    if format == "csv":
        si = StringIO()
        writer = csv.DictWriter(si, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(rows)
        return PlainTextResponse(si.getvalue(), media_type="text/csv")

    return rows


@app.post('/backup')
def trigger_backup(db: Session = Depends(get_db)):
    """Trigger writing the static snapshot and pushing to remote backup (if configured)."""
    try:
        write_catalog_snapshot()
        return {"detail": "snapshot written"}
    except Exception as e:
        logger.exception('backup failed')
        raise HTTPException(500, 'backup failed')


@app.get("/stats")
def stats(db: Session = Depends(get_db)):
    return crud.stats(db)

@app.get("/sales/stats")
def sales_stats(
    days: int = 30,
    source: Optional[str] = None,
    customer_type: Optional[str] = None,
    limit: int = 5000,
    db: Session = Depends(get_db),
):
    """Simple sales statistics based on recent orders.

    Returns aggregated metrics for the last `days` (inclusive), computed in the
    business timezone (ORDER_EMAIL_TIMEZONE or ORDER_DELIVERY_TIMEZONE).
    """
    try:
        days_int = int(days or 30)
    except Exception:
        days_int = 30
    days_int = max(1, min(365, days_int))

    try:
        limit_int = int(limit or 5000)
    except Exception:
        limit_int = 5000
    limit_int = max(50, min(20000, limit_int))

    tz_name = str(
        os.environ.get('ORDER_EMAIL_TIMEZONE')
        or os.environ.get('ORDER_DELIVERY_TIMEZONE')
        or 'America/Argentina/Buenos_Aires'
    ).strip() or 'America/Argentina/Buenos_Aires'
    try:
        from zoneinfo import ZoneInfo
        tzinfo = ZoneInfo(tz_name)
    except Exception:
        tzinfo = datetime.timezone.utc
        tz_name = 'UTC'

    now_local = datetime.datetime.now(tzinfo)
    end_date = now_local.date()
    start_date = end_date - datetime.timedelta(days=days_int - 1)

    def _normalize_customer_type(value: Any) -> str:
        v = str(value or '').strip().lower()
        return v if v in ('mayorista', 'minorista') else ''

    customer_type_filter = _normalize_customer_type(customer_type)

    orders_raw = crud.get_orders(db, skip=0, limit=limit_int, source=source, q=None, date=None) or []
    orders_list: List[Dict[str, Any]] = []
    for o in orders_raw:
        if isinstance(o, dict):
            orders_list.append(o)
        else:
            try:
                orders_list.append({k: getattr(o, k, None) for k in ('id', 'items', 'total', 'created_at', 'customer_type', 'source', 'status')})
            except Exception:
                continue

    by_day: Dict[str, Dict[str, Any]] = {}
    orders_count = 0
    revenue_total = 0.0
    top_products: Dict[str, Dict[str, Any]] = {}

    for order in orders_list:
        try:
            if customer_type_filter:
                if _normalize_customer_type(order.get('customer_type')) != customer_type_filter:
                    continue

            created_at = order.get('created_at')
            dt = None
            if isinstance(created_at, datetime.datetime):
                dt = created_at
            elif created_at:
                try:
                    raw = str(created_at).strip()
                    if raw.endswith('Z'):
                        raw = raw[:-1] + '+00:00'
                    dt = datetime.datetime.fromisoformat(raw)
                except Exception:
                    dt = None
            if dt is None:
                continue
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=datetime.timezone.utc)
            local_date = dt.astimezone(tzinfo).date()
            if local_date < start_date or local_date > end_date:
                continue

            try:
                total = float(order.get('total') or 0.0)
            except Exception:
                total = 0.0

            orders_count += 1
            revenue_total += total

            day_key = local_date.isoformat()
            bucket = by_day.get(day_key)
            if not bucket:
                bucket = {'date': day_key, 'orders': 0, 'revenue': 0.0}
                by_day[day_key] = bucket
            bucket['orders'] = int(bucket.get('orders') or 0) + 1
            bucket['revenue'] = float(bucket.get('revenue') or 0.0) + float(total or 0.0)

            items = order.get('items') if isinstance(order.get('items'), list) else []
            for it in items:
                if not isinstance(it, dict):
                    continue
                pid = str(it.get('id') or '').strip()
                if not pid:
                    continue
                try:
                    qty = float(it.get('qty') or 0.0)
                except Exception:
                    qty = 0.0
                meta = it.get('meta') if isinstance(it.get('meta'), dict) else {}
                name = meta.get('name') or meta.get('title') or meta.get('product_name') or ''
                price_raw = meta.get('price')
                if price_raw is None:
                    price_raw = meta.get('unit_price')
                try:
                    unit_price = float(price_raw) if price_raw is not None else 0.0
                except Exception:
                    unit_price = 0.0
                revenue = qty * unit_price
                rec = top_products.get(pid)
                if not rec:
                    rec = {'product_id': pid, 'name': name or pid, 'qty': 0.0, 'revenue': 0.0}
                    top_products[pid] = rec
                rec['qty'] = float(rec.get('qty') or 0.0) + float(qty or 0.0)
                rec['revenue'] = float(rec.get('revenue') or 0.0) + float(revenue or 0.0)
                if name and (not rec.get('name') or rec.get('name') == pid):
                    rec['name'] = name
        except Exception:
            continue

    # Fill missing days so charts can render contiguous series.
    series = []
    cursor = start_date
    while cursor <= end_date:
        key = cursor.isoformat()
        bucket = by_day.get(key) or {'date': key, 'orders': 0, 'revenue': 0.0}
        series.append({
            'date': key,
            'orders': int(bucket.get('orders') or 0),
            'revenue': float(round(float(bucket.get('revenue') or 0.0), 2)),
        })
        cursor = cursor + datetime.timedelta(days=1)

    top = sorted(top_products.values(), key=lambda x: float(x.get('revenue') or 0.0), reverse=True)[:20]
    for rec in top:
        try:
            rec['qty'] = float(round(float(rec.get('qty') or 0.0), 3))
        except Exception:
            rec['qty'] = 0.0
        try:
            rec['revenue'] = float(round(float(rec.get('revenue') or 0.0), 2))
        except Exception:
            rec['revenue'] = 0.0

    avg_ticket = (revenue_total / orders_count) if orders_count else 0.0
    return {
        'timezone': tz_name,
        'days': days_int,
        'from': start_date.isoformat(),
        'to': end_date.isoformat(),
        'orders': int(orders_count),
        'revenue': float(round(revenue_total, 2)),
        'avg_ticket': float(round(avg_ticket, 2)),
        'by_day': series,
        'top_products': top,
        'source': source,
        'customer_type': customer_type_filter or None,
        'sample_size': len(orders_list),
    }


# -----------------------------
# Orders
# -----------------------------
def _infer_frontend_origin(request: Request) -> Optional[str]:
    try:
        origin = (request.headers.get('origin') or '').strip()
        if origin.lower().startswith('http://') or origin.lower().startswith('https://'):
            return origin.rstrip('/')
    except Exception:
        pass
    try:
        from urllib.parse import urlparse
        referer = (request.headers.get('referer') or '').strip()
        if referer.lower().startswith('http://') or referer.lower().startswith('https://'):
            parsed = urlparse(referer)
            if parsed.scheme and parsed.netloc:
                return f"{parsed.scheme}://{parsed.netloc}"
    except Exception:
        pass
    try:
        return str(request.base_url).rstrip('/')
    except Exception:
        return None


def _update_order_payment_snapshot(order_id: Any, payment_method: Optional[str], payment_status: Optional[str], payment_reference: Optional[str]) -> None:
    try:
        insp = inspect(engine)
        existing_cols = {c['name'] for c in insp.get_columns('orders')}
    except Exception:
        existing_cols = set()

    # Self-heal legacy schemas: if payment columns are missing, attempt the
    # same migration used by debug tooling before writing the payment snapshot.
    required_payment_cols = {'payment_method', 'payment_status', 'payment_reference'}
    if required_payment_cols - existing_cols:
        try:
            _run_add_user_columns()
            insp = inspect(engine)
            existing_cols = {c['name'] for c in insp.get_columns('orders')}
        except Exception:
            logger.exception('Could not auto-migrate payment columns before snapshot update')

    set_parts = []
    params: Dict[str, Any] = {}
    if payment_method is not None and 'payment_method' in existing_cols:
        set_parts.append('payment_method = :payment_method')
        params['payment_method'] = str(payment_method)
    if payment_status is not None and 'payment_status' in existing_cols:
        set_parts.append('payment_status = :payment_status')
        params['payment_status'] = str(payment_status)
    if payment_reference is not None and 'payment_reference' in existing_cols:
        set_parts.append('payment_reference = :payment_reference')
        params['payment_reference'] = str(payment_reference)

    if not set_parts:
        return

    set_sql = ', '.join(set_parts)
    use_int_id = False
    try:
        oid_int = int(order_id)
        use_int_id = True
    except Exception:
        oid_int = None

    with engine.begin() as conn:
        if use_int_id:
            conn.execute(text(f"UPDATE orders SET {set_sql} WHERE id = :id"), {**params, 'id': oid_int})
        else:
            conn.execute(text(f"UPDATE orders SET {set_sql} WHERE CAST(id AS TEXT) = :id"), {**params, 'id': str(order_id)})


def _normalize_mp_payment_status(value: Optional[str]) -> str:
    raw = str(value or '').strip().lower()
    if raw in ('approved', 'accredited'):
        return 'approved'
    if raw in ('rejected', 'cancelled', 'cancelled_by_user'):
        return 'rejected'
    if raw in ('refunded', 'charged_back'):
        return 'refunded'
    if raw in ('authorized', 'in_process', 'inprocess', 'pending', 'in_mediation', 'action_required'):
        return 'in_process'
    if raw:
        return raw
    return 'mp_pending'


async def _sync_mercadopago_payment(
    payment_id: Optional[str],
    external_reference: Optional[str],
    raw_status: Optional[str],
    metadata: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    metadata = metadata or {}
    access_token = (os.environ.get('MERCADOPAGO_ACCESS_TOKEN') or '').strip()
    allow_unverified_raw = (os.environ.get('MERCADOPAGO_ALLOW_UNVERIFIED_SYNC') or 'false').strip().lower()
    allow_unverified = allow_unverified_raw in ('1', 'true', 'yes', 'on')

    resolved_order_id = str(external_reference or metadata.get('order_id') or '').strip()
    resolved_status = str(raw_status or '').strip().lower()
    payment_reference = str(payment_id or '').strip() or None
    mp_payload: Optional[Dict[str, Any]] = None

    # Security: by default, require payment_id so the backend can verify the
    # payment against Mercado Pago before updating order status.
    if not payment_id and not allow_unverified:
        return {
            'ok': True,
            'updated': False,
            'reason': 'missing_payment_id',
            'payment_status': _normalize_mp_payment_status(resolved_status),
        }

    # Prefer authoritative payment details from MP when payment_id is available.
    if payment_id and access_token:
        try:
            timeout = httpx.Timeout(20.0, connect=8.0)
            async with httpx.AsyncClient(timeout=timeout) as client:
                resp = await client.get(
                    f"https://api.mercadopago.com/v1/payments/{payment_id}",
                    headers={'Authorization': f'Bearer {access_token}'},
                )
            if resp.status_code < 400:
                mp_payload = resp.json()
                resolved_status = str(mp_payload.get('status') or resolved_status or '').strip().lower()
                mp_ext = str(mp_payload.get('external_reference') or '').strip()
                if mp_ext:
                    resolved_order_id = mp_ext
                try:
                    md = mp_payload.get('metadata') if isinstance(mp_payload, dict) else None
                    if not resolved_order_id and isinstance(md, dict):
                        resolved_order_id = str(md.get('order_id') or '').strip()
                except Exception:
                    pass
                pid = str(mp_payload.get('id') or payment_id).strip()
                payment_reference = pid or payment_reference
            else:
                logger.warning(
                    'Mercado Pago payment lookup failed status=%s body=%s',
                    resp.status_code,
                    (resp.text or '')[:500],
                )
        except Exception:
            logger.exception('Could not fetch Mercado Pago payment details for payment_id=%s', payment_id)

    if not resolved_order_id:
        return {
            'ok': True,
            'updated': False,
            'reason': 'missing_order_reference',
            'payment_status': _normalize_mp_payment_status(resolved_status),
        }

    normalized_status = _normalize_mp_payment_status(resolved_status)
    try:
        _update_order_payment_snapshot(
            order_id=resolved_order_id,
            payment_method='mercadopago',
            payment_status=normalized_status,
            payment_reference=payment_reference,
        )
    except Exception:
        logger.exception('Could not update order payment snapshot from MP event order_id=%s', resolved_order_id)
        return {
            'ok': False,
            'updated': False,
            'order_id': str(resolved_order_id),
            'payment_status': normalized_status,
            'payment_reference': payment_reference,
        }

    return {
        'ok': True,
        'updated': True,
        'order_id': str(resolved_order_id),
        'payment_status': normalized_status,
        'payment_reference': payment_reference,
        'has_payment_payload': bool(mp_payload),
    }


@app.get('/payments/mercadopago/health')
async def mercadopago_health(request: Request):
    configured = bool((os.environ.get('MERCADOPAGO_ACCESS_TOKEN') or '').strip())
    use_sdk_raw = (os.environ.get('MERCADOPAGO_USE_SDK') or 'true').strip().lower()
    use_sdk = use_sdk_raw not in ('0', 'false', 'no', 'off')
    explicit_success = (os.environ.get('MERCADOPAGO_SUCCESS_URL') or '').strip()
    explicit_failure = (os.environ.get('MERCADOPAGO_FAILURE_URL') or '').strip()
    explicit_pending = (os.environ.get('MERCADOPAGO_PENDING_URL') or '').strip()
    return_path = (os.environ.get('MERCADOPAGO_RETURN_PATH') or '/catalogo').strip() or '/catalogo'
    allow_unverified_raw = (os.environ.get('MERCADOPAGO_ALLOW_UNVERIFIED_SYNC') or 'false').strip().lower()
    allow_unverified = allow_unverified_raw in ('1', 'true', 'yes', 'on')

    try:
        insp = inspect(engine)
        cols = {c['name'] for c in insp.get_columns('orders')}
    except Exception:
        cols = set()

    required = {'payment_method', 'payment_status', 'payment_reference'}
    headers = _cors_headers_for_request(request)
    return JSONResponse(
        status_code=200,
        content={
            'configured': configured,
            'sdk_installed': mercadopago is not None,
            'sdk_enabled': use_sdk,
            'orders_payment_columns_ready': required.issubset(cols),
            'missing_order_columns': sorted(list(required - cols)),
            'explicit_back_urls': {
                'success': bool(explicit_success),
                'failure': bool(explicit_failure),
                'pending': bool(explicit_pending),
            },
            'notification_url_configured': bool((os.environ.get('MERCADOPAGO_NOTIFICATION_URL') or '').strip()),
            'allow_unverified_sync': allow_unverified,
            'return_path': return_path,
        },
        headers=headers,
    )


@app.api_route('/payments/mercadopago/webhook', methods=['POST', 'GET'])
async def mercadopago_webhook(request: Request):
    query = request.query_params
    payment_id = (
        query.get('payment_id')
        or query.get('id')
        or query.get('data.id')
    )
    external_reference = query.get('external_reference')
    status = query.get('status') or query.get('collection_status')

    body: Dict[str, Any] = {}
    try:
        maybe_body = await request.json()
        if isinstance(maybe_body, dict):
            body = maybe_body
    except Exception:
        body = {}

    if not payment_id:
        try:
            data_obj = body.get('data')
            if isinstance(data_obj, dict):
                payment_id = data_obj.get('id') or data_obj.get('payment_id')
        except Exception:
            pass

    if not external_reference:
        external_reference = str(body.get('external_reference') or '').strip() or None

    if not status:
        status = str(body.get('status') or body.get('collection_status') or '').strip() or None

    metadata = body.get('metadata') if isinstance(body.get('metadata'), dict) else {}
    result = await _sync_mercadopago_payment(
        payment_id=str(payment_id or '').strip() or None,
        external_reference=str(external_reference or '').strip() or None,
        raw_status=str(status or '').strip() or None,
        metadata=metadata,
    )

    headers = _cors_headers_for_request(request)
    return JSONResponse(status_code=200, content=result, headers=headers)


@app.post('/payments/mercadopago/sync')
async def mercadopago_sync(request: Request):
    body: Dict[str, Any] = {}
    try:
        maybe_body = await request.json()
        if isinstance(maybe_body, dict):
            body = maybe_body
    except Exception:
        body = {}

    payment_id = str(body.get('payment_id') or body.get('id') or '').strip() or None
    external_reference = str(body.get('external_reference') or body.get('order_id') or '').strip() or None
    status = str(body.get('status') or body.get('collection_status') or '').strip() or None
    metadata = body.get('metadata') if isinstance(body.get('metadata'), dict) else {}

    result = await _sync_mercadopago_payment(
        payment_id=payment_id,
        external_reference=external_reference,
        raw_status=status,
        metadata=metadata,
    )
    headers = _cors_headers_for_request(request)
    return JSONResponse(status_code=200, content=result, headers=headers)


@app.post('/payments/mercadopago/preference', response_model=schemas.MercadoPagoPreferenceResponse)
async def create_mercadopago_preference(request: Request, payload: schemas.MercadoPagoPreferenceCreate):
    access_token = (os.environ.get('MERCADOPAGO_ACCESS_TOKEN') or '').strip()
    if not access_token:
        raise HTTPException(status_code=503, detail='Mercado Pago no esta configurado')

    items = []
    for item in (payload.items or []):
        try:
            qty = int(item.quantity or 1)
        except Exception:
            qty = 1
        if qty <= 0:
            qty = 1

        try:
            unit_price = float(item.unit_price or 0)
        except Exception:
            unit_price = 0
        if unit_price <= 0:
            raise HTTPException(status_code=400, detail='Todos los items deben tener un precio valido')

        title = (item.title or '').strip()
        if not title:
            title = f"Producto {item.id}"

        mp_item = {
            'id': str(item.id),
            'title': title,
            'quantity': qty,
            'unit_price': round(unit_price, 2),
            'currency_id': str(item.currency_id or 'ARS').upper(),
        }
        if item.description:
            mp_item['description'] = str(item.description)
        items.append(mp_item)

    if not items:
        raise HTTPException(status_code=400, detail='No hay items para cobrar')

    mp_payload: Dict[str, Any] = {
        'items': items,
        'external_reference': str(payload.external_reference or payload.order_id),
        'metadata': {
            'order_id': str(payload.order_id)
        }
    }
    try:
        if payload.total is not None:
            mp_payload['metadata']['order_total'] = float(payload.total)
    except Exception:
        pass

    payer_obj = {}
    try:
        if payload.payer:
            if payload.payer.name:
                payer_obj['name'] = str(payload.payer.name)
            if payload.payer.email:
                payer_obj['email'] = str(payload.payer.email)
    except Exception:
        payer_obj = {}
    if payer_obj:
        mp_payload['payer'] = payer_obj

    def _is_http_url(v: Any) -> bool:
        try:
            s = str(v or '').strip().lower()
            return s.startswith('http://') or s.startswith('https://')
        except Exception:
            return False

    # 1) Explicit env URLs (highest priority in production)
    back_urls: Dict[str, str] = {}
    env_back_urls = {
        'success': (os.environ.get('MERCADOPAGO_SUCCESS_URL') or '').strip(),
        'failure': (os.environ.get('MERCADOPAGO_FAILURE_URL') or '').strip(),
        'pending': (os.environ.get('MERCADOPAGO_PENDING_URL') or '').strip(),
    }
    for key in ('success', 'failure', 'pending'):
        value = env_back_urls.get(key)
        if _is_http_url(value):
            back_urls[key] = str(value).strip()

    # 2) Request payload URLs (frontend-provided)
    try:
        raw_back_urls = payload.back_urls if isinstance(payload.back_urls, dict) else {}
    except Exception:
        raw_back_urls = {}
    for key in ('success', 'failure', 'pending'):
        if key in back_urls:
            continue
        value = raw_back_urls.get(key)
        if _is_http_url(value):
            back_urls[key] = str(value).strip()

    # 3) Inferred base URL fallback
    if 'success' not in back_urls:
        forced_origin = (
            (os.environ.get('FRONTEND_BASE_URL') or '')
            or (os.environ.get('PUBLIC_FRONTEND_URL') or '')
            or (os.environ.get('APP_BASE_URL') or '')
        ).strip().rstrip('/')
        origin = forced_origin if _is_http_url(forced_origin) else _infer_frontend_origin(request)
        if origin:
            return_path = (os.environ.get('MERCADOPAGO_RETURN_PATH') or '/catalogo').strip()
            if not return_path:
                return_path = '/catalogo'
            if not return_path.startswith('/'):
                return_path = '/' + return_path
            back_urls['success'] = back_urls.get('success') or f"{origin}{return_path}?payment=success"
            back_urls['failure'] = back_urls.get('failure') or f"{origin}{return_path}?payment=failure"
            back_urls['pending'] = back_urls.get('pending') or f"{origin}{return_path}?payment=pending"

    if back_urls:
        mp_payload['back_urls'] = back_urls
        # Mercado Pago requires back_urls.success when auto_return=approved.
        if back_urls.get('success'):
            mp_payload['auto_return'] = 'approved'

    notification_url = (os.environ.get('MERCADOPAGO_NOTIFICATION_URL') or '').strip()
    if notification_url:
        mp_payload['notification_url'] = notification_url

    data: Optional[Dict[str, Any]] = None
    use_sdk_raw = (os.environ.get('MERCADOPAGO_USE_SDK') or 'true').strip().lower()
    use_sdk = use_sdk_raw not in ('0', 'false', 'no', 'off')

    if use_sdk and mercadopago is not None:
        try:
            def _sdk_create_preference():
                client = mercadopago.SDK(access_token)
                return client.preference().create(mp_payload)

            sdk_resp = await anyio.to_thread.run_sync(_sdk_create_preference)
            if isinstance(sdk_resp, dict):
                try:
                    sdk_status = int(sdk_resp.get('status') or 0)
                except Exception:
                    sdk_status = 0
                if sdk_status >= 400:
                    raise RuntimeError(f'sdk_status={sdk_status}')
                sdk_payload = sdk_resp.get('response')
                if isinstance(sdk_payload, dict):
                    data = sdk_payload
                else:
                    data = sdk_resp
            if not isinstance(data, dict):
                raise RuntimeError('invalid_sdk_response')
        except Exception as sdk_err:
            data = None
            logger.warning('Mercado Pago SDK failed, using HTTP fallback: %s', sdk_err)

    if data is None:
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json',
            'X-Idempotency-Key': f"order-{payload.order_id}-{int(time.time() * 1000)}",
        }

        try:
            timeout = httpx.Timeout(25.0, connect=10.0)
            async with httpx.AsyncClient(timeout=timeout) as client:
                resp = await client.post(
                    'https://api.mercadopago.com/checkout/preferences',
                    json=mp_payload,
                    headers=headers,
                )
        except Exception as e:
            logger.exception('Mercado Pago request failed: %s', e)
            raise HTTPException(status_code=502, detail='No se pudo contactar a Mercado Pago')

    if data is not None:
        class _MPRespShim:
            def __init__(self, payload_obj):
                self.status_code = 200
                self.text = ''
                self._payload = payload_obj

            def json(self):
                return self._payload

        resp = _MPRespShim(data)


    if resp.status_code >= 400:
        raw = ''
        reason = ''
        try:
            raw = resp.text[:1000]
        except Exception:
            raw = 'error'
        try:
            err_json = resp.json()
        except Exception:
            err_json = None

        if isinstance(err_json, dict):
            candidates = [
                err_json.get('message'),
                err_json.get('error'),
                err_json.get('detail'),
            ]
            cause = err_json.get('cause')
            if isinstance(cause, list) and cause:
                first = cause[0]
                if isinstance(first, dict):
                    candidates.extend([
                        first.get('description'),
                        first.get('message'),
                        first.get('code'),
                    ])
                else:
                    candidates.append(str(first))
            for value in candidates:
                if isinstance(value, str) and value.strip():
                    reason = value.strip()
                    break
            if not reason:
                try:
                    reason = json.dumps(err_json, ensure_ascii=False)[:300]
                except Exception:
                    reason = ''
        else:
            reason = (raw or '').strip()[:300]

        logger.warning('Mercado Pago preference error %s: %s', resp.status_code, reason or raw)
        detail = f"Mercado Pago rechazo la preferencia ({resp.status_code})"
        if reason:
            detail = f"{detail}: {reason}"
        raise HTTPException(status_code=502, detail=detail)

    try:
        data = resp.json()
    except Exception:
        raise HTTPException(status_code=502, detail='Respuesta invalida de Mercado Pago')

    preference_id = str(data.get('id') or '').strip()
    init_point = str(data.get('init_point') or data.get('sandbox_init_point') or '').strip()
    sandbox_init_point = data.get('sandbox_init_point')

    if not preference_id or not init_point:
        logger.warning('Mercado Pago response missing checkout URLs: %s', data)
        raise HTTPException(status_code=502, detail='Mercado Pago no devolvio un link de pago')

    try:
        _update_order_payment_snapshot(
            order_id=payload.order_id,
            payment_method='mercadopago',
            payment_status='mp_pending',
            payment_reference=preference_id,
        )
    except Exception:
        logger.exception('Could not persist Mercado Pago preference in order %s', payload.order_id)

    return schemas.MercadoPagoPreferenceResponse(
        preference_id=preference_id,
        init_point=init_point,
        sandbox_init_point=sandbox_init_point,
    )


def _is_transient_db_error(exc: Exception) -> bool:
    """Detect serialization/deadlock/lock-timeout errors that can be retried safely."""
    try:
        if isinstance(exc, (OperationalError, DBAPIError)):
            orig = getattr(exc, 'orig', None)
            pgcode = getattr(orig, 'pgcode', None) or getattr(orig, 'code', None)
            if pgcode in {'40001', '40P01', '55P03', '57014'}:
                return True
    except Exception:
        pass
    msg = str(exc).lower()
    transient_snippets = (
        'deadlock detected',
        'could not serialize access',
        'serialization failure',
        'lock timeout',
        'could not obtain lock',
        'lock not available',
        'database is locked',
        'too many connections',
    )
    return any(snippet in msg for snippet in transient_snippets)


def _create_order_with_retry(payload: schemas.OrderCreate, token_payload: Optional[dict] = None):
    max_attempts = int(os.environ.get('ORDER_RETRY_ATTEMPTS') or 3)
    base_delay = float(os.environ.get('ORDER_RETRY_DELAY_SEC') or 0.08)
    last_exc = None
    for attempt in range(1, max_attempts + 1):
        db = SessionLocal()
        try:
            return crud.create_order(db, payload, current_user=token_payload)
        except HTTPException:
            # Business rule errors should not be retried.
            raise
        except Exception as exc:
            last_exc = exc
            try:
                db.rollback()
            except Exception:
                pass
            if attempt < max_attempts and _is_transient_db_error(exc):
                try:
                    logger.warning('create_order transient DB error, retrying (%s/%s): %s', attempt, max_attempts, str(exc)[:200])
                except Exception:
                    pass
                time.sleep(base_delay * attempt)
                continue
            raise
        finally:
            try:
                db.close()
            except Exception:
                pass
    if last_exc:
        raise last_exc


@app.post('/orders', response_model=schemas.OrderResponse)
async def create_order(request: Request, payload: schemas.OrderCreate):
    # Deep logging of payload for debugging (may include PII)
    try:
        encoded = jsonable_encoder(payload)
    except Exception:
        encoded = { 'items': getattr(payload, 'items', None), 'total': getattr(payload, 'total', None) }

    logger.info('create_order called; payload=%s', encoded)

    # Self-heal legacy schemas before persisting order snapshots (including
    # payment_method/payment_status for cash or Mercado Pago flows).
    try:
        _run_add_user_columns()
    except Exception:
        logger.exception('create_order: could not run orders optional-columns migration')

    # decode optional bearer token so the DB task can associate the order with the authenticated user
    auth = request.headers.get('authorization') or request.headers.get('Authorization')
    token_payload = None
    # Helpful debug: log whether the browser sent an Access-Control-Request-Headers header (preflight)
    try:
        acrh = request.headers.get('access-control-request-headers')
        if acrh:
            logger.debug('create_order: preflight requested headers: %s', acrh)
    except Exception:
        pass

    if auth and isinstance(auth, str) and auth.lower().startswith('bearer '):
        logger.info('create_order: Authorization header present')
        try:
            token = auth.split(' ', 1)[1]
            token_payload = utils.decode_access_token(token)
            logger.info('create_order: token_payload=%s', token_payload)
        except Exception as ex:
            logger.exception('create_order: failed to decode token: %s', ex)
            token_payload = None
    else:
        logger.info('create_order: no Authorization header present')

    # If a token payload was decoded, attach a safe preview to the request payload
    # so the CRUD layer can persist it into the orders table when columns exist.
    if token_payload:
        try:
            try:
                setattr(payload, '_token_received', True)
                existing_preview = getattr(payload, '_token_preview', None)
                merged_preview = dict(existing_preview) if isinstance(existing_preview, dict) else {}
                merged_preview.setdefault('sub', token_payload.get('sub') or token_payload.get('email') or None)
                merged_preview.setdefault('email', token_payload.get('email') or token_payload.get('sub') or None)
                merged_preview.setdefault('name', token_payload.get('full_name') or token_payload.get('name') or None)
                payload_postal = getattr(payload, 'user_postal_code', None)
                payload_department = getattr(payload, 'user_department', None)
                if payload_postal and not merged_preview.get('postal_code'):
                    merged_preview['postal_code'] = _normalize_postal_code(payload_postal)
                if payload_department and not merged_preview.get('department'):
                    merged_preview['department'] = str(payload_department).strip()
                setattr(payload, '_token_preview', merged_preview)
            except Exception:
                # Fallback: try direct attribute assignment
                payload._token_received = True
        except Exception:
            # Non-fatal: ignore if we cannot mutate the payload
            pass

    # Resolve customer email before creating the order.
    # Priority: payload/request -> token -> users table by user_id.
    try:
        encoded_for_email = encoded if isinstance(encoded, dict) else {}
        inferred_email = _extract_customer_email_for_order(
            order_data=encoded_for_email,
            request_data=encoded_for_email,
            token_payload=token_payload if isinstance(token_payload, dict) else None,
        )
        if not inferred_email:
            uid_candidate = None
            try:
                uid_candidate = getattr(payload, 'user_id', None)
            except Exception:
                uid_candidate = None
            if uid_candidate is None and isinstance(encoded_for_email, dict):
                uid_candidate = encoded_for_email.get('user_id')
            if uid_candidate is not None:
                inferred_email = _resolve_user_email_by_id(uid_candidate)

        if inferred_email:
            try:
                if not getattr(payload, 'user_email', None):
                    setattr(payload, 'user_email', inferred_email)
            except Exception:
                try:
                    payload.user_email = inferred_email
                except Exception:
                    pass
            try:
                if isinstance(encoded_for_email, dict) and not encoded_for_email.get('user_email'):
                    encoded_for_email['user_email'] = inferred_email
            except Exception:
                pass
        else:
            require_email_raw = (os.environ.get('ORDER_REQUIRE_CUSTOMER_EMAIL') or 'true').strip().lower()
            require_email = require_email_raw not in ('0', 'false', 'no', 'off')
            if require_email:
                logger.warning('create_order blocked: missing customer email in payload/token/user profile')
                raise HTTPException(
                    status_code=400,
                    detail='Customer email is required to confirm the order. Please provide user_email.',
                )
            logger.info('create_order continuing without customer email because ORDER_REQUIRE_CUSTOMER_EMAIL is disabled')
    except HTTPException:
        raise
    except Exception:
        logger.exception('create_order: could not normalize customer email')

    # Infer source from headers or payload: prefer explicit payload.source, then header 'X-Client-Platform' or 'X-Source'.
    # Si el payload no trae source, o viene vacío, se infiere SIEMPRE aquí y se fuerza el valor correcto.
    try:
        src = getattr(payload, 'source', None)
        if not src or not str(src).strip():
            src_hdr = request.headers.get('x-client-platform') or request.headers.get('x-source') or request.headers.get('x-client')
            if src_hdr:
                src = src_hdr
        if not src or not str(src).strip():
            ua = (request.headers.get('user-agent') or '').lower()
            # Detectar móvil por user-agent o headers
            if ua and ( 'okhttp' in ua or 'android' in ua or 'dalvik' in ua or 'retrofit' in ua or 'okhttp/' in ua ):
                src = 'app'
            # También considerar si el referer o origin contiene 'app' (por si hay proxy)
            ref = (request.headers.get('referer') or request.headers.get('origin') or '').lower()
            if 'app' in ref:
                src = 'app'
        # Si sigue sin source, default a web
        if not src or not str(src).strip():
            src = 'web'
        src = str(src).strip().lower()
        # Solo permitir 'app' o 'web'
        if src not in ('app', 'web'):
            src = 'web'
        # Forzar 'app' si user-agent es móvil aunque el cliente mande mal el campo
        ua = (request.headers.get('user-agent') or '').lower()
        if ua and ( 'okhttp' in ua or 'android' in ua or 'dalvik' in ua or 'retrofit' in ua or 'okhttp/' in ua ):
            src = 'app'
        try:
            setattr(payload, 'source', src)
        except Exception:
            try:
                payload.source = src
            except Exception:
                pass
        logger.info(f'[create_order] source inferido: {src}')
    except Exception as e:
        setattr(payload, 'source', 'web')
        logger.warning(f'[create_order] Error infiriendo source, se forzó a web: {e}')

    try:
        customer_type = getattr(payload, 'customer_type', None)
        if not customer_type or not str(customer_type).strip():
            customer_type = 'mayorista'
        customer_type = str(customer_type).strip().lower()
        if customer_type not in ('mayorista', 'minorista'):
            customer_type = 'mayorista'
        try:
            setattr(payload, 'customer_type', customer_type)
        except Exception:
            try:
                payload.customer_type = customer_type
            except Exception:
                pass
        try:
            if isinstance(encoded, dict):
                encoded['customer_type'] = customer_type
        except Exception:
            pass
        logger.info(f'[create_order] customer_type normalizado: {customer_type}')
    except Exception as e:
        try:
            setattr(payload, 'customer_type', 'mayorista')
        except Exception:
            pass
        logger.warning(f'[create_order] Error normalizando customer_type, se forzó a mayorista: {e}')

    def task():
        try:
            return _create_order_with_retry(payload, token_payload=token_payload)
        except Exception as inner_e:
            logger.exception('Error in create_order DB task: %s', inner_e)
            raise

    try:
        order = await anyio.to_thread.run_sync(task)
        try:
            # Publish full order payload to WS so admin clients can insert it immediately
            payload = jsonable_encoder(order)
        except Exception:
            payload = {"id": getattr(order, 'id', None)}
        # Ensure 'source' is present on pushed payload when available (from DB row or request)
        try:
            if not payload.get('source'):
                # prefer DB value if available
                try:
                    payload['source'] = getattr(order, 'source', None)
                except Exception:
                    pass
            # fallback to original request-encoded value if still missing
            if not payload.get('source'):
                try:
                    payload['source'] = encoded.get('source') if isinstance(encoded, dict) else None
                except Exception:
                    pass
        except Exception:
            pass
        # Ensure customer_type is present on pushed payload.
        try:
            if not payload.get('customer_type'):
                payload['customer_type'] = getattr(order, 'customer_type', None)
            if not payload.get('customer_type'):
                payload['customer_type'] = encoded.get('customer_type') if isinstance(encoded, dict) else None
            customer_type_payload = str(payload.get('customer_type') or '').strip().lower()
            payload['customer_type'] = customer_type_payload if customer_type_payload in ('mayorista', 'minorista') else 'mayorista'
        except Exception:
            payload['customer_type'] = 'mayorista'
        # Workaround: if the client provided a bearer token, supplement the
        # pushed payload with user info from the token so the admin sees the
        # user's email/name immediately even if the DB could not persist the
        # user_* columns yet (useful while migrations are pending).
        try:
            if token_payload:
                if not payload.get('user_full_name'):
                    payload['user_full_name'] = token_payload.get('full_name') or token_payload.get('name')
                if not payload.get('user_email'):
                    payload['user_email'] = token_payload.get('sub') or token_payload.get('email')
                if not payload.get('user_id'):
                    payload['user_id'] = token_payload.get('id') or token_payload.get('user_id')
                # Add a safe token preview to help diagnose missing token propagation without exposing secrets
                try:
                    payload['_token_received'] = True
                    preview_existing = payload.get('_token_preview') if isinstance(payload.get('_token_preview'), dict) else {}
                    customer_type_preview = str(payload.get('customer_type') or '').strip().lower()
                    payload['_token_preview'] = {
                        **preview_existing,
                        'sub': token_payload.get('sub') or token_payload.get('email') or None,
                        'email': token_payload.get('email') or token_payload.get('sub') or None,
                        'name': token_payload.get('full_name') or token_payload.get('name') or None,
                        'customer_type': customer_type_preview if customer_type_preview in ('mayorista', 'minorista') else None
                    }
                except Exception:
                    payload['_token_received'] = True
                # Try to fetch richer profile from the users table (if present)
                try:
                    uid = payload.get('user_id') or (token_payload.get('id') if isinstance(token_payload, dict) else None)
                    from app import models as _models
                    if uid is not None:
                        db_for_user = SessionLocal()
                        try:
                            # ensure we use integer id when possible
                            try:
                                uid_int = int(uid)
                            except Exception:
                                uid_int = None
                            if uid_int is not None:
                                u = db_for_user.query(_models.User).filter(_models.User.id == uid_int).first()
                            else:
                                u = db_for_user.query(_models.User).filter(_models.User.email == (payload.get('user_email') or token_payload.get('sub') or token_payload.get('email'))).first()
                            if u:
                                payload['user_full_name'] = payload.get('user_full_name') or getattr(u, 'full_name', None)
                                payload['user_email'] = payload.get('user_email') or getattr(u, 'email', None)
                                payload['user_barrio'] = payload.get('user_barrio') or getattr(u, 'barrio', None)
                                payload['user_calle'] = payload.get('user_calle') or getattr(u, 'calle', None)
                                payload['user_numeracion'] = payload.get('user_numeracion') or getattr(u, 'numeracion', None)
                        finally:
                            try: db_for_user.close()
                            except Exception: pass
                except Exception:
                    # non-fatal: if user table doesn't exist or query fails, ignore
                    pass
        except Exception:
            pass
        try:
            payload = _attach_maps_url(payload if isinstance(payload, dict) else {})
        except Exception:
            pass
        try:
            logger.info('create_order succeeded; id=%s created_at=%s payload_summary=%s', getattr(order, 'id', None), getattr(order, 'created_at', None), { 'user': payload.get('user_email') or payload.get('user_full_name'), 'total': payload.get('total') })
        except Exception:
            pass
        # Ensure durability: persist a token preview / contact snapshot into
        # `order_token_previews` so contact info survives service restarts and
        # schema mismatches where orders.user_* columns are missing.
        try:
            # Use the original request payload (logged as `encoded`) as the source
            # of contact fields and token preview, not the `payload` variable which
            # was overwritten with the created order representation.
            try:
                tp_val = encoded.get('_token_preview') if isinstance(encoded, dict) else None
            except Exception:
                tp_val = None
            # If no explicit token preview, build one from available request payload fields
            if not tp_val:
                try:
                    tp_val = {}
                    if isinstance(encoded, dict) and encoded.get('user_full_name'):
                        tp_val['name'] = encoded.get('user_full_name')
                    if isinstance(encoded, dict) and encoded.get('user_email'):
                        tp_val['email'] = encoded.get('user_email')
                    if isinstance(encoded, dict) and encoded.get('user_barrio'):
                        tp_val['barrio'] = encoded.get('user_barrio')
                    if isinstance(encoded, dict) and encoded.get('user_calle'):
                        tp_val['calle'] = encoded.get('user_calle')
                    if isinstance(encoded, dict) and encoded.get('user_numeracion'):
                        tp_val['numeracion'] = encoded.get('user_numeracion')
                    if isinstance(encoded, dict) and encoded.get('user_postal_code'):
                        tp_val['postal_code'] = encoded.get('user_postal_code')
                    if isinstance(encoded, dict) and encoded.get('user_department'):
                        tp_val['department'] = encoded.get('user_department')
                    if not tp_val:
                        tp_val = None
                except Exception:
                    tp_val = None
            # token preview constructed from the request is handled in CRUD layer
        except Exception:
            pass
        # Cache the pushed payload so list queries can show token-preview/user info even if DB lacks columns
        try:
            if payload and payload.get('id'):
                # Ensure cached payload always records the inferred `source` so admin can split Web/App correctly
                try:
                    if not payload.get('source'):
                        payload['source'] = src if 'src' in locals() else (getattr(order, 'source', None) or (encoded.get('source') if isinstance(encoded, dict) else None) or 'web')
                except Exception:
                    try:
                        payload['source'] = getattr(order, 'source', None) or (encoded.get('source') if isinstance(encoded, dict) else None) or 'web'
                    except Exception:
                        payload['source'] = 'web'
                try:
                    customer_type_cache = str(payload.get('customer_type') or '').strip().lower()
                    payload['customer_type'] = customer_type_cache if customer_type_cache in ('mayorista', 'minorista') else 'mayorista'
                except Exception:
                    payload['customer_type'] = 'mayorista'
                ORDER_PAYLOAD_CACHE[str(payload.get('id'))] = {'payload': payload, 'ts': time.time()}
                try:
                    logger.debug('create_order: cached payload for id=%s (has_user=%s, has_token_preview=%s)', payload.get('id'), bool(payload.get('user_full_name') or payload.get('user_email')), bool(payload.get('_token_preview')))
                except Exception:
                    pass
                _prune_order_cache()
        except Exception:
            pass
        await push_event({"action": "order_created", "order": payload})
        # If order creation decremented stock, update snapshot and notify product watchers.
        try:
            # Product stock updates
            updated = getattr(order, '_updated_product_ids', None)
            if updated:
                try:
                    await anyio.to_thread.run_sync(write_catalog_snapshot)
                except Exception:
                    logger.exception('write_catalog_snapshot after order failed')
                try:
                    for pid in updated:
                        await push_event({"action": "updated", "product": {"id": pid}})
                except Exception:
                    logger.exception('push_event product updates after order failed')
            # Consumptions consumed: reduce quantities in consumos.json on disk
            consumed = getattr(order, '_consumos_consumed', None)
            if consumed:
                try:
                    def _apply_consumos(consumed_map):
                        try:
                            path = os.path.join(CATALOG_DIR, 'consumos.json')
                            if not os.path.exists(path):
                                return None
                            try:
                                import fcntl  # Unix-only; best-effort lock
                            except Exception:
                                fcntl = None
                            with open(path, 'r+', encoding='utf-8') as f:
                                if fcntl:
                                    try:
                                        fcntl.flock(f.fileno(), fcntl.LOCK_EX)
                                    except Exception:
                                        pass
                                try:
                                    f.seek(0)
                                    cur = json.load(f) or []
                                except Exception:
                                    cur = []
                                # Build map of existing entries
                                emap = { int(c.get('id')): c for c in cur }
                                changed = False
                                for pid_s, delta in consumed_map.items():
                                    try:
                                        pid = int(pid_s)
                                    except Exception:
                                        continue
                                    if pid not in emap:
                                        continue
                                    try:
                                        curqty = int(emap[pid].get('qty', 0) or 0)
                                    except Exception:
                                        curqty = 0
                                    newqty = max(0, curqty - int(delta or 0))
                                    if newqty <= 0:
                                        del emap[pid]
                                        changed = True
                                    else:
                                        emap[pid]['qty'] = newqty
                                        changed = True
                                if changed:
                                    new_list = list(emap.values())
                                    try:
                                        f.seek(0)
                                        f.truncate()
                                        json.dump(new_list, f, ensure_ascii=False, indent=2)
                                        f.flush()
                                    except Exception:
                                        pass
                                    return new_list
                                return None
                        except Exception:
                            return None
                    new_consumos = await anyio.to_thread.run_sync(lambda: _apply_consumos(consumed))
                    if new_consumos is not None:
                        try:
                            await push_event({"action": "consumos-updated", "consumos": new_consumos})
                        except Exception:
                            logger.exception('push_event consumos-updated after order failed')
                        try:
                            await anyio.to_thread.run_sync(write_catalog_snapshot)
                        except Exception:
                            logger.exception('write_catalog_snapshot after consumos update failed')
                except Exception:
                    logger.exception('apply consumos after order failed')
        except Exception:
            logger.exception('post-order product update notifications failed')
        # Send order confirmation email in background so checkout response is not blocked.
        try:
            order_for_email = dict(payload) if isinstance(payload, dict) else {}
            request_for_email = dict(encoded) if isinstance(encoded, dict) else None
            token_for_email = dict(token_payload) if isinstance(token_payload, dict) else None
            asyncio.create_task(
                _send_order_confirmation_email(
                    order_data=order_for_email,
                    request_data=request_for_email,
                    token_payload=token_for_email,
                )
            )
        except Exception:
            logger.exception('Could not schedule order confirmation email for id=%s', getattr(order, 'id', None))
        # Return an explicit JSON response so a response-model mismatch never
        # turns a successfully persisted order into a 500 for the checkout.
        try:
            response_payload = payload if isinstance(payload, dict) else None
            if not isinstance(response_payload, dict):
                response_payload = jsonable_encoder(order)
            if not isinstance(response_payload, dict):
                response_payload = {
                    'id': getattr(order, 'id', None),
                    'items': getattr(order, 'items', []),
                    'total': getattr(order, 'total', 0),
                    'status': getattr(order, 'status', 'recibido'),
                    'created_at': getattr(order, 'created_at', None),
                }
            response_payload = _attach_maps_url(response_payload)
            headers = _cors_headers_for_request(request)
            return JSONResponse(status_code=200, content=response_payload, headers=headers)
        except Exception:
            return order
    except HTTPException:
        raise
    except Exception as e:
        logger.exception('Unexpected error creating order (outer): %s', e)
        # Emergency fallback: persist a minimal order snapshot so checkout does
        # not fail for the customer when the full order flow crashes.
        try:
            fallback_payload = encoded if isinstance(encoded, dict) else {
                'items': getattr(payload, 'items', []),
                'total': getattr(payload, 'total', 0),
                'user_email': getattr(payload, 'user_email', None),
                'user_full_name': getattr(payload, 'user_full_name', None),
                'user_barrio': getattr(payload, 'user_barrio', None),
                'user_calle': getattr(payload, 'user_calle', None),
                'user_numeracion': getattr(payload, 'user_numeracion', None),
                'user_postal_code': getattr(payload, 'user_postal_code', None),
                'user_department': getattr(payload, 'user_department', None),
            }

            def emergency_task():
                db = SessionLocal()
                try:
                    return crud.create_order_minimal(db, fallback_payload, current_user=token_payload if isinstance(token_payload, dict) else None)
                finally:
                    db.close()

            emergency_order = await anyio.to_thread.run_sync(emergency_task)
            emergency_payload = jsonable_encoder(emergency_order)
            if isinstance(emergency_payload, dict):
                emergency_payload['_fallback'] = True
                emergency_payload = _attach_maps_url(emergency_payload)
            headers = _cors_headers_for_request(request)
            return JSONResponse(status_code=200, content=emergency_payload, headers=headers)
        except Exception:
            logger.exception('create_order emergency fallback failed')
            raise HTTPException(status_code=500, detail='Could not create order')


@app.post('/debug/orders')
async def debug_create_order(payload: schemas.OrderCreate):
    """Debug-only endpoint: attempt to create order and return full details for diagnosis."""
    try:
        encoded = jsonable_encoder(payload)
    except Exception:
        encoded = { 'items': getattr(payload, 'items', None), 'total': getattr(payload, 'total', None) }
    logger.info('debug_create_order called; payload=%s', encoded)

    def task():
        db = SessionLocal()
        try:
            return crud.create_order(db, payload, current_user=None)
        finally:
            db.close()

    order = await anyio.to_thread.run_sync(task)
    return order


@app.post('/debug/orders-diagnose')
async def debug_create_order_diagnose(request: Request):
    """Verbose diagnostic endpoint: attempts to create an order and returns
    either the created object or a detailed error trace (appends trace to
    Backend/server_log.txt for further inspection).
    """
    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail='Invalid JSON body')

    # Try to coerce into OrderCreate schema for consistency with create flow
    try:
        order_schema = schemas.OrderCreate.parse_obj(data)
    except Exception:
        try:
            order_schema = schemas.OrderCreate(**data)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f'Invalid order payload: {e}')

    def task_payload(sch=order_schema):
        db = SessionLocal()
        try:
            return crud.create_order(db, sch, current_user=None)
        finally:
            try:
                db.close()
            except Exception:
                pass

    try:
        created = await anyio.to_thread.run_sync(task_payload)
        try:
            return JSONResponse(status_code=200, content=jsonable_encoder(created))
        except Exception:
            return JSONResponse(status_code=200, content={ 'id': getattr(created, 'id', None) })
    except Exception as e:
        import traceback, datetime, os as _os
        tb = traceback.format_exc()
        logger.exception('debug/orders-diagnose failed: %s', e)
        # Append trace to Backend/server_log.txt next to the repository Backend folder
        try:
            base = _os.path.dirname(_os.path.dirname(__file__))
            logpath = _os.path.join(base, 'server_log.txt')
            with open(logpath, 'a', encoding='utf-8') as f:
                f.write(f"{datetime.datetime.utcnow().isoformat()} - DEBUG_CREATE_ORDER_ERROR - {str(e)}\n")
                f.write(tb + "\n\n")
        except Exception:
            pass
        # Return limited trace in response to help local debugging (trim to 1000 chars)
        safe_tb = tb[:1000]
        raise HTTPException(status_code=500, detail={ 'error': str(e), 'trace': safe_tb })


@app.get('/debug/db-info')
def debug_db_info(request: Request):
    """Return basic DB diagnostic info: dialect, masked DATABASE_URL and orders columns."""
    try:
        db_env = os.environ.get('DATABASE_URL')
        masked = 'sqlite (local file)'
        if db_env:
            try:
                from urllib.parse import urlparse
                u = urlparse(db_env)
                user = u.username or ''
                host = u.hostname or ''
                port = u.port or ''
                path = u.path or ''
                masked = f"{u.scheme}://{user + ':****@' if user else ''}{host}{(':'+str(port)) if port else ''}{path}"
            except Exception:
                masked = 'postgres (masked)'
    except Exception:
        masked = 'unknown'

    try:
        insp = inspect(engine)
        cols = {c['name'] for c in insp.get_columns('orders')}
    except Exception:
        cols = set()

    try:
        dialect = getattr(engine, 'dialect', None)
        dialect_name = getattr(dialect, 'name', '') if dialect else ''
    except Exception:
        dialect_name = ''

    headers = _cors_headers_for_request(request)
    return JSONResponse(status_code=200, content={'database_url': masked, 'dialect': dialect_name, 'orders_columns': sorted(list(cols))}, headers=headers)


@app.get('/debug/db-test')
def debug_db_test(request: Request):
    """Perform a small set of DB connectivity checks and return results.
    Useful to diagnose connection/auth issues on PaaS (Render).
    """
    out = {'ok': False, 'checks': []}
    headers = _cors_headers_for_request(request)
    try:
        # Try basic connect
        try:
            conn = engine.connect()
            try:
                out['checks'].append({'connect': True})
            finally:
                conn.close()
        except Exception as e:
            out['checks'].append({'connect': False, 'error': str(e)})
            # append to server log
            try:
                import traceback, datetime, os as _os
                base = _os.path.dirname(_os.path.dirname(__file__))
                logpath = _os.path.join(base, 'server_log.txt')
                with open(logpath, 'a', encoding='utf-8') as f:
                    f.write(f"{datetime.datetime.utcnow().isoformat()} - DB_CONNECT_FAILED - {str(e)}\n")
                    f.write(traceback.format_exc() + "\n\n")
            except Exception:
                pass
            return JSONResponse(status_code=500, content=out, headers=headers)

        # Try select 1
        try:
            conn2 = engine.connect()
            try:
                try:
                    r = conn2.execute(text('SELECT 1')).fetchone()
                    out['checks'].append({'select_1': True, 'result': list(r) if r is not None else None})
                except Exception as e:
                    try:
                        _invalidate_conn(conn2)
                    except Exception:
                        pass
                    out['checks'].append({'select_1': False, 'error': str(e)})
            finally:
                conn2.close()
        except Exception as e:
            out['checks'].append({'select_1': False, 'error': str(e)})

        # Inspect orders table existence
        try:
            insp = inspect(engine)
            cols = {c['name'] for c in insp.get_columns('orders')} if insp and insp.get_columns('orders') is not None else set()
            out['checks'].append({'orders_table_columns_count': len(cols), 'orders_columns_sample': list(cols)[:10]})
        except Exception as e:
            out['checks'].append({'orders_table': False, 'error': str(e)})

        out['ok'] = True
        return JSONResponse(status_code=200, content=out, headers=headers)
    except Exception as e:
        out['checks'].append({'error': str(e)})
        return JSONResponse(status_code=500, content=out, headers=headers)


@app.get('/debug/token-previews')
def debug_token_previews():
    """Return recent rows from `order_token_previews` for debugging/verification."""
    try:
        rows = _safe_engine_fetchall('SELECT order_id, token_preview, token_received, created_at FROM order_token_previews ORDER BY created_at DESC LIMIT 50') or []
    except Exception:
        return []
    out = []
    import json as _json
    for r in rows:
        try:
            tp_raw = r[1]
            try:
                tp = _json.loads(tp_raw) if isinstance(tp_raw, str) else tp_raw
            except Exception:
                tp = tp_raw
        except Exception:
            tp = None
        out.append({'order_id': r[0], 'token_preview': tp, 'token_received': bool(r[2]), 'created_at': str(r[3])})
    return out


@app.get('/debug/server-log')
async def debug_server_log(request: Request):
    """Return last 200 lines of server_log.txt (requires MIGRATION_SECRET header)."""
    secret = os.environ.get('MIGRATION_SECRET')
    provided = request.headers.get('x-migrate-secret')
    if secret and provided != secret:
        return JSONResponse(status_code=403, content={'error': 'forbidden'})
    try:
        path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'server_log.txt')
        if not os.path.exists(path):
            return JSONResponse(status_code=200, content={'lines': []})
        with open(path, 'r', encoding='utf-8') as f:
            lines = f.read().splitlines()
        return JSONResponse(status_code=200, content={'lines': lines[-200:]})
    except Exception as e:
        logger.exception('debug_server_log failed: %s', e)
        return JSONResponse(status_code=500, content={'error': str(e)})


@app.post('/backup-orders')
async def backup_orders(request: Request):
    """Accept one or many order payloads and persist them to the DB so they are
    durable and visible to the admin panel. This endpoint is intentionally
    permissive (no auth) to allow offline/failed clients to sync back their local
    queue when connectivity is restored. Use with care in public deployments.
    """
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail='Invalid JSON body')
    items = body if isinstance(body, list) else [body]
    results = []
    for it in items:
        try:
            request_payload = dict(it) if isinstance(it, dict) else {}
            inferred_email = _extract_customer_email_for_order(
                order_data=request_payload,
                request_data=request_payload,
                token_payload=None,
            )
            if not inferred_email and isinstance(request_payload, dict) and request_payload.get('user_id') is not None:
                inferred_email = _resolve_user_email_by_id(request_payload.get('user_id'))
            require_email_raw = (os.environ.get('ORDER_REQUIRE_CUSTOMER_EMAIL') or 'true').strip().lower()
            require_email = require_email_raw not in ('0', 'false', 'no', 'off')
            if require_email and not inferred_email:
                results.append({'ok': False, 'error': 'missing_user_email'})
                continue
            if inferred_email and isinstance(request_payload, dict) and not request_payload.get('user_email'):
                request_payload['user_email'] = inferred_email
            payload_to_parse = request_payload if isinstance(request_payload, dict) and request_payload else it
            # Normalize into OrderCreate schema
            try:
                order_schema = schemas.OrderCreate.parse_obj(payload_to_parse)
            except Exception:
                # Try simple coercion for common shapes
                order_schema = schemas.OrderCreate(**payload_to_parse)
            def task_payload(sch=order_schema):
                db = SessionLocal()
                try:
                    return crud.create_order(db, sch, current_user=None)
                finally:
                    db.close()
            created = await anyio.to_thread.run_sync(task_payload)
            try:
                created_payload = jsonable_encoder(created)
            except Exception:
                created_payload = {
                    'id': getattr(created, 'id', None),
                    'items': getattr(created, 'items', None),
                    'total': getattr(created, 'total', None),
                    'user_email': getattr(created, 'user_email', None),
                    'created_at': getattr(created, 'created_at', None),
                }
            try:
                asyncio.create_task(
                    _send_order_confirmation_email(
                        order_data=created_payload if isinstance(created_payload, dict) else {},
                        request_data=request_payload if isinstance(request_payload, dict) else None,
                        token_payload=None,
                    )
                )
            except Exception:
                logger.exception(
                    'backup_orders: could not schedule confirmation email for id=%s',
                    created_payload.get('id') if isinstance(created_payload, dict) else None,
                )
            results.append({'ok': True, 'id': getattr(created, 'id', None)})
        except Exception as e:
            logger.exception('backup_orders: failed to persist order: %s', e)
            # Emergency fallback: persist a minimal snapshot so client-side
            # failed-order queues can still be drained.
            try:
                emergency_payload = request_payload if isinstance(request_payload, dict) else (it if isinstance(it, dict) else {})
                def task_emergency(p=emergency_payload):
                    db = SessionLocal()
                    try:
                        return crud.create_order_minimal(db, p, current_user=None)
                    finally:
                        db.close()
                emergency = await anyio.to_thread.run_sync(task_emergency)
                results.append({'ok': True, 'id': getattr(emergency, 'id', None), 'fallback': 'minimal'})
                continue
            except Exception as fallback_err:
                logger.exception('backup_orders: minimal fallback failed: %s', fallback_err)
            results.append({'ok': False, 'error': str(e)})
    return {'saved': len([r for r in results if r.get('ok')]), 'results': results}


@app.get('/orders', response_model=List[schemas.OrderResponse])
def list_orders(
    skip: int = 0,
    limit: int = 200,
    source: Optional[str] = None,
    q: Optional[str] = None,
    date: Optional[str] = None,
    db: Session = Depends(get_db),
):
    # Fetch recent orders (safe select in CRUD). Then merge any cached pushed payloads
    # (which may contain token preview) to surface user info when DB lacks columns.
    rows = crud.get_orders(db, skip, limit, source=source, q=q, date=date)
    try:
        _prune_order_cache()
    except Exception:
        pass
    out = []
    # Load status overrides (best-effort) so admin can show "visto" even if DB column is missing.
    status_overrides = {}
    try:
        status_overrides = crud.get_setting(db, 'order_status_overrides') or {}
    except Exception:
        status_overrides = {}
    try:
        _prune_status_cache()
    except Exception:
        pass

    now_utc = datetime.datetime.now(datetime.timezone.utc)
    status_updates: List[Tuple[int, str]] = []

    product_price_map = {}
    try:
        product_rows = db.query(models.Product.id, models.Product.price, models.Product.price_retail).all()
        for pr in (product_rows or []):
            try:
                pid = str(getattr(pr, 'id', '')).strip()
                if not pid:
                    continue
                wholesale = float(getattr(pr, 'price', 0) or 0)
                retail_raw = getattr(pr, 'price_retail', None)
                retail = None if retail_raw is None else float(retail_raw)
                product_price_map[pid] = {'wholesale': wholesale, 'retail': retail}
            except Exception:
                continue
    except Exception:
        product_price_map = {}

    def _infer_customer_type_from_items(items_value):
        try:
            items_arr = items_value if isinstance(items_value, list) else []
            retail_hits = 0
            wholesale_hits = 0
            for it in items_arr:
                if not isinstance(it, dict):
                    continue
                pid = str(it.get('id') or '').strip()
                if not pid:
                    continue
                meta = it.get('meta') if isinstance(it.get('meta'), dict) else {}
                line_price_raw = meta.get('price')
                if line_price_raw is None:
                    continue
                try:
                    line_price = float(line_price_raw)
                except Exception:
                    continue
                prod_prices = product_price_map.get(pid)
                if not prod_prices:
                    continue
                retail_price = prod_prices.get('retail')
                wholesale_price = prod_prices.get('wholesale')
                if retail_price is None:
                    continue
                try:
                    retail_price = float(retail_price)
                    wholesale_price = float(wholesale_price or 0)
                except Exception:
                    continue
                if abs(retail_price - wholesale_price) <= 1e-9:
                    continue
                dr = abs(line_price - retail_price)
                dw = abs(line_price - wholesale_price)
                tolerance = 0.05
                if dr <= tolerance and dr <= dw:
                    retail_hits += 1
                elif dw <= tolerance and dw < dr:
                    wholesale_hits += 1
            if retail_hits > 0 and retail_hits >= wholesale_hits:
                return 'minorista'
            if wholesale_hits > 0:
                return 'mayorista'
        except Exception:
            pass
        return None

    def _normalize_customer_type(value):
        v = str(value or '').strip().lower()
        return v if v in ('mayorista', 'minorista') else ''

    def _customer_type_from_preview(preview_value):
        try:
            preview = preview_value if isinstance(preview_value, dict) else {}
            for key in ('customer_type', 'perfil', 'profile', 'price_tier', 'tier'):
                ct = _normalize_customer_type(preview.get(key))
                if ct:
                    return ct
        except Exception:
            pass
        return ''

    cached_any = 0
    for r in (rows or []):
        try:
            # r may be a dict (raw select) or an object (ORM fallback). Normalize to dict
            od = r if isinstance(r, dict) else {
                k: getattr(r, k, None)
                for k in [
                    'id', 'items', 'total', 'status', 'customer_type', 'user_id', 'user_full_name', 'user_email',
                    'user_barrio', 'user_calle', 'user_numeracion', 'user_postal_code', 'user_department', 'created_at',
                    '_token_received', '_token_preview', 'source',
                    'payment_method', 'payment_status', 'payment_reference',
                    'scheduled_delivery_date', 'delivery_cutoff_applied', 'delivery_timezone', 'delivery_cutoff_hour',
                    'assigned_driver_id', 'assigned_driver_username', 'assigned_driver_name', 'assigned_driver_zone', 'assigned_at',
                    'delivery_lat', 'delivery_lon',
                    'route_id', 'route_order', 'route_generated_at',
                    'delivered_at', 'delivered_by_id', 'delivered_by_username',
                ]
            }
            customer_type_value = _normalize_customer_type(od.get('customer_type'))
            if not customer_type_value:
                customer_type_value = _customer_type_from_preview(od.get('_token_preview'))
            if not customer_type_value:
                try:
                    cached_ct = ORDER_PAYLOAD_CACHE.get(str(od.get('id'))) if od.get('id') else None
                    if cached_ct and isinstance(cached_ct.get('payload'), dict):
                        customer_type_value = _normalize_customer_type(cached_ct['payload'].get('customer_type'))
                        if not customer_type_value:
                            customer_type_value = _customer_type_from_preview(cached_ct['payload'].get('_token_preview'))
                except Exception:
                    pass
            if not customer_type_value:
                inferred_ct = _infer_customer_type_from_items(od.get('items'))
                if inferred_ct in ('mayorista', 'minorista'):
                    customer_type_value = inferred_ct
            if not od.get('scheduled_delivery_date') and od.get('id'):
                try:
                    cached_delivery = ORDER_PAYLOAD_CACHE.get(str(od.get('id')))
                    if cached_delivery and isinstance(cached_delivery.get('payload'), dict):
                        cp = cached_delivery['payload']
                        for f in (
                            'scheduled_delivery_date',
                            'delivery_cutoff_applied',
                            'delivery_timezone',
                            'delivery_cutoff_hour',
                        ):
                            if od.get(f) is None and cp.get(f) is not None:
                                od[f] = cp.get(f)
                except Exception:
                    pass
            cached_used = False
            # If user fields missing try to merge from cached pushed payload
            if (not od.get('user_full_name') and not od.get('user_email')) and od.get('id'):
                try:
                    cached = ORDER_PAYLOAD_CACHE.get(str(od.get('id')))
                    if cached and cached.get('payload'):
                        p = cached.get('payload')
                        cached_used = True
                        cached_any += 1
                        # prefer explicit user_* fields from cached payload
                        for f in (
                            'customer_type',
                            'user_full_name','user_email','user_barrio','user_calle','user_numeracion','user_postal_code','user_department','user_id',
                            'source','payment_method','payment_status','payment_reference',
                            'scheduled_delivery_date', 'delivery_cutoff_applied', 'delivery_timezone', 'delivery_cutoff_hour',
                            'assigned_driver_id', 'assigned_driver_username', 'assigned_driver_name', 'assigned_driver_zone', 'assigned_at',
                            'delivery_lat', 'delivery_lon',
                            'route_id', 'route_order', 'route_generated_at',
                            'delivered_at', 'delivered_by_id', 'delivered_by_username',
                        ):
                            if not od.get(f) and p.get(f):
                                od[f] = p.get(f)
                        # fallback to token preview when available
                        tp = p.get('_token_preview') or {}
                        tp_addr = tp.get('address') if isinstance(tp.get('address'), dict) else {}
                        if not od.get('user_full_name') and tp.get('name'):
                            od['user_full_name'] = tp.get('name')
                        if not od.get('user_email') and tp.get('email'):
                            od['user_email'] = tp.get('email')
                        if not od.get('user_postal_code'):
                            od['user_postal_code'] = (
                                tp.get('postal_code') or tp.get('user_postal_code') or
                                tp_addr.get('postal_code') or tp_addr.get('postcode')
                            )
                        if not od.get('user_department'):
                            od['user_department'] = tp.get('department') or tp.get('user_department') or tp_addr.get('department')
                        if not customer_type_value:
                            customer_type_value = _customer_type_from_preview(tp)
                        # include the token preview in the returned row so clients can show it explicitly
                        if tp:
                            od['_token_preview'] = tp
                            od['_token_received'] = True
                except Exception:
                    pass

                # If still missing user info try to read persisted token previews table
                # (durable across restarts) and merge into the row.
                if (not od.get('user_full_name') and not od.get('user_email')) and od.get('id'):
                    try:
                        # order_id stored as text in order_token_previews to support debug ids
                        row = _safe_engine_fetchone('SELECT token_preview, token_received FROM order_token_previews WHERE order_id = :id ORDER BY created_at DESC LIMIT 1', {'id': str(od.get('id'))})
                        if row:
                            try:
                                tp_raw = row[0]
                                tr_flag = row[1]
                                if tp_raw:
                                    try:
                                        tp = json.loads(tp_raw) if isinstance(tp_raw, str) else tp_raw
                                    except Exception:
                                        tp = {}
                                    if not od.get('user_full_name') and tp.get('name'):
                                        od['user_full_name'] = tp.get('name')
                                    if not od.get('user_email') and tp.get('email'):
                                        od['user_email'] = tp.get('email')
                                    tp_addr = tp.get('address') if isinstance(tp.get('address'), dict) else {}
                                    if not od.get('user_postal_code'):
                                        od['user_postal_code'] = (
                                            tp.get('postal_code') or tp.get('user_postal_code') or
                                            tp_addr.get('postal_code') or tp_addr.get('postcode')
                                        )
                                    if not od.get('user_department'):
                                        od['user_department'] = tp.get('department') or tp.get('user_department') or tp_addr.get('department')
                                    if not customer_type_value:
                                        customer_type_value = _customer_type_from_preview(tp)
                                    od['_token_preview'] = tp
                                    od['_token_received'] = bool(tr_flag)
                                    cached_used = True
                                    cached_any += 1
                            except Exception:
                                pass
                    except Exception:
                        pass
            od['customer_type'] = customer_type_value or 'mayorista'
            # Merge DB status + override + auto milestones (scheduled delivery day 09:00/16:00).
            try:
                oid = od.get('id')
                db_status_raw = od.get('status')
                override_raw = None
                if oid is not None:
                    try:
                        override_raw = (status_overrides or {}).get(str(oid))
                    except Exception:
                        override_raw = None
                    if override_raw is None:
                        override_raw = (ORDER_STATUS_CACHE.get(str(oid)) or {}).get('status')

                base_status = _merge_order_statuses(db_status_raw, override_raw)
                auto_status = _compute_order_auto_status(
                    od.get('created_at'),
                    scheduled_delivery_date_value=od.get('scheduled_delivery_date'),
                    status_value=base_status,
                    now_utc=now_utc,
                    tz_name_hint=str(od.get('delivery_timezone') or '').strip() or None,
                )
                effective_status = _merge_order_statuses(base_status, auto_status)
                od['status'] = effective_status

                # Best-effort persist auto advancement to DB (no schedulers required).
                try:
                    if auto_status and effective_status == auto_status and effective_status in ('enviado', 'entregado'):
                        if _order_status_rank(effective_status) > _order_status_rank(db_status_raw):
                            status_updates.append((int(oid), str(effective_status)))
                except Exception:
                    pass
            except Exception:
                pass
            try:
                od = _attach_maps_url(od)
            except Exception:
                pass
            # log per-row diagnostics for debugging clarity
            try:
                logger.debug('list_orders row id=%s user_full_name=%s user_email=%s cached_used=%s', od.get('id'), od.get('user_full_name'), od.get('user_email'), cached_used)
            except Exception:
                pass
            out.append(od)
        except Exception:
            out.append(r)

    # Persist any auto-advanced statuses so subsequent reads are consistent.
    try:
        if status_updates:
            uniq: Dict[int, str] = {}
            for oid, st in (status_updates or []):
                try:
                    uniq[int(oid)] = str(st)
                except Exception:
                    continue
            if uniq:
                with engine.begin() as conn:
                    for oid, st in uniq.items():
                        try:
                            conn.execute(text('UPDATE orders SET status = :status WHERE id = :id'), {'status': st, 'id': oid})
                        except Exception:
                            continue
    except Exception:
        pass
    try:
        logger.info('list_orders returning %d rows (sample ids %s)', len(out), [str(x.get('id')) for x in (out or [])][:10])
        if cached_any:
            logger.info('list_orders merged cached payloads for %d rows', cached_any)
    except Exception:
        pass
    return out


@app.get('/admin/orders', response_model=List[schemas.OrderResponse])
def admin_list_orders(
    skip: int = 0,
    limit: int = 200,
    source: Optional[str] = None,
    q: Optional[str] = None,
    date: Optional[str] = None,
    status: Optional[str] = None,
    zone: Optional[str] = None,
    driver_id: Optional[str] = None,
    driver_username: Optional[str] = None,
    auto: Optional[str] = None,
    current_admin=Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    # Reuse existing list_orders logic, then apply role-based filters.
    orders = list_orders(skip=skip, limit=limit, source=source, q=q, date=date, db=db)
    role = str(getattr(current_admin, 'role', '') or '').strip().lower()
    admin_id = getattr(current_admin, 'id', None)
    admin_username = getattr(current_admin, 'username', None)

    # Status filter
    status_values: List[str] = []
    if status:
        for part in str(status).split(','):
            st = _normalize_order_status_strict(part)
            if st:
                status_values.append(st)
    if not status_values and role == 'repartidor':
        # Default: show delivery-stage orders once prepared.
        status_values = ['preparado', 'enviado']

    if status_values:
        orders = [o for o in (orders or []) if _normalize_order_status(o.get('status')) in status_values]

    # Assignment filter
    target_driver_id = None
    target_driver_username = None
    if role == 'repartidor':
        target_driver_id = admin_id
        target_driver_username = admin_username
    else:
        if driver_id:
            try:
                target_driver_id = int(driver_id)
            except Exception:
                target_driver_id = None
        if driver_username:
            target_driver_username = str(driver_username or '').strip()

    auto_flag = True
    if auto is not None:
        auto_flag = str(auto).strip().lower() not in ('0', 'false', 'no', 'off')

    if role == 'repartidor':
        # Auto-assign prepared orders in zone when requested (default true).
        if auto_flag:
            try:
                _auto_assign_routes_for_driver(current_admin, orders, include_assigned=True, db=db)
                # refresh list after auto-assign
                orders = list_orders(skip=skip, limit=limit, source=source, q=q, date=date, db=db)
            except Exception:
                logger.exception('auto-assign for repartidor failed')

    if target_driver_id is not None or target_driver_username:
        orders = [
            o for o in (orders or [])
            if (
                (target_driver_id is not None and str(o.get('assigned_driver_id') or '') == str(target_driver_id)) or
                (target_driver_username and str(o.get('assigned_driver_username') or '').strip() == str(target_driver_username))
            )
        ]
    elif role == 'repartidor':
        return []

    # Zone filter (admins/owners only). Repartidores are filtered by assignment.
    if role != 'repartidor':
        zone_tokens: List[str] = _parse_admin_zones(zone) if zone else []
        if zone_tokens:
            orders = [o for o in (orders or []) if _order_matches_zone(o, zone_tokens)]

    # Prefer route order when present
    try:
        if orders and any(o.get('route_order') is not None for o in orders):
            orders = sorted(
                orders,
                key=lambda o: (o.get('route_order') is None, int(o.get('route_order') or 0), str(o.get('id') or '')),
            )
    except Exception:
        pass

    return orders


@app.get('/admin/zones')
def admin_zones(current_admin=Depends(get_current_admin_user)):
    return {'zones': _MENDOZA_DEPARTMENTS}


def _auto_assign_routes_for_driver(driver, orders: List[Dict[str, Any]], include_assigned: bool = True, db: Optional[Session] = None) -> Dict[str, Any]:
    driver_id = getattr(driver, 'id', None)
    driver_username = getattr(driver, 'username', None)
    driver_zone = str(getattr(driver, 'zone', '') or '').strip()
    if not driver_id or not driver_zone:
        return {'driver_id': driver_id, 'assigned': 0, 'route_id': None, 'assigned_ids': []}
    zone_tokens = _parse_admin_zones(driver_zone)
    now = datetime.datetime.now(datetime.timezone.utc)
    candidates = []
    for od in (orders or []):
        try:
            status_val = _normalize_order_status(od.get('status'))
            if status_val not in ('preparado', 'enviado'):
                continue
            if status_val == 'entregado':
                continue
            # Skip if assigned to other driver
            assigned_id = od.get('assigned_driver_id')
            assigned_user = str(od.get('assigned_driver_username') or '').strip()
            if assigned_id and str(assigned_id) != str(driver_id):
                continue
            if assigned_user and assigned_user != str(driver_username or ''):
                continue
            if not include_assigned and (assigned_id or assigned_user):
                continue
            if zone_tokens and not _order_matches_zone(od, zone_tokens):
                continue
            candidates.append(od)
        except Exception:
            continue

    if not candidates:
        return {'driver_id': driver_id, 'assigned': 0, 'route_id': None, 'assigned_ids': []}

    # Build nodes with coords for optimization
    nodes = []
    no_coords = []
    for od in candidates:
        lat = None
        lon = None
        try:
            snap = _extract_order_address_snapshot(od if isinstance(od, dict) else {})
            lat = snap.get('lat')
            lon = snap.get('lon')
        except Exception:
            lat = None
            lon = None
        if not _is_mendoza_point(lat, lon):
            try:
                lat, lon = _resolve_order_coords(od, db=db)
            except Exception:
                lat, lon = None, None
        if not _is_mendoza_point(lat, lon):
            no_coords.append(od)
            continue
        nodes.append({'order': od, 'lat': float(lat), 'lon': float(lon)})

    if nodes:
        start_lat, start_lon = _get_route_start_point(db=db)
        if _is_mendoza_point(start_lat, start_lon):
            optimized_nodes = _optimize_route_order_with_start(nodes, float(start_lat), float(start_lon))
        else:
            optimized_nodes = _optimize_route_order(nodes)
        ordered_orders = [n['order'] for n in optimized_nodes] + no_coords
    else:
        ordered_orders = candidates

    route_id = f"route-{driver_id}-{now.strftime('%Y%m%d')}-{int(time.time())}"
    updated_ids = []
    try:
        with engine.begin() as conn:
            for idx, od in enumerate(ordered_orders, start=1):
                oid = od.get('id')
                if oid is None:
                    continue
                params = {
                    'assigned_driver_id': int(driver_id),
                    'assigned_driver_username': str(driver_username or '').strip(),
                    'assigned_driver_name': str(getattr(driver, 'full_name', '') or driver_username or '').strip() or None,
                    'assigned_driver_zone': driver_zone,
                    'assigned_at': now if not od.get('assigned_driver_id') else od.get('assigned_at') or now,
                    'route_id': route_id,
                    'route_order': int(idx),
                    'route_generated_at': now,
                    'id': oid,
                }
                if str(oid).isdigit():
                    conn.execute(
                        text(
                            "UPDATE orders SET assigned_driver_id = :assigned_driver_id, "
                            "assigned_driver_username = :assigned_driver_username, "
                            "assigned_driver_name = :assigned_driver_name, "
                            "assigned_driver_zone = :assigned_driver_zone, "
                            "assigned_at = :assigned_at, "
                            "route_id = :route_id, route_order = :route_order, route_generated_at = :route_generated_at "
                            "WHERE id = :id"
                        ),
                        params,
                    )
                else:
                    conn.execute(
                        text(
                            "UPDATE orders SET assigned_driver_id = :assigned_driver_id, "
                            "assigned_driver_username = :assigned_driver_username, "
                            "assigned_driver_name = :assigned_driver_name, "
                            "assigned_driver_zone = :assigned_driver_zone, "
                            "assigned_at = :assigned_at, "
                            "route_id = :route_id, route_order = :route_order, route_generated_at = :route_generated_at "
                            "WHERE CAST(id AS TEXT) = :id"
                        ),
                        params,
                    )
                updated_ids.append(oid)
    except Exception:
        logger.exception('auto-assign route update failed')
    return {'driver_id': driver_id, 'assigned': len(updated_ids), 'route_id': route_id, 'assigned_ids': updated_ids}


def _select_driver_for_order(order: Dict[str, Any], db: Session) -> Optional[Any]:
    """Pick the best repartidor for an order based on zone and current load."""
    try:
        snapshot = _extract_order_address_snapshot(order if isinstance(order, dict) else {})
    except Exception:
        snapshot = {}
    dept = _normalize_region_token(snapshot.get('department'))
    barrio = _normalize_region_token(snapshot.get('barrio'))
    raw_addr = _normalize_region_token(snapshot.get('raw_address'))
    postal = _normalize_postal_code(snapshot.get('postal_code'))
    if not (dept or barrio or raw_addr or postal):
        return None

    drivers = [u for u in (crud.list_admin_users(db) or []) if str(getattr(u, 'role', '') or '').strip().lower() == 'repartidor']
    if not drivers:
        return None

    matching = []
    for d in drivers:
        zone_tokens = _parse_admin_zones(getattr(d, 'zone', None))
        if not zone_tokens:
            continue
        # quick match by department or barrio
        for z in zone_tokens:
            if dept and z == _normalize_region_token(dept):
                matching.append(d)
                break
            if barrio and z in barrio:
                matching.append(d)
                break
            if raw_addr and z in raw_addr:
                matching.append(d)
                break
            deps_for_postal = _departments_for_postal(postal)
            if any(_normalize_region_token(dep) == z for dep in deps_for_postal):
                matching.append(d)
                break

    if not matching:
        return None
    if len(matching) == 1:
        return matching[0]

    # Compute load per driver
    orders = list_orders(skip=0, limit=2000, source=None, q=None, date=None, db=db)
    loads: Dict[str, int] = {}
    for o in (orders or []):
        try:
            st = _normalize_order_status(o.get('status'))
            if st not in ('preparado', 'enviado'):
                continue
            did = o.get('assigned_driver_id')
            if did is None:
                continue
            key = str(did)
            loads[key] = loads.get(key, 0) + 1
        except Exception:
            continue
    matching.sort(key=lambda d: loads.get(str(getattr(d, 'id', '')), 0))
    return matching[0]


@app.post('/admin/routes/auto-assign')
async def admin_auto_assign_routes(request: Request, current_admin=Depends(get_current_admin_user), db: Session = Depends(get_db)):
    role = str(getattr(current_admin, 'role', '') or '').strip().lower()
    if role not in ('owner', 'admin'):
        raise HTTPException(status_code=403, detail='No autorizado')
    try:
        body = await request.json()
    except Exception:
        body = {}
    driver_id = body.get('driver_id')
    driver_username = body.get('driver_username')
    include_assigned = body.get('include_assigned', True)

    drivers = []
    if driver_id or driver_username:
        driver = None
        if driver_id is not None:
            try:
                driver = crud.get_admin_user_by_id(db, int(driver_id))
            except Exception:
                driver = None
        if not driver and driver_username:
            driver = crud.get_admin_user_by_username(db, str(driver_username))
        if not driver:
            raise HTTPException(status_code=404, detail='Repartidor no encontrado')
        if str(getattr(driver, 'role', '') or '').strip().lower() != 'repartidor':
            raise HTTPException(status_code=400, detail='Usuario no es repartidor')
        drivers = [driver]
    else:
        drivers = [u for u in (crud.list_admin_users(db) or []) if str(getattr(u, 'role', '') or '').strip().lower() == 'repartidor']

    orders = list_orders(skip=0, limit=2000, source=None, q=None, date=None, db=db)
    results = []
    for d in drivers:
        try:
            result = _auto_assign_routes_for_driver(d, orders, include_assigned=bool(include_assigned), db=db)
            results.append(result)
            try:
                assigned_ids = result.get('assigned_ids') if isinstance(result, dict) else None
                if assigned_ids:
                    for od in orders:
                        try:
                            if od.get('id') in assigned_ids:
                                od['assigned_driver_id'] = getattr(d, 'id', None)
                                od['assigned_driver_username'] = getattr(d, 'username', None)
                                od['assigned_driver_name'] = getattr(d, 'full_name', None) or getattr(d, 'username', None)
                                od['assigned_driver_zone'] = getattr(d, 'zone', None)
                        except Exception:
                            continue
            except Exception:
                pass
        except Exception:
            logger.exception('auto-assign driver failed')
            results.append({'driver_id': getattr(d, 'id', None), 'assigned': 0, 'route_id': None})
    headers = _cors_headers_for_request(request)
    return JSONResponse(status_code=200, content={'results': results}, headers=headers)


@app.get('/admin/deliveries', response_model=List[schemas.OrderResponse])
def admin_deliveries(
    skip: int = 0,
    limit: int = 200,
    driver_id: Optional[str] = None,
    driver_username: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_admin=Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    role = str(getattr(current_admin, 'role', '') or '').strip().lower()
    if role not in ('owner', 'admin'):
        raise HTTPException(status_code=403, detail='No autorizado')
    orders = list_orders(skip=skip, limit=limit, source=None, q=None, date=None, db=db)
    out = []
    df = str(date_from or '').strip()
    dt = str(date_to or '').strip()
    for od in (orders or []):
        status_val = _normalize_order_status(od.get('status'))
        delivered_at = od.get('delivered_at')
        if status_val != 'entregado' and not delivered_at:
            continue
        if driver_id or driver_username:
            did = str(od.get('delivered_by_id') or od.get('assigned_driver_id') or '').strip()
            duser = str(od.get('delivered_by_username') or od.get('assigned_driver_username') or '').strip()
            if driver_id and did != str(driver_id):
                continue
            if driver_username and duser != str(driver_username):
                continue
        if df or dt:
            try:
                dt_val = delivered_at or od.get('created_at')
                if isinstance(dt_val, str):
                    dt_val = _coerce_datetime(dt_val)
                if isinstance(dt_val, datetime.datetime):
                    date_key = dt_val.date().isoformat()
                else:
                    date_key = ''
            except Exception:
                date_key = ''
            if df and date_key and date_key < df:
                continue
            if dt and date_key and date_key > dt:
                continue
        out.append(od)
    out.sort(key=lambda o: (o.get('delivered_at') is None, o.get('delivered_at') or o.get('created_at') or datetime.datetime.min), reverse=True)
    return out


@app.post('/admin/driver-location', response_model=schemas.DriverLocationOut)
def admin_driver_location(
    payload: schemas.DriverLocationIn,
    current_admin=Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    role = str(getattr(current_admin, 'role', '') or '').strip().lower()
    if role not in ('owner', 'admin', 'repartidor'):
        raise HTTPException(status_code=403, detail='No autorizado')
    lat = _coerce_coord(getattr(payload, 'lat', None))
    lon = _coerce_coord(getattr(payload, 'lon', None))
    if lat is None or lon is None:
        raise HTTPException(status_code=400, detail='Coordenadas inválidas')

    accuracy = getattr(payload, 'accuracy', None)
    speed = getattr(payload, 'speed', None)
    heading = getattr(payload, 'heading', None)
    battery = getattr(payload, 'battery', None)
    recorded_at = _parse_epoch_to_datetime(getattr(payload, 'timestamp', None)) or datetime.datetime.now(datetime.timezone.utc)

    driver_id = getattr(current_admin, 'id', None)
    driver_username = getattr(current_admin, 'username', None)
    driver_name = getattr(current_admin, 'full_name', None) or driver_username

    payload_out = {
        'driver_id': driver_id,
        'driver_username': str(driver_username or '').strip() or None,
        'driver_name': str(driver_name or '').strip() or None,
        'lat': float(lat),
        'lon': float(lon),
        'accuracy': float(accuracy) if accuracy is not None else None,
        'speed': float(speed) if speed is not None else None,
        'heading': float(heading) if heading is not None else None,
        'battery': float(battery) if battery is not None else None,
        'recorded_at': recorded_at,
    }

    try:
        key = str(driver_id if driver_id is not None else driver_username or 'unknown')
        with DRIVER_LOCATIONS_LOCK:
            DRIVER_LOCATIONS[key] = dict(payload_out, updated_at=time.time())
    except Exception:
        pass

    if DRIVER_LOCATION_STORE_HISTORY:
        try:
            row = models.AdminDriverLocation(
                admin_user_id=int(driver_id) if driver_id is not None else None,
                username=str(driver_username or '').strip() or None,
                full_name=str(driver_name or '').strip() or None,
                lat=float(lat),
                lon=float(lon),
                accuracy=float(accuracy) if accuracy is not None else None,
                speed=float(speed) if speed is not None else None,
                heading=float(heading) if heading is not None else None,
                battery=float(battery) if battery is not None else None,
                recorded_at=recorded_at,
            )
            db.add(row)
            db.commit()
        except Exception:
            try:
                db.rollback()
            except Exception:
                pass

    try:
        asyncio.create_task(push_event({"action": "driver_location", "driver": payload_out}))
    except Exception:
        pass

    return payload_out


@app.get('/admin/driver-locations', response_model=List[schemas.DriverLocationOut])
def admin_driver_locations(current_admin=Depends(get_current_admin_user)):
    role = str(getattr(current_admin, 'role', '') or '').strip().lower()
    if role not in ('owner', 'admin'):
        raise HTTPException(status_code=403, detail='No autorizado')
    now = time.time()
    out = []
    try:
        with DRIVER_LOCATIONS_LOCK:
            for entry in DRIVER_LOCATIONS.values():
                try:
                    age = None
                    ts = entry.get('updated_at')
                    if ts:
                        age = max(0.0, now - float(ts))
                    if age is not None and DRIVER_LOCATION_MAX_AGE_SEC > 0 and age > DRIVER_LOCATION_MAX_AGE_SEC:
                        continue
                    payload = dict(entry)
                    payload.pop('updated_at', None)
                    payload['age_sec'] = age
                    out.append(payload)
                except Exception:
                    continue
    except Exception:
        pass
    return out


@app.patch('/admin/orders/{order_id}/assign')
async def admin_assign_order(order_id: str, request: Request, current_admin=Depends(get_current_admin_user)):
    role = str(getattr(current_admin, 'role', '') or '').strip().lower()
    if role not in ('owner', 'admin'):
        raise HTTPException(status_code=403, detail='No autorizado')
    try:
        body = await request.json()
    except Exception:
        body = {}
    driver_id = body.get('driver_id')
    driver_username = body.get('driver_username')
    if driver_id is None and not driver_username:
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=400, content={'error': 'missing_driver'}, headers=headers)

    # Resolve driver
    db = SessionLocal()
    try:
        driver = None
        if driver_id is not None:
            try:
                driver = crud.get_admin_user_by_id(db, int(driver_id))
            except Exception:
                driver = None
        if not driver and driver_username:
            driver = crud.get_admin_user_by_username(db, str(driver_username))
        if not driver:
            raise HTTPException(status_code=404, detail='Repartidor no encontrado')
        driver_role = str(getattr(driver, 'role', '') or '').strip().lower()
        if driver_role != 'repartidor':
            raise HTTPException(status_code=400, detail='Usuario no es repartidor')
        driver_zone = str(getattr(driver, 'zone', '') or '').strip()
        if not driver_zone:
            raise HTTPException(status_code=400, detail='Repartidor sin zona')
    finally:
        try:
            db.close()
        except Exception:
            pass

    # Ensure columns exist
    try:
        _run_add_user_columns()
    except Exception:
        pass

    # Fetch order for validation (zone)
    try:
        insp = inspect(engine)
        existing = {c['name'] for c in insp.get_columns('orders')}
    except Exception:
        existing = set()
    cols = ['id']
    for c in ('status', 'user_barrio', 'user_calle', 'user_numeracion', 'user_postal_code', 'user_department', '_token_preview', 'items', 'created_at'):
        if c in existing:
            cols.append(c)
    cols_sql = ', '.join(cols)
    row = None
    try:
        if str(order_id).isdigit():
            row = _safe_engine_fetchone(f"SELECT {cols_sql} FROM orders WHERE id = :id LIMIT 1", {'id': int(order_id)})
        else:
            row = _safe_engine_fetchone(f"SELECT {cols_sql} FROM orders WHERE CAST(id AS TEXT) = :id LIMIT 1", {'id': str(order_id)})
    except Exception:
        row = None
    if not row:
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=404, content={'error': 'not_found'}, headers=headers)

    order_data = {k: row[idx] for idx, k in enumerate(cols)}
    try:
        if isinstance(order_data.get('_token_preview'), str):
            order_data['_token_preview'] = json.loads(order_data['_token_preview'])
    except Exception:
        pass

    zone_tokens = _parse_admin_zones(driver_zone)
    try:
        snapshot = _extract_order_address_snapshot(order_data if isinstance(order_data, dict) else {})
        has_hint = bool(
            str(snapshot.get('department') or '').strip() or
            str(snapshot.get('barrio') or '').strip() or
            str(snapshot.get('raw_address') or '').strip() or
            str(snapshot.get('postal_code') or '').strip()
        )
        if zone_tokens and has_hint and not _order_matches_zone(order_data, zone_tokens):
            headers = _cors_headers_for_request(request)
            return JSONResponse(status_code=409, content={'error': 'zona_no_coincide'}, headers=headers)
    except Exception:
        pass

    # Persist assignment
    try:
        assigned_at = datetime.datetime.now(datetime.timezone.utc)
        params = {
            'assigned_driver_id': int(getattr(driver, 'id')),
            'assigned_driver_username': str(getattr(driver, 'username') or '').strip(),
            'assigned_driver_name': str(getattr(driver, 'full_name') or getattr(driver, 'username') or '').strip() or None,
            'assigned_driver_zone': driver_zone,
            'assigned_at': assigned_at,
            'id': int(order_id) if str(order_id).isdigit() else str(order_id),
        }
        with engine.begin() as conn:
            if str(order_id).isdigit():
                conn.execute(
                    text(
                        "UPDATE orders SET assigned_driver_id = :assigned_driver_id, "
                        "assigned_driver_username = :assigned_driver_username, "
                        "assigned_driver_name = :assigned_driver_name, "
                        "assigned_driver_zone = :assigned_driver_zone, "
                        "assigned_at = :assigned_at "
                        "WHERE id = :id"
                    ),
                    params,
                )
            else:
                conn.execute(
                    text(
                        "UPDATE orders SET assigned_driver_id = :assigned_driver_id, "
                        "assigned_driver_username = :assigned_driver_username, "
                        "assigned_driver_name = :assigned_driver_name, "
                        "assigned_driver_zone = :assigned_driver_zone, "
                        "assigned_at = :assigned_at "
                        "WHERE CAST(id AS TEXT) = :id"
                    ),
                    params,
                )
    except Exception as e:
        logger.exception('admin_assign_order update failed: %s', e)
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=500, content={'error': 'update_failed'}, headers=headers)

    # Recompute optimized route order for this driver
    try:
        db_local = SessionLocal()
        try:
            all_orders = list_orders(skip=0, limit=2000, source=None, q=None, date=None, db=db_local)
            _auto_assign_routes_for_driver(driver, all_orders, include_assigned=True, db=db_local)
        finally:
            db_local.close()
    except Exception:
        pass

    # Return updated order (best effort)
    try:
        insp = inspect(engine)
        existing = {c['name'] for c in insp.get_columns('orders')}
    except Exception:
        existing = set()
    cols = ['id']
    for c in ('items', 'total', 'created_at', 'status'):
        if c in existing:
            cols.append(c)
    optional = [
        'customer_type',
        'user_id','user_full_name','user_email','user_barrio','user_calle','user_numeracion','user_postal_code','user_department',
        '_token_received','_token_preview','source',
        'payment_method','payment_status','payment_reference',
        'scheduled_delivery_date', 'delivery_cutoff_applied', 'delivery_timezone', 'delivery_cutoff_hour',
        'assigned_driver_id', 'assigned_driver_username', 'assigned_driver_name', 'assigned_driver_zone', 'assigned_at',
        'delivery_lat', 'delivery_lon',
        'route_id', 'route_order', 'route_generated_at',
        'delivered_at', 'delivered_by_id', 'delivered_by_username',
    ]
    for c in optional:
        if c in existing:
            cols.append(c)
    cols_sql = ', '.join(cols)
    updated_row = None
    try:
        if str(order_id).isdigit():
            updated_row = _safe_engine_fetchone(f"SELECT {cols_sql} FROM orders WHERE id = :id LIMIT 1", {'id': int(order_id)})
        else:
            updated_row = _safe_engine_fetchone(f"SELECT {cols_sql} FROM orders WHERE CAST(id AS TEXT) = :id LIMIT 1", {'id': str(order_id)})
    except Exception:
        updated_row = None
    if not updated_row:
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=200, content={'ok': True}, headers=headers)
    od = {k: updated_row[idx] for idx, k in enumerate(cols)}
    try:
        if isinstance(od.get('items'), str):
            od['items'] = json.loads(od['items'])
    except Exception:
        od['items'] = []
    try:
        if isinstance(od.get('_token_preview'), str):
            od['_token_preview'] = json.loads(od['_token_preview'])
    except Exception:
        pass
    try:
        od = _attach_maps_url(od)
    except Exception:
        pass
    try:
        await push_event({"action": "order_updated", "order": od})
    except Exception:
        pass
    headers = _cors_headers_for_request(request)
    return JSONResponse(status_code=200, content=jsonable_encoder(od), headers=headers)


# Small HTML injector for admin/repartidor pages.
def _inject_admin_config(content: str) -> str:
    try:
        maps_key = str(GOOGLE_MAPS_JS_API_KEY or '').strip()
    except Exception:
        maps_key = ''
    return (content or '').replace('__GOOGLE_MAPS_API_KEY__', maps_key)


# -------------------------------------------------------------------
# STATIC FILES (mounted last so API routes take precedence)
# -------------------------------------------------------------------
if os.path.exists(FRONTEND_DIR):
    @app.get("/")
    def root():
        return RedirectResponse("/admin/index.html")

    # Serve a versioned `index.html` to help cache-bust client assets after deploys.
    @app.get("/admin/index.html")
    async def admin_index():
        path = os.path.join(FRONTEND_DIR, 'index.html')
        if not os.path.exists(path):
            raise HTTPException(status_code=404, detail='Admin UI not found')
        try:
            with open(path, 'r', encoding='utf-8') as f:
                content = f.read()
        except Exception as e:
            logger.exception('Failed reading admin index.html: %s', e)
            raise HTTPException(status_code=500, detail='Failed to read admin UI')
        # Determine a short version token: prefer DEPLOY_COMMIT if present, otherwise use file mtime
        ver = os.environ.get('DEPLOY_COMMIT') or os.environ.get('RENDER_GIT_COMMIT') or os.environ.get('HEROKU_SLUG_COMMIT') or str(int(os.path.getmtime(path)))
        # Inject config + cache-busting query param for the main JS and CSS references
        content = _inject_admin_config(content)
        content = content.replace('app.js"', f'app.js?v={ver}"').replace('styles.css"', f'styles.css?v={ver}"')
        return Response(content=content, media_type='text/html')

    @app.get("/admin/repartidor.html")
    async def admin_repartidor():
        path = os.path.join(FRONTEND_DIR, 'repartidor.html')
        if not os.path.exists(path):
            raise HTTPException(status_code=404, detail='Repartidor UI not found')
        try:
            with open(path, 'r', encoding='utf-8') as f:
                content = f.read()
        except Exception as e:
            logger.exception('Failed reading admin repartidor.html: %s', e)
            raise HTTPException(status_code=500, detail='Failed to read repartidor UI')
        ver = os.environ.get('DEPLOY_COMMIT') or os.environ.get('RENDER_GIT_COMMIT') or os.environ.get('HEROKU_SLUG_COMMIT') or str(int(os.path.getmtime(path)))
        content = _inject_admin_config(content)
        content = content.replace('repartidor.js"', f'repartidor.js?v={ver}"').replace('styles.css"', f'styles.css?v={ver}"')
        return Response(content=content, media_type='text/html')

    app.mount("/admin", StaticFiles(directory=FRONTEND_DIR), name="admin")

app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

if os.path.exists(CATALOG_DIR):
    app.mount("/catalogo", StaticFiles(directory=CATALOG_DIR), name="catalogo")


@app.patch('/orders/{order_id}/status')
async def update_order_status(order_id: str, request: Request):
    """Update the `status` field for an order and broadcast the change via WS.
    Accepts numeric or non-numeric IDs (debug events may use string IDs)."""
    def _ensure_status_column():
        try:
            insp_local = inspect(engine)
            existing_local = {c['name'] for c in insp_local.get_columns('orders')}
        except Exception:
            existing_local = set()
        if 'status' in existing_local:
            return True, existing_local
        try:
            dialect_local = getattr(engine, 'dialect', None)
            dialect_name_local = getattr(dialect_local, 'name', '') if dialect_local else ''
            with engine.begin() as conn_local:
                if 'postgres' in dialect_name_local:
                    try:
                        conn_local.execute(text("ALTER TABLE orders ADD COLUMN IF NOT EXISTS status VARCHAR(50) DEFAULT 'recibido'"))
                    except Exception:
                        conn_local.execute(text("ALTER TABLE orders ADD COLUMN status VARCHAR(50) DEFAULT 'recibido'"))
                else:
                    conn_local.execute(text("ALTER TABLE orders ADD COLUMN status VARCHAR(50) DEFAULT 'recibido'"))
            try:
                insp_local = inspect(engine)
                existing_local = {c['name'] for c in insp_local.get_columns('orders')}
            except Exception:
                existing_local.add('status')
            return True, existing_local
        except Exception as e_local:
            logger.exception('ensure status column failed: %s', e_local)
            return False, existing_local

    def _persist_status_override(order_id_value, status_value):
        try:
            db_local = SessionLocal()
            try:
                overrides = crud.get_setting(db_local, 'order_status_overrides') or {}
                overrides[str(order_id_value)] = status_value
                crud.set_setting(db_local, 'order_status_overrides', overrides)
            finally:
                try:
                    db_local.close()
                except Exception:
                    pass
            return True
        except Exception:
            # fall back to in-memory cache
            try:
                ORDER_STATUS_CACHE[str(order_id_value)] = { 'status': status_value, 'ts': time.time() }
                return True
            except Exception:
                return False

    def _get_status_override(order_id_value):
        try:
            db_local = SessionLocal()
            try:
                overrides = crud.get_setting(db_local, 'order_status_overrides') or {}
                return overrides.get(str(order_id_value))
            finally:
                try:
                    db_local.close()
                except Exception:
                    pass
        except Exception:
            try:
                return (ORDER_STATUS_CACHE.get(str(order_id_value)) or {}).get('status')
            except Exception:
                return None

    def _clear_status_override(order_id_value):
        key = str(order_id_value)
        try:
            db_local = SessionLocal()
            try:
                overrides = crud.get_setting(db_local, 'order_status_overrides') or {}
                if key in overrides:
                    del overrides[key]
                    crud.set_setting(db_local, 'order_status_overrides', overrides)
            finally:
                try:
                    db_local.close()
                except Exception:
                    pass
        except Exception:
            pass
        try:
            if key in ORDER_STATUS_CACHE:
                del ORDER_STATUS_CACHE[key]
        except Exception:
            pass

    try:
        body = await request.json()
    except Exception:
        body = {}
    status_raw = body.get('status')
    if status_raw is None or not str(status_raw).strip():
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=400, content={'error': 'missing_status'}, headers=headers)
    status_norm = _normalize_order_status_strict(status_raw)
    if not status_norm:
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=400, content={'error': 'invalid_status'}, headers=headers)
    delivered_actor_id, delivered_actor_username = _extract_admin_actor_from_request(request)

    # Decide whether we should treat order_id as int or text for SQL WHERE
    use_id_int = False
    id_param = order_id
    try:
        id_int = int(order_id)
        use_id_int = True
        id_param = id_int
    except Exception:
        use_id_int = False
        id_param = str(order_id)

    # Best-effort: ensure status column exists before updating
    ok_status_col, existing_cols = _ensure_status_column()

    # Guard against backwards transitions (e.g. already enviado -> visto).
    try:
        now_utc = datetime.datetime.now(datetime.timezone.utc)
        current_override_raw = _get_status_override(order_id)
        current_status_raw = None
        current_created_at = None
        current_scheduled_delivery_date = None
        current_delivery_tz = None
        try:
            # created_at is always expected; include it so we can apply auto milestones.
            pre_cols = ['status', 'created_at']
            if 'scheduled_delivery_date' in (existing_cols or set()):
                pre_cols.append('scheduled_delivery_date')
            if 'delivery_timezone' in (existing_cols or set()):
                pre_cols.append('delivery_timezone')
            pre_cols_sql = ', '.join(pre_cols)
            if use_id_int:
                row_pre = _safe_engine_fetchone(f"SELECT {pre_cols_sql} FROM orders WHERE id = :id LIMIT 1", {'id': id_param})
            else:
                row_pre = _safe_engine_fetchone(f"SELECT {pre_cols_sql} FROM orders WHERE CAST(id AS TEXT) = :id LIMIT 1", {'id': id_param})
            if row_pre:
                try:
                    idx_status = pre_cols.index('status') if 'status' in pre_cols else -1
                except Exception:
                    idx_status = -1
                try:
                    idx_created = pre_cols.index('created_at') if 'created_at' in pre_cols else -1
                except Exception:
                    idx_created = -1
                try:
                    idx_tz = pre_cols.index('delivery_timezone') if 'delivery_timezone' in pre_cols else -1
                except Exception:
                    idx_tz = -1
                try:
                    idx_sched = pre_cols.index('scheduled_delivery_date') if 'scheduled_delivery_date' in pre_cols else -1
                except Exception:
                    idx_sched = -1
                if idx_status >= 0:
                    current_status_raw = row_pre[idx_status]
                if idx_created >= 0:
                    current_created_at = row_pre[idx_created]
                if idx_sched >= 0:
                    current_scheduled_delivery_date = row_pre[idx_sched]
                if idx_tz >= 0:
                    current_delivery_tz = row_pre[idx_tz]
        except Exception:
            pass

        base_current = _merge_order_statuses(current_status_raw, current_override_raw)
        auto_status = _compute_order_auto_status(
            current_created_at,
            scheduled_delivery_date_value=current_scheduled_delivery_date,
            status_value=base_current,
            now_utc=now_utc,
            tz_name_hint=str(current_delivery_tz or '').strip() or None,
        )
        effective_current = _merge_order_statuses(base_current, auto_status)
        if effective_current == 'cancelado' and status_norm != 'cancelado':
            headers = _cors_headers_for_request(request)
            return JSONResponse(
                status_code=409,
                content={'error': 'order_cancelled', 'current': effective_current, 'requested': status_norm},
                headers=headers,
            )
        if status_norm != 'cancelado' and _order_status_rank(status_norm) < _order_status_rank(effective_current):
            headers = _cors_headers_for_request(request)
            return JSONResponse(
                status_code=409,
                content={'error': 'backwards_transition', 'current': effective_current, 'requested': status_norm},
                headers=headers,
            )
    except Exception:
        pass

    db_update_ok = False
    try:
        # Use a transactional context so the UPDATE is committed immediately
        try:
            with engine.begin() as conn:
                if use_id_int:
                    conn.execute(text('UPDATE orders SET status = :status WHERE id = :id'), {'status': status_norm, 'id': id_param})
                else:
                    # Fall back to text comparison; CAST to TEXT works on SQLite/Postgres
                    conn.execute(text('UPDATE orders SET status = :status WHERE CAST(id AS TEXT) = :id'), {'status': status_norm, 'id': id_param})
            db_update_ok = True
        except Exception as inner_e:
            # If status column was missing, attempt to add and retry once.
            msg = str(inner_e).lower()
            if (('status' in msg) and ('does not exist' in msg or 'no such column' in msg)):
                ok_status_col, existing_cols = _ensure_status_column()
                if ok_status_col:
                    try:
                        with engine.begin() as conn2:
                            if use_id_int:
                                conn2.execute(text('UPDATE orders SET status = :status WHERE id = :id'), {'status': status_norm, 'id': id_param})
                            else:
                                conn2.execute(text('UPDATE orders SET status = :status WHERE CAST(id AS TEXT) = :id'), {'status': status_norm, 'id': id_param})
                        db_update_ok = True
                    except Exception as inner_e2:
                        logger.exception('update_order_status retry failed: %s', inner_e2)
                        headers = _cors_headers_for_request(request)
                        # fall through to override persistence below
                        db_update_ok = False
                else:
                    logger.exception('update_order_status DB execute failed: %s', inner_e)
                    headers = _cors_headers_for_request(request)
                    db_update_ok = False
            else:
                logger.exception('update_order_status DB execute failed: %s', inner_e)
                headers = _cors_headers_for_request(request)
                db_update_ok = False
    except Exception as e:
        logger.exception('update_order_status failed: %s', e)
        headers = _cors_headers_for_request(request)
        db_update_ok = False

    # Persist an override only when the DB update failed. When DB update succeeded,
    # clear any previous override so list endpoints don't get stuck on stale values.
    override_ok = False
    try:
        if db_update_ok:
            _clear_status_override(order_id)
        else:
            override_ok = _persist_status_override(order_id, status_norm)
    except Exception:
        override_ok = False

    # Fetch updated row safely (only request existing columns)
    try:
        insp = inspect(engine)
        existing = {c['name'] for c in insp.get_columns('orders')}
    except Exception:
        existing = set()

    cols = ['id']
    for c in ('items', 'total', 'created_at', 'status'):
        if c in existing:
            cols.append(c)
    optional = [
        'customer_type',
        'user_id','user_full_name','user_email','user_barrio','user_calle','user_numeracion','user_postal_code','user_department',
        '_token_received','_token_preview','source',
        'payment_method','payment_status','payment_reference',
        'scheduled_delivery_date', 'delivery_cutoff_applied', 'delivery_timezone', 'delivery_cutoff_hour',
        'assigned_driver_id', 'assigned_driver_username', 'assigned_driver_name', 'assigned_driver_zone', 'assigned_at',
    ]
    for c in optional:
        if c in existing:
            cols.append(c)

    cols_sql = ', '.join(cols)
    try:
        if use_id_int:
            row = _safe_engine_fetchone(f"SELECT {cols_sql} FROM orders WHERE id = :id LIMIT 1", {'id': id_param})
        else:
            row = _safe_engine_fetchone(f"SELECT {cols_sql} FROM orders WHERE CAST(id AS TEXT) = :id LIMIT 1", {'id': id_param})
        if row is None:
            # treat as not found or as a DB failure depending on caller
            pass
    except Exception as e:
        logger.exception('fetch updated order failed: %s', e)
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=500, content={'error': str(e)}, headers=headers)

    if not row:
        # If update failed but we could persist override, return a minimal response.
        if override_ok:
            headers = _cors_headers_for_request(request)
            return JSONResponse(status_code=200, content=jsonable_encoder({'id': id_param, 'status': status_norm, 'status_fallback': True}), headers=headers)
        headers = _cors_headers_for_request(request)
        return JSONResponse(status_code=404, content={'error': 'not_found'}, headers=headers)

    od = {k: row[idx] for idx, k in enumerate(cols)}
    if 'items' not in od:
        od['items'] = []
    if 'total' not in od:
        od['total'] = 0
    if 'status' not in od:
        od['status'] = status_norm
    # parse items and token_preview if present
    try:
        if isinstance(od.get('items'), str):
            import json as _json
            od['items'] = _json.loads(od['items'])
    except Exception:
        od['items'] = []
    try:
        if isinstance(od.get('_token_preview'), str):
            import json as _json
            try:
                od['_token_preview'] = _json.loads(od['_token_preview'])
            except Exception:
                pass
    except Exception:
        pass
    try:
        od = _attach_maps_url(od)
    except Exception:
        pass

    # Auto-assign prepared orders to a repartidor based on zone (if not assigned yet)
    try:
        if status_norm == 'preparado' and not od.get('assigned_driver_id') and not od.get('assigned_driver_username'):
            db_local = SessionLocal()
            try:
                driver = _select_driver_for_order(od, db_local)
                if driver:
                    all_orders = list_orders(skip=0, limit=2000, source=None, q=None, date=None, db=db_local)
                    _auto_assign_routes_for_driver(driver, all_orders, include_assigned=True, db=db_local)
                    # re-fetch updated order to include assignment + route fields
                    try:
                        insp = inspect(engine)
                        existing = {c['name'] for c in insp.get_columns('orders')}
                    except Exception:
                        existing = set()
                    cols = ['id']
                    for c in ('items', 'total', 'created_at', 'status'):
                        if c in existing:
                            cols.append(c)
                    optional = [
                        'customer_type',
                        'user_id','user_full_name','user_email','user_barrio','user_calle','user_numeracion','user_postal_code','user_department',
                        '_token_received','_token_preview','source',
                        'payment_method','payment_status','payment_reference',
                        'scheduled_delivery_date', 'delivery_cutoff_applied', 'delivery_timezone', 'delivery_cutoff_hour',
                        'assigned_driver_id', 'assigned_driver_username', 'assigned_driver_name', 'assigned_driver_zone', 'assigned_at',
                        'delivery_lat', 'delivery_lon',
                        'route_id', 'route_order', 'route_generated_at',
                        'delivered_at', 'delivered_by_id', 'delivered_by_username',
                    ]
                    for c in optional:
                        if c in existing:
                            cols.append(c)
                    cols_sql = ', '.join(cols)
                    if use_id_int:
                        row2 = _safe_engine_fetchone(f"SELECT {cols_sql} FROM orders WHERE id = :id LIMIT 1", {'id': id_param})
                    else:
                        row2 = _safe_engine_fetchone(f"SELECT {cols_sql} FROM orders WHERE CAST(id AS TEXT) = :id LIMIT 1", {'id': id_param})
                    if row2:
                        od = {k: row2[idx] for idx, k in enumerate(cols)}
                        try:
                            if isinstance(od.get('items'), str):
                                od['items'] = json.loads(od['items'])
                        except Exception:
                            od['items'] = []
                        try:
                            if isinstance(od.get('_token_preview'), str):
                                od['_token_preview'] = json.loads(od['_token_preview'])
                        except Exception:
                            pass
                        try:
                            od = _attach_maps_url(od)
                        except Exception:
                            pass
            finally:
                db_local.close()
    except Exception:
        pass

    # Broadcast the updated order to connected admin clients
    try:
        await push_event({"action": "order_updated", "order": od})
    except Exception:
        pass

    # If delivered, persist delivery metadata
    try:
        if status_norm == 'entregado':
            delivered_at = od.get('delivered_at')
            if not delivered_at:
                assign_id = od.get('assigned_driver_id')
                assign_user = od.get('assigned_driver_username')
                use_id = delivered_actor_id if delivered_actor_id is not None else (assign_id if assign_id is not None else None)
                use_user = delivered_actor_username if delivered_actor_username else (assign_user if assign_user else None)
                now_ts = datetime.datetime.now(datetime.timezone.utc)
                try:
                    with engine.begin() as conn:
                        if use_id is not None or use_user:
                            if use_id_int:
                                conn.execute(
                                    text("UPDATE orders SET delivered_at = :delivered_at, delivered_by_id = :delivered_by_id, delivered_by_username = :delivered_by_username WHERE id = :id"),
                                    {'delivered_at': now_ts, 'delivered_by_id': use_id, 'delivered_by_username': use_user, 'id': id_param},
                                )
                            else:
                                conn.execute(
                                    text("UPDATE orders SET delivered_at = :delivered_at, delivered_by_id = :delivered_by_id, delivered_by_username = :delivered_by_username WHERE CAST(id AS TEXT) = :id"),
                                    {'delivered_at': now_ts, 'delivered_by_id': use_id, 'delivered_by_username': use_user, 'id': id_param},
                                )
                        else:
                            if use_id_int:
                                conn.execute(
                                    text("UPDATE orders SET delivered_at = :delivered_at WHERE id = :id"),
                                    {'delivered_at': now_ts, 'id': id_param},
                                )
                            else:
                                conn.execute(
                                    text("UPDATE orders SET delivered_at = :delivered_at WHERE CAST(id AS TEXT) = :id"),
                                    {'delivered_at': now_ts, 'id': id_param},
                                )
                    od['delivered_at'] = now_ts
                    if use_id is not None:
                        od['delivered_by_id'] = use_id
                    if use_user:
                        od['delivered_by_username'] = use_user
                except Exception:
                    pass
    except Exception:
        pass

    # Ensure status is present in response
    try:
        effective_status = _normalize_order_status(od.get('status') or status_norm)
        od['status'] = effective_status
        if override_ok and not db_update_ok:
            od['status_fallback'] = True
    except Exception:
        pass

    # Notify customer when order status changes to user-facing milestones.
    try:
        status_effective = _normalize_order_status(od.get('status') or status_norm)
        if status_effective == 'visto':
            payload_for_seen = _enrich_order_contact_fields(
                dict(od) if isinstance(od, dict) else {'id': id_param, 'status': status_effective}
            )
            send_ok = await _send_order_seen_notification_email(payload_for_seen)
            try:
                od['_seen_email_sent'] = bool(send_ok)
            except Exception:
                pass
        elif status_effective == 'preparado':
            payload_for_prepared = _enrich_order_contact_fields(
                dict(od) if isinstance(od, dict) else {'id': id_param, 'status': status_effective}
            )
            send_ok = await _send_order_prepared_notification_email(payload_for_prepared)
            try:
                od['_prepared_email_sent'] = bool(send_ok)
            except Exception:
                pass
    except Exception:
        logger.exception('Could not schedule status notification email for order id=%s', id_param)

    headers = _cors_headers_for_request(request)
    return JSONResponse(status_code=200, content=jsonable_encoder(od), headers=headers)
